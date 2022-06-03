
from array import array
from openpyxl import load_workbook
from openpyxl import Workbook
import apsw
import time
import os
import openpyxl
from PyQt5.QtCore import QObject, QThread, pyqtSignal
#QProgressBar
from PyQt5.QtWidgets import QProgressBar


from loguru import logger
#import QVBoxLayout
from PyQt5.QtWidgets import  QPushButton,   QFormLayout, QLineEdit,  QLabel, QPushButton, QDialog, QMessageBox, QGridLayout

class Excel_Popup(QDialog):
    def __init__(self, logger, ctypes, parent_, mode, excelpath, enginepath, dbpath, sqldir, sourcetable, userange, preserve):
        
        super().__init__(parent_)
        modestr= "import" if mode == 0 else "export"
        self.resize(600, 300)
        self.setWindowTitle("Import/Export DB")
        self.layout = QFormLayout()
        self.setLayout(self.layout)
        #SELECT formname, tab, tabsequence, grid, tabdesc, treepath, tabgroup, tabsize, taburl FROM tabs; where formname = parent_title and tabname = tab_name
        #add button
        self.start_button = QPushButton("start")
        self.infobox = QLabel(modestr)
        self.start_button.clicked.connect(self.runTask)
        self.layout.addRow(self.start_button)
        self.layout.addRow(self.infobox)
        self.mode = mode
        self.excelpath = excelpath
        self.enginepath = enginepath
        self.dbpath = dbpath
        self.sqldir = sqldir
        self.sourcetable = sourcetable
        self.userange = userange
        self.preserve = preserve
        self.logger = logger
        self.ctypes = ctypes
        self.progress = 0
        self.finished = False
        self.i1 = 0 
        self.i2 = 0
        self.progressbar = []
       
    def runTask(self):
        #destroy the button
        self.start_button.deleteLater()
        self.thread = QThread()
        self.worker = Worker(self.mode, self.excelpath, self.enginepath, self.dbpath, self.sqldir, self.sourcetable, self.userange, self.preserve, self.logger, self.ctypes)
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        


        self.thread.start()


class Worker(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(list)

    def __init__(self, mode,excelpath,enginepath,dbpath,sqldir,sourcetable,userange,preserve,logger,ctypes):
        super().__init__()
        self.mode = mode
        self.modestr = "import" if mode == 0 else "export"
        self.excelpath = excelpath
        self.enginepath = enginepath
        self.dbpath = dbpath
        self.sqldir = sqldir
        self.sourcetable = sourcetable
        self.userange = userange
        self.preserve = preserve
        self.logger = logger
        self.ctypes = ctypes
        
        self.i1 = 0 
        self.i2 = 0

    def add1(self):
        self.i1 += 1
        return(self.i1)
    def add2(self):
        self.i2 += 1
        return(self.i2)

    def run(self):
        if self.mode==1:
            print("export")
            if os.path.exists(self.excelpath)==True:
                book = Workbook()
                
            filepath = self.excelpath
            if os.path.exists(filepath)==False:
                self.logger.error("excel file not found")
                self.ctypes.windll.user32.MessageBoxW(0,"excel file not found","Error",1)
                return
            if self.preserve==1:
                wb = openpyxl.load_workbook(filepath)
            else:
                wb = openpyxl.Workbook(filepath)
            
            attachedconnection = apsw.Connection(self.dbpath)
            attachedcursor = attachedconnection.cursor()
            append = ""
            if self.sourcetable!="" and self.sourcetable!="None" and self.sourcetable!=None:
                append = " and tbl_name = '"+self.sourcetable+"'"
            table_names = self.aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master where tbl_name <> 'pyexcelmenu' and tbl_name <> 'sqlengine'"+append)
            table_con = []
            for tn in range(len(table_names)):
                table_names[tn] = str(table_names[tn][0])
            
            for tn in range(len(table_names)):
                if table_names[tn] not in wb.sheetnames:
                    wb.create_sheet(table_names[tn])
                    db_con = self.aReadSQL(attachedcursor,"select * from "+table_names[tn])
                    pragma = self.aReadSQL(attachedcursor,"pragma table_info('"+table_names[tn]+"')")
                    field_names = []
                    for p in range(len(pragma)):
                        field_names.append(str(pragma[p][1]))
                    db_con.insert(0,field_names)
                    table_con.append(db_con)
                else:
                    ws = wb[str(table_names[tn])]
                    for x in range(wb[str(table_names[tn])].min_row-1,wb[str(table_names[tn])].max_row):
                        for y in range(wb[str(table_names[tn])].min_column-1,wb[str(table_names[tn])].max_column):
                            ws.cell(x+1,y+1).value = None
                    db_con = self.aReadSQL(attachedcursor,"select * from "+table_names[tn])
                    pragma = self.aReadSQL(attachedcursor,"pragma table_info('"+table_names[tn]+"')")
                    field_names = []
                    for p in range(len(pragma)):
                        field_names.append(str(pragma[p][1]))
                    db_con.insert(0,field_names)
                    table_con.append(db_con)
            attachedconnection.close(True)
            wb.save(filepath)
            print("saved 1")
            
            book = load_workbook(filename=filepath,data_only=True)
            for tn in range(len(book.sheetnames)):
                ws = book[book.sheetnames[tn]]
                if book.sheetnames[tn] in table_names:
                    index = table_names.index(book.sheetnames[tn])
                    for i in range(len(table_con[index])):
                        for j in range(len(table_con[index][i])):
                            ws.cell(row=i+1,column=j+1).value = table_con[index][i][j]
            book.save(filepath)
            print("saved 2")
        elif self.mode==0:
            print("import")
            sqldir = self.addSlash(self.sqldir)
            attachedconnection = apsw.Connection(self.dbpath)
            attachedcursor = attachedconnection.cursor()
            table_names = self.aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master where tbl_name <> 'pyexcelmenu' and tbl_name <> 'sqlengine' order by tbl_name asc",12,0.2,True)
            build_statement = []
            insert_statement = []
            orig_statement = []
            temp_table_names = []
            for etd0 in range(len(table_names)):
                sqlite_stat = table_names[etd0][0]
                if table_names[etd0][0][-4:]!="_old" and sqlite_stat.find("sqlite_stat")==-1:
                    temp_table_names.append(table_names[etd0])
            table_names = temp_table_names
            idir = os.path.dirname(self.excelpath)
            ibase = os.path.basename(self.excelpath)
            print(sqldir+"\\FK_rem_"+ibase[:-5]+".sql")
            if os.path.exists(sqldir+"\\FK_rem_"+ibase[:-5]+".sql")==False:
                f = open(sqldir+"\\FK_rem_"+ibase[:-5]+".sql","w+")
                for etd1 in range(len(table_names)):
                    temporig = self.aReadSQL(attachedcursor,"select sql from sqlite_master where tbl_name = '"+str(table_names[etd1][0])+"'",12,0.2,True)
                    if len(temporig)>0:
                        orig_statement.append(temporig[0][0])
                    pragma_info = self.aReadSQL(attachedcursor,"pragma table_info('"+str(table_names[etd1][0])+"')",12,0.2,True)
                    entry = []
                    pkeys = []
                    insert_statement2 = []
                    for pragrow in range(len(pragma_info)):
                        if pragma_info[pragrow][1].lower()=="table":
                            pragma_info[pragrow][1] = "["+pragma_info[pragrow][1]+"]"
                        insert_statement2.append(str(pragma_info[pragrow][1]))
                        entry.append(str(pragma_info[pragrow][1])+" "+str(pragma_info[pragrow][2]))
                        if int(pragma_info[pragrow][5])>0:
                            pkeys.append(str(pragma_info[pragrow][1]))
                    if len(entry)>0:
                        build_statement.append("create table if not exists "+str(table_names[etd1][0])+"("+','.join(entry))
                        if len(pkeys)>0:
                            build_statement[len(build_statement)-1] += ",primary key("+','.join(pkeys)+")); pragma foreign_keys = on"
                        else:
                            build_statement[len(build_statement)-1] += "); pragma foreign_keys = on"
                    insert_statement.append(insert_statement2)
                f.write("pragma foreign_keys = off;\n\n")
                #aWriteSQL(attachedcursor,"pragma foreign_keys = off")
                f.write("begin transaction;\n\n")
                #aWriteSQL(attachedcursor,"begin transaction")
                for etd2 in range(len(table_names)):
                    f.write("drop table if exists _"+str(table_names[etd2][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"drop table if exists _"+str(table_names[etd2][0])+"_old")
                    f.write("alter table "+str(table_names[etd2][0])+" rename to _"+str(table_names[etd2][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"alter table "+str(table_names[etd2][0])+" rename to _"+str(table_names[etd2][0])+"_old")
                    f.write(build_statement[etd2]+";\n\n")
                    #aWriteSQL(attachedcursor,build_statement[etd2])
                    f.write("insert into "+str(table_names[etd2][0])+"("+','.join(insert_statement[etd2])+")\n\n")
                    f.write("select "+','.join(insert_statement[etd2])+" from _"+str(table_names[etd2][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"insert into "+str(table_names[etd2][0])+"("+','.join(insert_statement[etd2])+") "+"select "+
                    #         ','.join(insert_statement[etd2])+" from _"+str(table_names[etd2][0])+"_old")
                    f.write("drop table _"+str(table_names[etd2][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"drop table _"+str(table_names[etd2][0])+"_old")
                f.write("commit;\n\n")
                #aWriteSQL(attachedcursor,"commit")
                f.write("pragma foreign_keys = on;")
                #aWriteSQL(attachedcursor,"pragma foreign_keys = on")
                f.close()
            else:
                for etd1 in range(len(table_names)):
                    temporig = self.aReadSQL(attachedcursor,"select sql from sqlite_master where tbl_name = '"+str(table_names[etd1][0])+"'",12,0.2,True)
                    if len(temporig)>0:
                        orig_statement.append(temporig[0][0])
                    pragma_info = self.aReadSQL(attachedcursor,"pragma table_info('"+str(table_names[etd1][0])+"')",12,0.2,True)
                    entry = []
                    pkeys = []
                    insert_statement2 = []
                    for pragrow in range(len(pragma_info)):
                        if pragma_info[pragrow][1].lower()=="table":
                            pragma_info[pragrow][1] = "["+pragma_info[pragrow][1]+"]"
                        insert_statement2.append(str(pragma_info[pragrow][1]))
                        entry.append(str(pragma_info[pragrow][1])+" "+str(pragma_info[pragrow][2]))
                        if int(pragma_info[pragrow][5])>0:
                            pkeys.append(str(pragma_info[pragrow][1]))
                    if len(entry)>0:
                        build_statement.append("create table if not exists "+str(table_names[etd1][0])+"("+','.join(entry))
                        if len(pkeys)>0:
                            build_statement[len(build_statement)-1] += ",primary key("+','.join(pkeys)+")); pragma foreign_keys = on"
                        else:
                            build_statement[len(build_statement)-1] += "); pragma foreign_keys = on"
                    insert_statement.append(insert_statement2)

            oldsqldir = sqldir
            sqldir = self.addSlash(sqldir)
            path = self.enginepath+" "+self.dbpath+" \".read '"+sqldir+"\\\\FK_rem_"+ibase[:-5]+".sql'\""
            var = os.system(path)
            if int(var)!=0 and str(var)!=None and str(var)!="":
                self.ctypes.windll.user32.MessageBoxW(0,str(var),"SQL Query Error",1)

            if os.path.exists(self.excelpath)==True:
                book = load_workbook(filename=self.excelpath,data_only=True)
                content_keys = book.sheetnames
            else:
                content_keys = []

            if self.userange==1:
                if self.sourcetable!="" and self.sourcetable!=None:
                    self.RangeToDB(attachedcursor,self.sourcetable,self.excelpath)
            else:
                temp_table_names = []
                for tn in range(len(table_names)):
                    temp_table_names.append(table_names[tn][0])
                for etd3 in range(len(content_keys)):
                    if self.sourcetable!="" and self.sourcetable!=None and self.sourcetable!="None":
                        if str(content_keys[etd3])==self.sourcetable:
                            self.ExcelToDB(attachedcursor,str(content_keys[etd3]),self.excelpath)
                    else:
                        if str(content_keys[etd3]) in temp_table_names:
                            self.ExcelToDB(attachedcursor,str(content_keys[etd3]),self.excelpath)
            if os.path.exists(oldsqldir+"\\FK_add_"+ibase[:-5]+".sql")==False:
                f = open(oldsqldir+"\\FK_add_"+ibase[:-5]+".sql","w+")
                f.write("pragma foreign_keys = off;\n\n")
                #aWriteSQL(attachedcursor,"pragma foreign_keys = off")
                f.write("begin transaction;\n\n")
                #aWriteSQL(attachedcursor,"begin transaction")
                for etd4 in range(len(table_names)):
                    if table_names[etd4][0][-4:]!="_old":
                        f.write("alter table "+str(table_names[etd4][0])+" rename to _"+str(table_names[etd4][0])+"_old;\n\n")
                        #aWriteSQL(attachedcursor,"alter table "+str(table_names[etd4][0])+" rename to _"+str(table_names[etd4][0])+"_old")
                        f.write(orig_statement[etd4]+";\n\n")
                        #aWriteSQL(attachedcursor,orig_statement[etd4])
                        f.write("insert or replace into "+str(table_names[etd4][0])+"("+','.join(insert_statement[etd4])+")\n\n")
                        f.write("select "+','.join(insert_statement[etd4])+" from _"+str(table_names[etd4][0])+"_old;\n\n")
                        #aWriteSQL(attachedcursor,"insert or replace into "+str(table_names[etd4][0])+"("+','.join(insert_statement[etd4])+") "+"select "+
                        #         ','.join(insert_statement[etd4])+" from _"+str(table_names[etd4][0])+"_old")
                        f.write("drop table _"+str(table_names[etd4][0])+"_old;\n\n")
                        #aWriteSQL(attachedcursor,"drop table _"+str(table_names[etd4][0])+"_old")
                        
                f.write("commit;\n\n")
                #aWriteSQL(attachedcursor,"commit")
                f.write("pragma foreign_keys = on;\n\n")
                #aWriteSQL(attachedcursor,"pragma foreign_keys = on")
                f.close()
            path = self.enginepath+" "+self.dbpath+" \".read '"+oldsqldir+"\\FK_add_"+ibase[:-5]+".sql'\""
            os.system(path)
            attachedconnection.close(True)
        self.finished.emit()
    def aWriteSQL(self, a,query,attempts=12,sec=0.2,msg=False):
        try:
            a.execute(query)
        except apsw.SQLError as e:
            err = str(e)
            for i in range(0,attempts):
                try:
                    a.execute(query)
                    break
                except apsw.SQLError as e:
                    if err.find("database is locked")>=0:
                        time.sleep(sec)
                        i -= 1
                    if i==attempts-1:
                        self.logger.debug(query)
                        self.logger.error(str(e))
                        if msg==True:
                            self.ctypes.windll.user32.MessageBoxW(0,str(e),"SQL Query Error",1)
                        break
                    else:
                        time.sleep(sec)
                        continue
        except apsw.ConstraintError as e2:
            err = str(e2)
            for i in range(0,attempts):
                try:
                    a.execute(query)
                    break
                except apsw.ConstraintError as e2:
                    if i==attempts-1:
                        self.logger.debug(query)
                        self.logger.error(str(e2))
                        if msg==True:
                            self.ctypes.windll.user32.MessageBoxW(0,str(e2),"SQL Query Error",1)
                        break
                    else:
                        time.sleep(sec)
                        continue
        except apsw.BusyError as e3:
            err = str(e3)
            while err.find("database is locked")>=0:
                time.sleep(sec)


    def ExcelToDB(self, a,sheet,excelpath):
        self.logger.debug("sheet attempted = "+sheet + " excelpath = "+excelpath)
        db_sheet = []
        book = load_workbook(filename=excelpath,data_only=True)
        ws = book[sheet]
        for row in ws.iter_rows():
            temp_db_sheet = []
            for cell in row:
                temp_db_sheet.append(cell.value)
            db_sheet.append(temp_db_sheet)
        db_sheet = db_sheet[1:]
        '''
        if(sheet=="batchsequence"):
            print(db_sheet)
            print("\n\n")
        '''

        self.aWriteSQL(a,"delete from "+sheet,12,0.2,True)
        pragma = self.aReadSQL(a,"pragma table_info('"+sheet+"')",12,0.2,True)
        cast_types = []
        for i in range(len(pragma)):
            pragma[i][2] = pragma[i][2].lower()
            if pragma[i][2].find("char")>=0 or pragma[i][2].find("text")>=0 or pragma[i][2].find("date")>=0 or pragma[i][2].find("datetime")>=0 or pragma[i][2].find("")>=0:
                cast_types.append(1)
            else:
                cast_types.append(0)
        for i in range(len(db_sheet)):
            entry = []
            for j in range(len(pragma)):
                '''
                if sheet=="batchsequence":
                    print("lenI = "+str(len(db_sheet)))
                    print("i = "+str(i))
                    print("lenJ = "+str(len(db_sheet[i])))
                    print("j = "+str(j))
                '''
                #DUNNO
                #print(ni(db_sheet[i][j],cast_types[j]))
                try:
                    entry.append(self.ni(db_sheet[i][j],cast_types[j]))
                except IndexError as e:
                    self.ctypes.windll.user32.MessageBoxW(0,sheet+": "+str(e)+"; check Excel structure of "+excelpath,"Table Error",1)
            entry = "insert or replace into "+sheet+" values("+','.join(entry)+")"
            #if sheet=="batchsequence":
            #    print(entry)
            self.aWriteSQL(a,entry,12,0.2,True)
        #if sheet=="batchsequence":
    #     print("finished")

    def addSlash(self, s,orig="\\"):
        prefix = s[0:2]
        suffix = s[2:]
        suffix = suffix.replace(orig,"\\\\")
        return(prefix+suffix)

    def ni(self, n,a):
        #print(str(n=='')+" ; "+str(n==None)+" ; "+str(n))
        if n=='' or n==None or n=='None':
            return("null")
        else:
            if a==1:
                n = str(n)
                n = n.replace("\'","\'\'")
                return("'"+str(n)+"'")
            else:
                return(str(n))

    def getExcelRange(self, f,rngname):
        wb = load_workbook(filename=f,data_only=True)
        
        #print("rngname = "+rngname)
        sheetname = list(wb.defined_names[rngname].destinations)[0][0]
        rangecells = list(wb.defined_names[rngname].destinations)[0][1].replace('$','')
        
        sheet = wb[str(sheetname)]
        rangecells = rangecells.split(':')
        
        cells = sheet[str(rangecells[0]):str(rangecells[1])]
        end_range = []
        for row in cells:
            row_range = []
            for cell in row:
                row_range.append(cell.value)
            end_range.append(row_range)
        wb.close()
        return(end_range)

    def getNewTableQuery(self, name,n):
        header = n[0]
        query = "create table if not exists "+name+"("
        for i in range(len(header)):
            header[i] += " text"
        query += ','.join(header)+"); pragma foreign_keys = on"
        return(query)

    def aReadSQL(self, a,query,attempts=12,sec=0.2,msg=False):
        val = []
        try:
            results = a.execute(query)
            for result in results:
                val.append(list(result))
            return(val)
        except apsw.SQLError as e:
            err = str(e)
            for i in range(0,attempts):
                try:
                    results = a.execute(query)
                    for result in results:
                        val.append(list(result))
                    return(val)
                except apsw.SQLError as e:
                    if err.find("database is locked")>=0:
                        time.sleep(sec)
                        i -= 1
                    if i==attempts-1:
                        self.logger.debug(query)
                        self.logger.error(str(e))
                        if msg==True:
                            self.ctypes.windll.user32.MessageBoxW(0,str(e),"SQL Query Error",1)
                        return(val)
                    else:
                        time.sleep(sec)
                        continue
        except apsw.ConstraintError as e2:
            err = str(e2)
            for i in range(0,attempts):
                try:
                    results = a.execute(query)
                    for result in results:
                        val.append(list(result))
                    return(val)
                except apsw.ConstraintError as e2:
                    if i==attempts-1:
                        self.logger.debug(query)
                        self.logger.error(str(e2))
                        if msg==True:
                            self.ctypes.windll.user32.MessageBoxW(0,str(e2),"SQL Query Error",1)
                        return(val)
                    else:
                        time.sleep(sec)
                        continue
        except apsw.BusyError as e3:
            err = str(e3)
            while err.find("database is locked")>=0:
                time.sleep(sec)