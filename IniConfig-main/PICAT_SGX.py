#from PyQt5.QtWidgets import QApplication, QMainWindow

from tkinter import Button
from PyQt5 import QtCore,QtGui,QtWidgets
from PyQt5.QtWidgets import *
from PyQt5 import QtGui
import os,sys,time,datetime,ctypes,shutil,winshell,win32api,subprocess,psutil
import glob,math,ctypes,openpyxl,apsw,csv,math,win32api,psycopg2,urllib.parse
from openpyxl import load_workbook
from openpyxl import Workbook
from subprocess import call
from functools import partial
from win32com.client import Dispatch
from bs4 import BeautifulSoup
import win32api,requests,webbrowser
import Components.PICAT_gui as PICAT_gui
from Components.PICAT_gui import Ui_PICAT_SM
from Components.PICAT_ini import Ui_Dialog
from Components.PICAT_dialog import Ui_ButtonDialog
from loguru import logger
from iniconfig import Parse

version = "08 march 2022 15:16"

db_cfg = Parse("SQLCONN").cfg

current_filename = os.path.basename(sys.argv[0]).split(".")[0]

file_cfg = Parse(current_filename.upper()).cfg
logger.debug(f"filename: {current_filename.upper()} + filecfg_arr = {file_cfg}")
form_name = None
if not file_cfg:
    logger.error(f"No [{current_filename}] header in config.ini")
    #pop up with ctypes
    ctypes.windll.user32.MessageBoxW(0,"No [{}] header in config.ini, please check file name and capitalization within config.ini".format(current_filename.upper()),"PICAT Error",1)
    #exit progrm
    sys.exit()
else:
    form_name = file_cfg["form"]

connectionpath = db_cfg["connectionpath"] if db_cfg["connectionpath"] else os.getcwd()
connection_db ="\\" + db_cfg["connection"]
logger.debug(f"connectionpath = {connectionpath}, connection_db = {connection_db}")


logger.debug(f"Form name is: {form_name}")

picat_cfg = Parse("APPLICATION VARIABLES").cfg
PG_PASS = picat_cfg["pg_pass"]

signalsprefix = picat_cfg["signalsprefix"]

#
# SQLite Functions



def ReadSQL(query,attempts=12,sec=0.2,msg=False):
    val = []####
    try:
        results = cursor.execute(query)
        for result in results:
            val.append(list(result))
        return(val)
    except apsw.SQLError as e:
        err = str(e)
        for i in range(0,attempts):
            try:
                results = cursor.execute(query)
                for result in results:
                    val.append(list(result))
                return(val)
            except apsw.SQLError as e:
                if err.find("database is locked")>=0:
                    time.sleep(sec)
                    i -= 1
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e),"SQL Query Error",1)
                    return(val)
                else:
                    time.sleep(sec)
                    continue
    except apsw.ConstraintError as e2:
        err = str(e2)
        for i in range(0,attempts):
            try:
                results = cursor.execute(query)
                for result in results:
                    val.append(list(result))
                return(val)
            except apsw.ConstraintError as e2:
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e2))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e2),"SQL Query Error",1)
                    return(val)
                else:
                    time.sleep(sec)
                    continue
    except apsw.BusyError as e3:
        err = str(e3)
        while err.find("database is locked")>=0:
            time.sleep(sec)
    
def WriteSQL(query,attempts=12,sec=0.2,msg=False):
    try:
        cursor.execute(query)
    except apsw.SQLError as e:
        err = str(e)
        for i in range(0,attempts):
            try:
                cursor.execute(query)
                break
            except apsw.SQLError as e:
                if err.find("database is locked")>=0:
                    time.sleep(sec)
                    i -= 1
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e),"SQL Query Error",1)
                    break
                else:
                    time.sleep(sec)
                    continue
    except apsw.ConstraintError as e2:
        err = str(e2)
        for i in range(0,attempts):
            try:
                cursor.execute(query)
                break
            except apsw.ConstraintError as e2:
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e2))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e2),"SQL Query Error",1)
                    break
                else:
                    time.sleep(sec)
                    continue
    except apsw.BusyError as e3:
        err = str(e3)
        while err.find("database is locked")>=0:
            time.sleep(sec)

def aReadSQL(a,query,attempts=12,sec=0.2,msg=False):
    val = []
    try:
        results = a.execute(query)
        for result in results:
            val.append(list(result))
        return(val)
    except apsw.SQLError as e:
        err = str(e)
        for i in range(0,attempts):
            try:
                results = a.execute(query)
                for result in results:
                    val.append(list(result))
                return(val)
            except apsw.SQLError as e:
                if err.find("database is locked")>=0:
                    time.sleep(sec)
                    i -= 1
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e),"SQL Query Error",1)
                    return(val)
                else:
                    time.sleep(sec)
                    continue
    except apsw.ConstraintError as e2:
        err = str(e2)
        for i in range(0,attempts):
            try:
                results = a.execute(query)
                for result in results:
                    val.append(list(result))
                return(val)
            except apsw.ConstraintError as e2:
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e2))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e2),"SQL Query Error",1)
                    return(val)
                else:
                    time.sleep(sec)
                    continue
    except apsw.BusyError as e3:
        err = str(e3)
        while err.find("database is locked")>=0:
            time.sleep(sec)
    
def aWriteSQL(a,query,attempts=12,sec=0.2,msg=False):
    try:
        a.execute(query)
    except apsw.SQLError as e:
        err = str(e)
        for i in range(0,attempts):
            try:
                a.execute(query)
                break
            except apsw.SQLError as e:
                if err.find("database is locked")>=0:
                    time.sleep(sec)
                    i -= 1
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e),"SQL Query Error",1)
                    break
                else:
                    time.sleep(sec)
                    continue
    except apsw.ConstraintError as e2:
        err = str(e2)
        for i in range(0,attempts):
            try:
                a.execute(query)
                break
            except apsw.ConstraintError as e2:
                if i==attempts-1:
                    logger.debug(query)
                    logger.error(str(e2))
                    if msg==True:
                        ctypes.windll.user32.MessageBoxW(0,str(e2),"SQL Query Error",1)
                    break
                else:
                    time.sleep(sec)
                    continue
    except apsw.BusyError as e3:
        err = str(e3)
        while err.find("database is locked")>=0:
            time.sleep(sec)

def addSlash(s,orig="\\"):
    prefix = s[0:2]
    suffix = s[2:]
    suffix = suffix.replace(orig,"\\\\")
    return(prefix+suffix)

scriptpath = os.path.realpath(sys.argv[0])
scriptpath = os.path.dirname(scriptpath)+"\\"
scriptpath = addSlash(scriptpath)

# Format Functions
def FormatMR(x,d,comma=True):
    neg = ""
    if x<0:
        neg = "-"
        x = -x
    val = str(x)
    val = val.split(".")

    if comma==True:
        temp = ""
        frag = val[0]
        while len(frag)>0:
            if len(frag)-3>0:
                temp = ","+frag[-3:]+temp
                frag = frag[0:len(frag)-3]
            else:
                temp = frag+temp
                frag = ""
        if temp!="":
            val[0] = temp
    
    if len(val)>1:
        dec = float("0."+val[1])
        dec = '{:.{prec}f}'.format(dec,prec=d)
        dec = str(dec)
        dec = dec.split(".")[1]
        #print("a = "+dec)
    else:
        dec = "0"*d
        #print("b = "+str(dec))
    val[0] = int(val[0])
    val[0] = str(val[0])
    val = neg+val[0]+"."+dec
    return(val)

def copyDir(src,dest,neweronly=False):
    if os.path.isdir(src):
        if not os.path.isdir(dest):
            os.makedirs(dest)
        files = os.listdir(src)
        for f in files:
            newsrc = os.path.join(src,f)
            newdest = os.path.join(dest,f)
            if neweronly==True:
                try:
                    if os.path.exists(newdest)==True:
                        if os.stat(newsrc).st_mtime-os.stat(newdest).st_mtime>=1:
                            copyDir(newsrc,newdest,neweronly)
                    else:
                        copyDir(newsrc,newdest,neweronly)
                except:
                    pass
            else:
                copyDir(newsrc,newdest,neweronly)
    else:
        if neweronly==True:
            try:
                if os.path.exists(dest)==True:
                    if os.stat(src).st_mtime-os.stat(dest).st_mtime>=1:
                        try:
                            shutil.copyfile(src,dest)
                        except Exception as e:
                            ctypes.windll.user32.MessageBoxW(0,src+" Error: "+str(e),"",1)
                else:
                    shutil.copyfile(src,dest)
            except:
                pass
        else:
            try:
                shutil.copyfile(src,dest)
            except Exception as e:
                ctypes.windll.user32.MessageBoxW(0,src+" Error: "+str(e),"",1)

def deleteExt(path,ext,bname,row):
    if os.path.isdir(path):
        files = os.listdir(path)
        for f in files:
            newpath = os.path.join(path,f)
            deleteExt(newpath,ext,bname,row)
    else:
        if getExt(path,ext)==True:
            try:
                os.remove(path)
            except:
                ctypes.windll.user32.MessageBoxW(0,"Problem deleting "+path+"?","Failed cleardirext: "+bname+"! \\"+str(row),0)

def getExt(baseN,fileN):
    usebase = os.path.basename(baseN)
    #more than two asterisks
    if fileN.count('*')>2:
        return(False)
    #two asterisks no dot sandwich
    elif fileN.count('*')==2 and fileN.find("*.*")==-1:
        return(False)
    #pure "*.*"
    if fileN=="*.*":
        return(True)
    #no asterisk, pure string search
    elif fileN.find("*")==-1 and usebase.find(fileN)>=0:
        return(True)
    #asterisk(s)
    elif fileN.find("*")>=0:
        # some.something*
        if fileN.endswith("*")==True and usebase.startswith(fileN[:-1])==True:
            return(True)
        # *some.something
        elif fileN.startswith("*")==True and usebase.endswith(fileN[1:])==True:
            return(True)
        #incomplete *.*
        elif fileN.endswith("*.")==True:
            fileN = fileN+"*"
        elif fileN.startswith(".*")==True:
            fileN = "*"+fileN
        #uses *.*
        if fileN.find("*.*")>=0:
            #any name & partial extension
            if fileN.startswith("*.*")==True:
                return(usebase.endswith(fileN[3:]))
            #partial name & any extension
            elif fileN.endswith("*.*")==True:
                return(usebase.startswith(fileN[0:-3]))
            #partial name & partial extension
            else:
                arr = fileN.split('*.*')
                return(usebase.startswith(arr[0]) and usebase.endswith(arr[1]))
        #uses *.
        elif fileN.find("*.")>=0:
            arr = fileN.split('*.')
            return(usebase.startswith(arr[0]) and usebase.endswith("."+arr[1]))
        #uses .*
        elif fileN.find(".*")>=0:
            arr = fileN.split('.*')
            return(usebase.startswith(arr[0]+".") and usebase.endswith(arr[1]))
        #no other scenario
        else:
            return(False)
    #no other scenario
    else:
        return(False)

if not connection_db and not connectionpath:
    result = ctypes.windll.user32.MessageBoxW("SQLITE DB Connection details not found in config.ini","",1)
    if result==1:
        sys.exit()

db_conn = connectionpath + connection_db
if os.path.exists(db_conn):
    connection = apsw.Connection(db_conn)
    cursor = connection.cursor()
else:
    result = ctypes.windll.user32.MessageBoxW(0,f"failed to locate {db_conn}","",1)
    if result==1:
        sys.exit()


# Create tables
WriteSQL("create table if not exists forms(formname char(63) "+
         "primary key,formdesc text); pragma foreign_keys = on")

'''
WriteSQL("create table if not exists tabs(formname char(63),"+
         "tab char(63),tabsequence integer,tabdesc text,treepath char(1023),"+
         "primary key(formname,tab),foreign key(formname) references forms"+
         "(formname) on delete no action on update no action); pragma foreign_keys = on")
'''

WriteSQL("create table if not exists tabs(formname char(63),"+
         "tab char(63),tabsequence integer,grid integer,"+
         "tabdesc text,treepath char(1023),primary key(formname,tab),foreign key(formname) "+
         "references forms(formname) on delete no action on update no action); pragma foreign_keys = on")


WriteSQL("create table if not exists buttons(formname char(63),"+
         "tab char(63),buttonname char(63),buttonsequence "+
         "integer,columnnum integer,buttondesc text,buttongroup char(63),active char(63),"+
         "treepath char(1023),primary key(formname,tab,"+
         "buttonname),foreign key(formname,tab) references tabs(formname,"+
         "tab) on delete no action on update no action); pragma "+
         "foreign_keys = on")

WriteSQL("create table if not exists sqlengine(enginepathfile char(1023) "+
         "primary key); pragma foreign_keys = on")

WriteSQL("create table if not exists batchsequence(formname char(63),tab "+
         "char(63),buttonname char(63),runsequence integer,"+
         "folderpath char(1023),filename char(255),type char(63),"+
         "source char(1023),target char(1023),databasepath char(1023),"+
         "databasename char(255),keypath char(1023),"+
         "keyfile char(255),treepath char(1023),primary key(formname,tab,buttonname,runsequence),"+
         "foreign key(formname,tab,buttonname) references buttons(formname,"+
         "tab,buttonname) on delete no action on update no action); pragma "+
         "foreign_keys = on")

WriteSQL("create table if not exists mt4config(formname char(63),"+
         "tab char(63),buttonname char(63),affinity integer,"+
         "profile char(63),marketwatch char(255),login integer,password "+
         "char(255),server char(255),suffix text,autoconfig integer,sym char(20),period char(4),"+
         "template char(63),trademode char(63),createdon datetime,active char(63),"+
         "primary key(formname,tab,buttonname,login),foreign key(formname,tab,buttonname) "+
         "references batchsequence(formname,tab,buttonname,runsequence) on delete no action on "+
         "update no action); pragma foreign_keys = on")

def writeTable(path,file):
    if os.path.exists(path)==True:
        if path[0:1]=="\\":
            path = path[1:]
        newpath = path.replace("\\","\\\\")
        WriteSQL("attach database '"+newpath+"\\\\"+file+"' as DB")
        sql_state = ReadSQL("select tbl_name,sql as sql_create from DB.sqlite_master "+
                            "where type = 'table'")
        WriteSQL("begin")
        for i in range(len(sql_state)):
            get_fields = ReadSQL("pragma table_info('"+sql_state[i][0]+"')")
            fields = ""
            for j in range(len(get_fields)):
                if j>0:
                    fields += ","+get_fields[j][1]
                else:
                    fields += get_fields[j][1]
            WriteSQL("update masttbls set fields = '"+fields+"' where tbl_name = '"+sql_state[i][0]+"'")
            WriteSQL("update mastdbtbls set fields = '"+fields+"', reqpopulate = null where tbl_name = '"+
                     sql_state[i][0]+"' and databasefile = '"+file+"'")
        WriteSQL("commit")
        WriteSQL("detach database DB")
    else:
        ctypes.windll.user32.MessageBoxW(0,path+" does not exist?","",1)

def ni(n,a):
    #print(str(n=='')+" ; "+str(n==None)+" ; "+str(n))
    if n=='' or n==None or n=='None':
        return("null")
    else:
        if a==1:
            n = str(n)
            n = n.replace("\'","\'\'")
            return("'"+str(n)+"'")
        else:
            return(str(n))

def getExcelRange(f,rngname):
    wb = load_workbook(filename=f,data_only=True)
    
    #print("rngname = "+rngname)
    sheetname = list(wb.defined_names[rngname].destinations)[0][0]
    rangecells = list(wb.defined_names[rngname].destinations)[0][1].replace('$','')
    
    sheet = wb[str(sheetname)]
    rangecells = rangecells.split(':')
    
    cells = sheet[str(rangecells[0]):str(rangecells[1])]
    end_range = []
    for row in cells:
        row_range = []
        for cell in row:
            row_range.append(cell.value)
        end_range.append(row_range)
    wb.close()
    return(end_range)

def getNewTableQuery(name,n):
    header = n[0]
    query = "create table if not exists "+name+"("
    for i in range(len(header)):
        header[i] += " text"
    query += ','.join(header)+"); pragma foreign_keys = on"
    return(query)

def ExcelToDB(a,sheet,excelpath):
    logger.debug("sheet attempted = "+sheet)
    db_sheet = []
    book = load_workbook(filename=excelpath,data_only=True)
    ws = book[sheet]
    for row in ws.iter_rows():
        temp_db_sheet = []
        for cell in row:
            temp_db_sheet.append(cell.value)
        db_sheet.append(temp_db_sheet)
    db_sheet = db_sheet[1:]
    '''
    if(sheet=="batchsequence"):
        print(db_sheet)
        print("\n\n")
    '''

    aWriteSQL(a,"delete from "+sheet,12,0.2,True)
    pragma = aReadSQL(a,"pragma table_info('"+sheet+"')",12,0.2,True)
    cast_types = []
    for i in range(len(pragma)):
        pragma[i][2] = pragma[i][2].lower()
        if pragma[i][2].find("char")>=0 or pragma[i][2].find("text")>=0 or pragma[i][2].find("date")>=0 or pragma[i][2].find("datetime")>=0 or pragma[i][2].find("")>=0:
            cast_types.append(1)
        else:
            cast_types.append(0)
    for i in range(len(db_sheet)):
        entry = []
        for j in range(len(pragma)):
            '''
            if sheet=="batchsequence":
                print("lenI = "+str(len(db_sheet)))
                print("i = "+str(i))
                print("lenJ = "+str(len(db_sheet[i])))
                print("j = "+str(j))
            '''
            #DUNNO
            #print(ni(db_sheet[i][j],cast_types[j]))
            try:
                entry.append(ni(db_sheet[i][j],cast_types[j]))
            except IndexError as e:
                ctypes.windll.user32.MessageBoxW(0,sheet+": "+str(e)+"; check Excel structure of "+excelpath,"Table Error",1)
        entry = "insert or replace into "+sheet+" values("+','.join(entry)+")"
        #if sheet=="batchsequence":
        #    print(entry)
        aWriteSQL(a,entry,12,0.2,True)
    #if sheet=="batchsequence":
   #     print("finished")

def RangeToDB(a,rngname,excelpath):
    logger.debug("begin range name function")
    rng = getExcelRange(excelpath,rngname)
    logger.debug("range name "+rngname+" retrieved")

    aWriteSQL(a,getNewTableQuery(rngname,rng))
    logger.debug("create statement created")

    aWriteSQL(a,"delete from "+rngname,12,0.2,True)
    logger.debug(rngname+" table purged")
    pragma = aReadSQL(a,"pragma table_info('"+rngname+"')",12,0.2,True)
    cast_types = []
    logger.debug("retrieving cast types")
    for i in range(len(pragma)):
        pragma[i][2] = pragma[i][2].lower()
        if pragma[i][2].find("char")>=0 or pragma[i][2].find("text")>=0 or pragma[i][2].find("date")>=0 or pragma[i][2].find("datetime")>=0 or pragma[i][2].find("")>=0:
            cast_types.append(1)
        else:
            cast_types.append(0)
    logger.debug("begin writing to DB")
    logger.debug("data X = "+str(len(rng))+" ; data Y = "+str(len(rng[0]))+" ; cast_types = "+str(len(cast_types)))
    for i in range(1,len(rng)):
        entry = []
        for j in range(len(pragma)):
            entry.append(ni(rng[i][j],cast_types[j]))
        entry = "insert or replace into "+rngname+" values("+','.join(entry)+")"
        aWriteSQL(a,entry,12,0.2,True)
    logger.debug("end writing to DB")

def runQuery(excelpath,enginepath,sourcepath,dbpath,queryname):
    #if os.path.exists(excelpath)==True:
    book = Workbook()
    filepath = excelpath
    wb = openpyxl.Workbook(filepath)
    
    queryresult = ""
    query = open(sourcepath).read()
    query = query.split('\n\n')
    
    attachedconnection = apsw.Connection(dbpath)
    attachedcursor = attachedconnection.cursor()
    if len(query)>0:
        for i in range(len(query)):
            prefix = query[i][:6].upper()
            if prefix=="ATTACH" or prefix=="DETACH":
                aWriteSQL(attachedcursor,query[i])
            elif prefix=="SELECT":
                queryresult = aReadSQL(attachedcursor,query[i])
        
    #queryresult = aReadSQL(attachedcursor,query)
    attachedconnection.close(True)
    
    if queryname not in wb.sheetnames:
        wb.create_sheet(queryname)
    wb.save(filepath)
    
    book = load_workbook(filename=filepath,data_only=True)
    ws = book[queryname]
    for i in range(len(queryresult)):
        for j in range(len(queryresult[i])):
           ws.cell(row=i+1,column=j+1).value = queryresult[i][j]
    book.save(filepath)

def runRange(dbpath,sourcetable,excelpath):
    attachedconnection = apsw.Connection(dbpath)
    attachedcursor = attachedconnection.cursor()
    if sourcetable!="" and sourcetable!=None:
        RangeToDB(attachedcursor,sourcetable,excelpath)
    attachedconnection.close(True)
    
def  runExcel(mode,excelpath,enginepath,dbpath,sqldir,sourcetable,userange,preserve):
    if mode==1:
        if os.path.exists(excelpath)==True:
            book = Workbook()
        filepath = excelpath
        if preserve==1:
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook(filepath)
        
        attachedconnection = apsw.Connection(dbpath)
        attachedcursor = attachedconnection.cursor()
        append = ""
        if sourcetable!="" and sourcetable!="None" and sourcetable!=None:
            append = " and tbl_name = '"+sourcetable+"'"
        table_names = aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master where tbl_name <> 'pyexcelmenu' and tbl_name <> 'sqlengine'"+append)
        table_con = []
        for tn in range(len(table_names)):
            table_names[tn] = str(table_names[tn][0])
        
        for tn in range(len(table_names)):
            if table_names[tn] not in wb.sheetnames:
                wb.create_sheet(table_names[tn])
                db_con = aReadSQL(attachedcursor,"select * from "+table_names[tn])
                pragma = aReadSQL(attachedcursor,"pragma table_info('"+table_names[tn]+"')")
                field_names = []
                for p in range(len(pragma)):
                    field_names.append(str(pragma[p][1]))
                db_con.insert(0,field_names)
                table_con.append(db_con)
            else:
                ws = wb[str(table_names[tn])]
                for x in range(wb[str(table_names[tn])].min_row-1,wb[str(table_names[tn])].max_row):
                    for y in range(wb[str(table_names[tn])].min_column-1,wb[str(table_names[tn])].max_column):
                        ws.cell(x+1,y+1).value = None
                db_con = aReadSQL(attachedcursor,"select * from "+table_names[tn])
                pragma = aReadSQL(attachedcursor,"pragma table_info('"+table_names[tn]+"')")
                field_names = []
                for p in range(len(pragma)):
                    field_names.append(str(pragma[p][1]))
                db_con.insert(0,field_names)
                table_con.append(db_con)
        attachedconnection.close(True)
        while True:
            try:
                wb.save(filepath)
                logger.debug("file has been saved to: " + filepath)
                ctypes.windll.user32.MessageBoxW(0,"File has been saved to: " + filepath, "File Saved!",0)

            except PermissionError:
                logger.warning(f"{filepath}  is open, please close and try again")
                ctypes.windll.user32.MessageBoxW(0,filepath + " is open, please close and try again", "Failed exe!",0)

                continue
            else:
                break
        
        book = load_workbook(filename=filepath,data_only=True)
        for tn in range(len(book.sheetnames)):
            ws = book[book.sheetnames[tn]]
            if book.sheetnames[tn] in table_names:
                index = table_names.index(book.sheetnames[tn])
                for i in range(len(table_con[index])):
                    for j in range(len(table_con[index][i])):
                        ws.cell(row=i+1,column=j+1).value = table_con[index][i][j]
        book.save(filepath)
    elif mode==0:
        sqldir = addSlash(sqldir)
        attachedconnection = apsw.Connection(dbpath)
        attachedcursor = attachedconnection.cursor()
        table_names = aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master where tbl_name <> 'pyexcelmenu' and tbl_name <> 'sqlengine' order by tbl_name asc",12,0.2,True)
        build_statement = []
        insert_statement = []
        orig_statement = []
        temp_table_names = []
        for etd0 in range(len(table_names)):
            sqlite_stat = table_names[etd0][0]
            if table_names[etd0][0][-4:]!="_old" and sqlite_stat.find("sqlite_stat")==-1:
                temp_table_names.append(table_names[etd0])
        table_names = temp_table_names
        idir = os.path.dirname(excelpath)
        ibase = os.path.basename(excelpath)
        logger.debug(sqldir+"\\FK_rem_"+ibase[:-5]+".sql")
        if os.path.exists(sqldir+"\\FK_rem_"+ibase[:-5]+".sql")==False:
            f = open(sqldir+"\\FK_rem_"+ibase[:-5]+".sql","w+")
            for etd1 in range(len(table_names)):
                temporig = aReadSQL(attachedcursor,"select sql from sqlite_master where tbl_name = '"+str(table_names[etd1][0])+"'",12,0.2,True)
                if len(temporig)>0:
                    orig_statement.append(temporig[0][0])
                pragma_info = aReadSQL(attachedcursor,"pragma table_info('"+str(table_names[etd1][0])+"')",12,0.2,True)
                entry = []
                pkeys = []
                insert_statement2 = []
                for pragrow in range(len(pragma_info)):
                    if pragma_info[pragrow][1].lower()=="table":
                        pragma_info[pragrow][1] = "["+pragma_info[pragrow][1]+"]"
                    insert_statement2.append(str(pragma_info[pragrow][1]))
                    entry.append(str(pragma_info[pragrow][1])+" "+str(pragma_info[pragrow][2]))
                    if int(pragma_info[pragrow][5])>0:
                        pkeys.append(str(pragma_info[pragrow][1]))
                if len(entry)>0:
                    build_statement.append("create table if not exists "+str(table_names[etd1][0])+"("+','.join(entry))
                    if len(pkeys)>0:
                        build_statement[len(build_statement)-1] += ",primary key("+','.join(pkeys)+")); pragma foreign_keys = on"
                    else:
                        build_statement[len(build_statement)-1] += "); pragma foreign_keys = on"
                insert_statement.append(insert_statement2)
            f.write("pragma foreign_keys = off;\n\n")
            #aWriteSQL(attachedcursor,"pragma foreign_keys = off")
            f.write("begin transaction;\n\n")
            #aWriteSQL(attachedcursor,"begin transaction")
            for etd2 in range(len(table_names)):
                f.write("drop table if exists _"+str(table_names[etd2][0])+"_old;\n\n")
                #aWriteSQL(attachedcursor,"drop table if exists _"+str(table_names[etd2][0])+"_old")
                f.write("alter table "+str(table_names[etd2][0])+" rename to _"+str(table_names[etd2][0])+"_old;\n\n")
                #aWriteSQL(attachedcursor,"alter table "+str(table_names[etd2][0])+" rename to _"+str(table_names[etd2][0])+"_old")
                f.write(build_statement[etd2]+";\n\n")
                #aWriteSQL(attachedcursor,build_statement[etd2])
                f.write("insert into "+str(table_names[etd2][0])+"("+','.join(insert_statement[etd2])+")\n\n")
                f.write("select "+','.join(insert_statement[etd2])+" from _"+str(table_names[etd2][0])+"_old;\n\n")
                #aWriteSQL(attachedcursor,"insert into "+str(table_names[etd2][0])+"("+','.join(insert_statement[etd2])+") "+"select "+
                #         ','.join(insert_statement[etd2])+" from _"+str(table_names[etd2][0])+"_old")
                f.write("drop table _"+str(table_names[etd2][0])+"_old;\n\n")
                #aWriteSQL(attachedcursor,"drop table _"+str(table_names[etd2][0])+"_old")
            f.write("commit;\n\n")
            #aWriteSQL(attachedcursor,"commit")
            f.write("pragma foreign_keys = on;")
            #aWriteSQL(attachedcursor,"pragma foreign_keys = on")
            f.close()


        else:
            for etd1 in range(len(table_names)):
                temporig = aReadSQL(attachedcursor,"select sql from sqlite_master where tbl_name = '"+str(table_names[etd1][0])+"'",12,0.2,True)
                if len(temporig)>0:
                    orig_statement.append(temporig[0][0])
                pragma_info = aReadSQL(attachedcursor,"pragma table_info('"+str(table_names[etd1][0])+"')",12,0.2,True)
                entry = []
                pkeys = []
                insert_statement2 = []
                for pragrow in range(len(pragma_info)):
                    if pragma_info[pragrow][1].lower()=="table":
                        pragma_info[pragrow][1] = "["+pragma_info[pragrow][1]+"]"
                    insert_statement2.append(str(pragma_info[pragrow][1]))
                    entry.append(str(pragma_info[pragrow][1])+" "+str(pragma_info[pragrow][2]))
                    if int(pragma_info[pragrow][5])>0:
                        pkeys.append(str(pragma_info[pragrow][1]))
                if len(entry)>0:
                    build_statement.append("create table if not exists "+str(table_names[etd1][0])+"("+','.join(entry))
                    if len(pkeys)>0:
                        build_statement[len(build_statement)-1] += ",primary key("+','.join(pkeys)+")); pragma foreign_keys = on"
                    else:
                        build_statement[len(build_statement)-1] += "); pragma foreign_keys = on"
                insert_statement.append(insert_statement2)

        oldsqldir = sqldir
        sqldir = addSlash(sqldir)
        path = enginepath+" "+dbpath+" \".read '"+sqldir+"\\\\FK_rem_"+ibase[:-5]+".sql'\""
        var = os.system(path)
        if int(var)!=0 and str(var)!=None and str(var)!="":
            ctypes.windll.user32.MessageBoxW(0,str(var),"SQL Query Error",1)

        if os.path.exists(excelpath)==True:
            book = load_workbook(filename=excelpath,data_only=True)
            content_keys = book.sheetnames
        else:
            content_keys = []

        if userange==1:
            if sourcetable!="" and sourcetable!=None:
                RangeToDB(attachedcursor,sourcetable,excelpath)
        else:
            temp_table_names = []
            for tn in range(len(table_names)):
                temp_table_names.append(table_names[tn][0])
            for etd3 in range(len(content_keys)):
                if sourcetable!="" and sourcetable!=None and sourcetable!="None":
                    if str(content_keys[etd3])==sourcetable:
                        ExcelToDB(attachedcursor,str(content_keys[etd3]),excelpath)
                else:
                    if str(content_keys[etd3]) in temp_table_names:
                        ExcelToDB(attachedcursor,str(content_keys[etd3]),excelpath)
        if os.path.exists(oldsqldir+"\\FK_add_"+ibase[:-5]+".sql")==False:
            f = open(oldsqldir+"\\FK_add_"+ibase[:-5]+".sql","w+")
            f.write("pragma foreign_keys = off;\n\n")
            #aWriteSQL(attachedcursor,"pragma foreign_keys = off")
            f.write("begin transaction;\n\n")
            #aWriteSQL(attachedcursor,"begin transaction")
            for etd4 in range(len(table_names)):
                if table_names[etd4][0][-4:]!="_old":
                    f.write("alter table "+str(table_names[etd4][0])+" rename to _"+str(table_names[etd4][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"alter table "+str(table_names[etd4][0])+" rename to _"+str(table_names[etd4][0])+"_old")
                    f.write(orig_statement[etd4]+";\n\n")
                    #aWriteSQL(attachedcursor,orig_statement[etd4])
                    f.write("insert or replace into "+str(table_names[etd4][0])+"("+','.join(insert_statement[etd4])+")\n\n")
                    f.write("select "+','.join(insert_statement[etd4])+" from _"+str(table_names[etd4][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"insert or replace into "+str(table_names[etd4][0])+"("+','.join(insert_statement[etd4])+") "+"select "+
                    #         ','.join(insert_statement[etd4])+" from _"+str(table_names[etd4][0])+"_old")
                    f.write("drop table _"+str(table_names[etd4][0])+"_old;\n\n")
                    #aWriteSQL(attachedcursor,"drop table _"+str(table_names[etd4][0])+"_old")
                    
            f.write("commit;\n\n")
            #aWriteSQL(attachedcursor,"commit")
            f.write("pragma foreign_keys = on;\n\n")
            #aWriteSQL(attachedcursor,"pragma foreign_keys = on")
            f.close()
        path = enginepath+" "+dbpath+" \".read '"+oldsqldir+"\\FK_add_"+ibase[:-5]+".sql'\""
        os.system(path)
        attachedconnection.close(True)
        logger.success(f"{excelpath} has been imported")
        ctypes.windll.user32.MessageBoxW(0,excelpath +" excelpath","File Imported",0)

class DB_Window(QMainWindow,PICAT_gui.Ui_PICAT_SM):
    def __init__(self,parent=None):
        super(DB_Window,self).__init__(parent)
        self.setupUi(self)
        QtWidgets.QApplication.setStyle('windowsvista')

        # Get Form Title & Description
        formtitle = ReadSQL("select formname from forms")
        if len(formtitle)<1:
            formtitle = 'Untitled'
        else:
            formtitle = []
            for item in ReadSQL("select formname from forms"):
                formtitle.append(item[0])

        if form_name:
            if form_name in formtitle:
                self.title = form_name
            else:
                new_formdesc = str(QInputDialog.getText(None,"New form description","New form description:",QLineEdit.Normal,"")[0])
                if not new_formdesc:
                    ctypes.windll.user32.MessageBoxW(0,"Form name not found in DB. exiting.","FAIL",0)
                    sys.exit()
                #INSERT INTO forms (formname, formdesc)
                WriteSQL("insert into forms (formname, formdesc) values ('"+form_name+"','"+new_formdesc+"')")
                self.title = form_name


        else:
            self.title = formtitle


        formdesc = ReadSQL(f"select formdesc from forms where formname='{self.title}'")
        if len(formdesc)<1:
            formdesc = 'Untitled'
        else:
            formdesc = ReadSQL(f"select formdesc from forms where formname='{self.title}'")[0][0]
        self.setWindowTitle(formdesc)
        self.edit_mode = False
        self.cur_seq = None

        # File Tree
        self.actionOpen_Files_Explorer.triggered.connect(self.Open_File_Tree)

        # Add EXE
        self.actionAdd_exe.triggered.connect(self.Add_EXE)

        self.actionAdd_Seq.triggered.connect(self.Add_Seq)

        
        #add URL
        self.actionAdd_url.triggered.connect(self.Add_url)

        self.actionAdd_Folder.triggered.connect(self.Add_Folder)

        self.actionEdit_mode.toggled.connect(self.actionEdit_mode_toggled)
        self.actionEdit_mode.setChecked(self.edit_mode)

        #copy file
        self.actionCopy.triggered.connect(self.Copy_File)

        # Refresh
        self.actionRefresh.triggered.connect(self.Refresh)
        
        self.actionTabsize.triggered.connect(self.ChangeTabSize)

        self.actionTab.triggered.connect(self.NewTab)

        self.actionTabUrl.triggered.connect(self.NewTabUrl)

        self.actionTabFolder.triggered.connect(self.NewTabFolder)

        self.actionOpenTabUrl.triggered.connect(self.OpenTabUrl)

        # Import Excel
        self.actionImport_from_Excel.triggered.connect(self.ImportExcel)

        # Export Excel
        self.actionExport_to_Excel.triggered.connect(self.ExportExcel)

        # Last Saved Tab
        self.lasttab = ""
        
        #Main: Set Up Tabs Filter
        #self.Main_CB_Tabs.addItem("")
        tablist = ReadSQL("select tab from tabs where formname = '"+self.title+"'")
        if len(tablist)>0:
            for tl in range(len(tablist)):
                self.Main_CB_Tabs.addItem(str(tablist[tl][0]))
        self.Main_CB_Tabs.currentIndexChanged.connect(self.on_tab_change)

        #Main: Set Up Buttons Filter
        self.Main_CB_Buttons.currentIndexChanged.connect(self.on_button_change)

        #Main: Insert Record
        self.Main_Button_InsertRec.clicked.connect(self.on_InsertRec)

        #Main: Delete Record
        self.Main_Button_DeleteRec.clicked.connect(self.on_DeleteRec)

        #Main: Main Table
        self.Main_Table.itemChanged.connect(self.on_CellChange)



        #Create Tabs & Buttons
        self.Refresh()

        self.actionAbout.triggered.connect(self.AboutMsg_Try)

        self.Refresh()

    

    def actionEdit_mode_toggled(self,state):
        if state:
            self.edit_mode = True
        else:
            self.edit_mode = False
    def Add_Seq(self):
        if self.cur_seq:
            ctypes.windll.user32.MessageBoxW(0,"Please stop current sequence first","FAIL",0)
            return
        self.cur_seq = Create_sequence(self)
        self.cur_seq.show()

    def Copy_File(self):
        curtab = self.SM_Tabs.currentIndex()
        if curtab<1:
            logger.error("current tab is first tab")
            ctypes.windll.user32.MessageBoxW(0,"current tab is first tab","Error",0)

            return
        new_file_master = QFileDialog.getOpenFileName(None,'Open file','')[0]
        if not new_file_master:
            return
        #new file without extension and path
        new_file = new_file_master.split("/")[-1]
        #remove file name and extension from path
        start_location = os.path.dirname(new_file_master)

        #combine elements in new_file_location
        
        #open select folder dialog
        end_location = QFileDialog.getExistingDirectory(None,'Open Folder','')
        if not end_location:
            return

        dest_name =  str(QInputDialog.getText(None,"Destination File Name","Destination File Name:",QLineEdit.Normal, new_file)[0])
        if not dest_name:
            return
        #copy file
        new_button = str(QInputDialog.getText(None,"New Copy Button Name","New Copy Button Name:",QLineEdit.Normal,"Copy " + dest_name)[0])
        if not new_button:
            return

  
        copy_exec = self.type_copy("Copy " + dest_name,new_file,start_location,end_location, 1, dest_name)
        
        if copy_exec:
            logger.success(f"{dest_name} has been copied to {end_location}")
            ctypes.windll.user32.MessageBoxW(0,dest_name +" has been copied to "+end_location,"File Copied",0)
        else:
            logger.error(f"{dest_name} has not been copied to {end_location}")
            return

        self.Add_EXE_FOR_COPY(start_location, end_location, new_file, new_button, dest_name)
        self.Refresh()


    def Add_EXE_FOR_COPY(self, start_location, end_location, new_file_name, new_button, dest_name):
        curtab = self.SM_Tabs.currentIndex()
        if curtab<1:
            
            logger.error("current tab is first tab")
            ctypes.windll.user32.MessageBoxW(0,"current tab is first tab","Error",0)
            return

        curtab = str(self.SM_Tabs.tabText(curtab))

       
        if not new_button:
            logger.error("new_button is None or ''")
            return

        if not new_file_name:
            logger.error("new_file_name is None or ''")
            return
        if not start_location:
            logger.error("start_location is None or ''")
            return
        if not end_location:
            logger.error("end_location is None or ''")
            return

        does_exist = ReadSQL("select buttonname from buttons where tab = '"+curtab+"' and buttonname = '"+new_button+"'")
        
        while len(does_exist)>0:
            new_button = str(QInputDialog.getText(None,"New EXE ButtonName","Name:",QLineEdit.Normal,"")[0])
            if new_button!=None and new_button!="":
                does_exist = ReadSQL("select buttonname from buttons where tab = '"+curtab+"' and buttonname = '"+new_button+"'")

        maxbut = ReadSQL("select max(buttonsequence) from buttons where tab = '"+curtab+"'")
        if len(maxbut) < 1:
            logger.error("Max button is less then 1")

        print(end_location)
        idir = os.path.dirname(end_location)
        if idir[-1:]!="\\": #this is redundant as he is using os.path.dirname.
            idir = idir+"\\"
        ibase = os.path.basename(end_location)
        maxbut = maxbut[0][0] + 1 if maxbut[0][0] else 1
        maxcolumn = ReadSQL("select max(columnnum) from buttons where tab = '"+curtab+"'")
        if not maxcolumn:
            maxcolumn = "null"
        elif not maxcolumn[0][0]:
            maxcolumn = "null"
        else:
            maxcolumn = "1" #putting button automatically in column 1
        formtitle = self.title
        WriteSQL(f"""insert into buttons(formname,tab,buttonname,buttonsequence,columnnum) values('{formtitle}', '{curtab}','{new_button}',{maxbut},{maxcolumn}) """)
        start_location = start_location.replace("/","\\")
        end_location = end_location.replace("/","\\")
        WriteSQL(f"""insert into batchsequence(formname,tab,buttonname,runsequence, filename,type, source, target, databasename) values('{formtitle}', '{curtab}','{new_button}',1,'{new_file_name}','copy','{start_location}','{end_location}','{dest_name}') """) 


    def NewTab(self):
        new_tab = str(QInputDialog.getText(None,"New Tab Name","Name:",QLineEdit.Normal,"")[0])
        if not new_tab:
            return
        if new_tab in self.SM_Tabs.tabText(1):
            logger.error("Tab already exists")
            ctypes.windll.user32.MessageBoxW(0,"Tab already exists","Error",0)
            return
        #get largest tabsequence based on formname\
        max_tab = ReadSQL("select max(tabsequence) from tabs where formname = '"+self.title+"'")
        
        #write to sql database the tab
        if not max_tab:
            max_tab =  1
            WriteSQL(f"""insert into tabs(formname,tab,tabsequence) values('{self.title}','{new_tab}',{max_tab})""")
        else:
            WriteSQL(f"""insert into tabs(formname,tab,tabsequence) values('{self.title}','{new_tab}',{max_tab[0][0]+1 if max_tab[0][0] else 1})""")
        self.Refresh()

  
    def ChangeTabSize(self):
        #get current tab size and check if it is above 900 px, then set to 800 px

        
        curtab = self.SM_Tabs.currentIndex()    
        curtab = str(self.SM_Tabs.tabText(curtab))
        cur_width = self.width()
        cur_height = self.height()
        WriteSQL(f"""update tabs set tabsize = '{cur_width},{cur_height}' where tab = '{curtab}' and formname = '{self.title}'""")
        self.Refresh()
        self.Refresh()

    def NewTabUrl(self):
            curtab = self.SM_Tabs.currentIndex()
            curtab = str(self.SM_Tabs.tabText(curtab))
            #select taburl from tabs where formname and tab
            taburl = ReadSQL("select taburl from tabs where formname = '"+self.title+"' and tab = '"+curtab+"'")
            if len(taburl)>0:
                taburl = taburl[0][0]
            else:
                taburl = ""
            new_url = str(QInputDialog.getText(None,"New URL","URL:",QLineEdit.Normal,taburl)[0])
            #update tab change tab_url to new_url
            WriteSQL(f"""update tabs set taburl = '{new_url}' where formname = '{self.title}' and tab = '{curtab}'""")
            
    def NewTabFolder(self):
            curtab = self.SM_Tabs.currentIndex()
            curtab = str(self.SM_Tabs.tabText(curtab))
            #select taburl from tabs where formname and tab
            taburl = ReadSQL("select taburl from tabs where formname = '"+self.title+"' and tab = '"+curtab+"'")
            if len(taburl)>0:
                taburl = taburl[0][0]
            else:
                taburl = ""
            new_folder = str(QFileDialog.getExistingDirectory(None,"Select Folder",taburl))
            #update tab change tab_url to new_url
            WriteSQL(f"""update tabs set treepath = '{new_folder}' where formname = '{self.title}' and tab = '{curtab}'""")
            self.Refresh()
            #update tab change tab_url to new_url
            

    def OpenTabUrl(self):
        curtab = self.SM_Tabs.currentIndex()
        curtab = str(self.SM_Tabs.tabText(curtab))
        #select taburl from tabs where formname = '"+self.title+"' and tab = '"+self.Main_CB_Tabs.currentText()+"'")
        print("select taburl from tabs where formname = '"+self.title+"' and tab = '"+curtab+"'")
        taburl = ReadSQL("select taburl from tabs where formname = '"+self.title+"' and tab = '"+curtab+"'")
        if taburl[0][0]:
            taburl = taburl[0][0]
            webbrowser.open(taburl)
        else:
            self.NewTabUrl()

    def Add_Folder(self):
        curtab = self.SM_Tabs.currentIndex()
        curtab = str(self.SM_Tabs.tabText(curtab))
        new_file = QFileDialog.getExistingDirectory(None,'Open Folder','')
        if len(new_file[0])<1:
            logger.error("No folder selected")
            return
        maxbut = ReadSQL("select max(buttonsequence) from buttons where tab = '"+curtab+"'")
        print(new_file)
        idir = os.path.dirname(new_file)
        if idir[-1:]!="\\": #this is redundant as he is using os.path.dirname.
            idir = idir+"\\"
        ibase = os.path.basename(new_file)
        logger.success(ibase + " " + idir)
        new_button = str(QInputDialog.getText(None,"New Folder Button Name","Name:",QLineEdit.Normal,"Folder " + ibase)[0])
        if not new_button:
            logger.error("ButtonName is None")
            return
        does_exist = ReadSQL("select buttonname from buttons where tab = '"+curtab+"' and buttonname = '"+new_button+"'")
        while len(does_exist)>0:
            ctypes.windll.user32.MessageBoxW(0,"Button already exists","Error",0)
            new_button = str(QInputDialog.getText(None,"New Folder Button Name","Name:",QLineEdit.Normal,"")[0])
            if new_button!=None and new_button!="":
                does_exist = ReadSQL("select buttonname from buttons where tab = '"+curtab+"' and buttonname = '"+new_button+"'")
            else:
                return
        maxbut = maxbut[0][0] + 1 if maxbut[0][0] else 1
        maxcolumn = ReadSQL("select max(columnnum) from buttons where tab = '"+curtab+"'")
        if not maxcolumn:
            maxcolumn = "null"
        elif not maxcolumn[0][0]:
            maxcolumn = "null"
        else:
            maxcolumn = "1" #putting button automatically in column 1
        formtitle = self.title
        WriteSQL(f"""insert into buttons(formname,tab,buttonname,buttonsequence,columnnum) values('{formtitle}', '{curtab}','{new_button}',{maxbut},{maxcolumn}) """)
        WriteSQL("insert into batchsequence(formname,tab,buttonname,runsequence,folderpath,filename,type) values('"+formtitle+"','"+curtab+"','"+str(new_button)+"',1,'"+idir+"','"+ibase+"','exe')")    
        self.Refresh()
    
    def Add_url(self):
        curtab = self.SM_Tabs.currentIndex()
        curtab = str(self.SM_Tabs.tabText(curtab))
        new_url = str(QInputDialog.getText(None,"New URL","URL:",QLineEdit.Normal,"")[0])
        #check for text onenote: in new_url and remove all text before it
        if new_url.find("onenote:")>-1:
            new_url = new_url[new_url.find("onenote:"):]
            print(new_url)

        new_button = str(QInputDialog.getText(None,"New URL Button Name","Name:",QLineEdit.Normal,"")[0])
        if not new_button:
            return
        maxbut = ReadSQL("select max(buttonsequence) from buttons where tab = '"+curtab+"'")
        if len(maxbut) < 1:
            logger.error("Max button is less then 1")
        maxbut = maxbut[0][0] + 1 if maxbut[0][0] else 1
        maxcolumn = ReadSQL("select max(columnnum) from buttons where tab = '"+curtab+"'")
        if not maxcolumn:
            maxcolumn = "null"
        elif not maxcolumn[0][0]:
            maxcolumn = "null"
        else:
            maxcolumn = "1"
        formtitle = self.title
        WriteSQL(f"""insert into buttons(formname,tab,buttonname,buttonsequence,columnnum) values('{formtitle}', '{curtab}','{new_button}',{maxbut},{maxcolumn}) """)

        WriteSQL(f"""insert into batchsequence(formname,tab,buttonname,runsequence,type, source) values('{formtitle}', '{curtab}','{new_button}',1,'url','{new_url}') """)
        self.Refresh()
        
    def AboutMsg_Try(self):
        try:
            self.AboutMsg()
        except Exception as e:
            logger.error("Error: "+str(e))

    def AboutMsg(self):
        ctypes.windll.user32.MessageBoxW(0,"Version: "+version,"About",0)


    def RefreshHandler(self):
        self.Refresh()
        self.Refresh()


    def Refresh(self):
        tabsize = ReadSQL("select tabsize from tabs where formname = '"+
            self.title+"' order by tabsequence asc")
        self.curTab = self.SM_Tabs.currentIndex()
        for h in reversed(range(0,self.SM_Tabs.count())):
            self.SM_Tabs.removeTab(h)
        
        formtabs = ReadSQL("select tab from tabs where formname = '"+
                           self.title+"' order by tabsequence asc")
        tabgrid = ReadSQL("select grid from tabs where formname = '"+
                           self.title+"' order by tabsequence asc")
        tabsdesc = ReadSQL("select tabdesc from tabs where formname = '"+
                           self.title+"' order by tabsequence asc")
        tabsize = ReadSQL("select tabsize from tabs where formname = '"+
                          self.title+"' order by tabsequence asc")

        if len(formtabs)>0:
            for ft in range(len(formtabs)):
                if tabgrid[ft][0]==None or tabgrid[ft][0]=="":
                    tabgrid[ft][0] = 1
                self.tab = QtWidgets.QWidget()
                self.tab.setToolTip(str(tabsdesc[ft][0]))
                self.tab.setObjectName("Tab_"+str(ft+1))
                self.TabGrid = QtWidgets.QGridLayout(self.tab)
                self.TabGrid.setObjectName("TabGrid_"+str(ft+1))
                self.SM_ScrollArea = QtWidgets.QScrollArea(self.tab)
                self.SM_ScrollArea.setWidgetResizable(True)
                self.SM_ScrollArea.setObjectName("SM_ScrollArea_"+str(ft+1))
                self.SM_ScrollAreaContents = QtWidgets.QWidget()
                self.SM_ScrollAreaContents.setGeometry(QtCore.QRect(0, 0, 248, 174))
                self.SM_ScrollAreaContents.setObjectName("SM_ScrollAreaContents_"+str(ft+1))
                self.SM_ScrollGrid = QtWidgets.QGridLayout(self.SM_ScrollAreaContents)
                self.SM_ScrollGrid.setObjectName("SM_ScrollGrid_"+str(ft+1))
                self.SM_Grid = QtWidgets.QGridLayout()
                self.SM_Grid.setObjectName("SM_Grid_"+str(ft+1))
                allbuttons = ReadSQL("select buttonname from buttons where "+
                                     "formname = '"+self.title+"' and tab = '"+
                                     str(formtabs[ft][0])+"' order by buttonsequence asc")
                if len(allbuttons)>0:
                    formbuttons = ReadSQL("select buttonname from buttons where "+
                                          "formname = '"+self.title+"' and tab = '"+
                                          str(formtabs[ft][0])+"' and columnnum is null "+
                                          "order by buttonsequence asc")
                    formbuttons2 = ReadSQL("select buttonname from buttons where "+
                                           "formname = '"+self.title+"' and tab = '"+
                                           str(formtabs[ft][0])+"' and columnnum is not null "+
                                           "order by buttonsequence asc")
                    buttonsordered = []
                    for i in range(len(formbuttons)):
                        buttonsordered.append(formbuttons[i])
                    for i in range(len(formbuttons2)):
                        buttonsordered.append(formbuttons2[i])
                    buttoncol = ReadSQL("select columnnum from buttons where "+
                                        "formname = '"+self.title+"' and tab = '"+
                                        str(formtabs[ft][0])+"' and columnnum is not null "+
                                        "order by buttonsequence asc")
                    buttoncolordered = []
                    for i in range(len(formbuttons)):
                        buttoncolordered.append(None)
                    for i in range(len(formbuttons2)):
                        buttoncolordered.append(buttoncol[i])
                    buttonsdesc = ReadSQL("select buttondesc from buttons where "+
                                          "formname = '"+self.title+"' and tab = '"+
                                          str(formtabs[ft][0])+"' and columnnum is null "+
                                          "order by buttonsequence asc")
                    buttonsdesc2 = ReadSQL("select buttondesc from buttons where "+
                                           "formname = '"+self.title+"' and tab = '"+
                                           str(formtabs[ft][0])+"' and columnnum is not null "+
                                           "order by buttonsequence asc")
                    buttondescordered = []
                    for i in range(len(buttonsdesc)):
                        buttondescordered.append(buttonsdesc[i])
                    for i in range(len(buttonsdesc2)):
                        buttondescordered.append(buttonsdesc2[i])

                    accumbuttons = []
                    for bn in range(len(buttonsordered)):
                        if bn!=0:
                            self.SM_Grid.setRowStretch(bn,3)
                        self.button = QtWidgets.QPushButton(self.SM_ScrollAreaContents)
                        self.button.setToolTip(str(buttondescordered[bn][0]))
                        self.button.setObjectName("SM_Button_"+str(ft+1)+"_"+str(bn+1))
                        self.button.setText(str(buttonsordered[bn][0]))

                        if buttoncolordered[bn] in (None, [None], ['']):
                            x = 0
                            y = 0
                            if tabgrid[ft][0]==None or tabgrid[ft][0]=="":
                                tabgrid[ft][0] = 1
                                x = bn//tabgrid[ft][0]
                                y = bn%tabgrid[ft][0]
                                while len(accumbuttons)<1:
                                    accumbuttons.append(0)
                                accumbuttons[0] += 1
                            else:
                                if tabgrid[ft][0]<2:
                                    x = bn//tabgrid[ft][0]
                                    y = bn%tabgrid[ft][0]
                                    while len(accumbuttons)<1:
                                        accumbuttons.append(0)
                                    accumbuttons[0] += 1
                                else:
                                    x = bn%math.ceil(len(buttonsordered)/tabgrid[ft][0])
                                    y = bn//math.ceil(len(buttonsordered)/tabgrid[ft][0])
                                    while len(accumbuttons)<y+1:
                                        accumbuttons.append(0)
                                    accumbuttons[y] += 1
                        else:
                            if len(formbuttons2)>0:
                                if len(accumbuttons)<tabgrid[ft][0]:
                                    lenaccumbuttons = len(accumbuttons)
                                    for i in range(tabgrid[ft][0]):
                                        if i>=lenaccumbuttons:
                                            accumbuttons.append(0)
                                
                                y = buttoncolordered[bn][0]-1
                                if tabgrid[ft][0]==None or tabgrid[ft][0]<2:
                                    tabgrid[ft][0] = 1
                                if y>tabgrid[ft][0]-1:
                                    y = tabgrid[ft][0]-1
                                elif y<0:
                                    y = 0
                                x = accumbuttons[y]
                                
                                accumbuttons[y] += 1

                        self.SM_Grid.addWidget(self.button, x, y, 1, 1)
                        self.button.setStyleSheet("QPushButton { background-color: none }"
                                                  "QPushButton:hover { background-color: lightblue }"
                                                  "QPushButton:focus { background-color: tomato }" )
                        self.button.clicked.connect(partial(self.on_click_button,
                                                    self.title,
                                                    str(formtabs[ft][0]),
                                                    str(buttonsordered[bn][0]),
                                                    "SM_Button_"+str(ft+1)+"_"+str(bn+1)))
                self.SM_ScrollGrid.addLayout(self.SM_Grid, 0, 0, 1, 1)
                self.SM_ScrollArea.setWidget(self.SM_ScrollAreaContents)
                self.TabGrid.addWidget(self.SM_ScrollArea, 0, 0, 1, 1)
                self.SM_Tabs.addTab(self.tab, "")
                self.SM_Tabs.setTabText(self.SM_Tabs.indexOf(self.tab),formtabs[ft][0])
        if self.curTab<0 or self.SM_Tabs.count()<=self.curTab:
            self.SM_Tabs.setCurrentIndex(0)
        else:
            self.SM_Tabs.setCurrentIndex(self.curTab)

        self.SM_Tabs.currentChanged.connect(partial(self.get_tab_change,tabsize))
        logger.success("Refreshed code")
    def get_tab_change(self,n):
        if len(n)>0 and self.SM_Tabs.currentIndex() < len(n):
            n = n[self.SM_Tabs.currentIndex()]
            n = n[0]
            if n!=None and n!="":
                n = n.split(',')
                if len(n)>1:
                    try:
                        n[0] = int(n[0])
                        n[1] = int(n[1])
                        self.resize(n[0],n[1])
                    except Exception as e:
                        logger.error(str(e))
            else:
                self.resize(650,300)
        else:
            self.resize(650,300)
    
    #Main: Change Tab ComboBox
    def on_tab_change(self):
        n = str(self.Main_CB_Tabs.currentText())
        logger.debug(n)
        buttonlist = ReadSQL("select buttonname from buttons where formname = '"+
                             self.title+"' and tab = '"+n+"'")
        self.Main_CB_Buttons.clear()
        self.Main_CB_Buttons.addItem("")
        if len(buttonlist)>0:
            for bl in range(len(buttonlist)):
                self.Main_CB_Buttons.addItem(str(buttonlist[bl][0]))

    #Main: Change Button ComboBox
    def on_button_change(self):
        n = str(self.Main_CB_Buttons.currentText())
        if len(n)>0:
            bh = ReadSQL("pragma table_info('batchsequence')")
            bs = ReadSQL("select runsequence,folderpath,filename,type,source,"+
                         "target,databasepath,databasename,keypath,"+
                         "keyfile from batchsequence where formname = '"+self.title+
                         "' and tab = '"+str(self.Main_CB_Tabs.currentText())+
                         "' and buttonname = '"+n+"'")
            header = []
            for h in range(3,len(bh)):
                header.append(bh[h][1])
            self.Main_Table.setColumnCount(len(bs[0]))
            self.Main_Table.setRowCount(len(bs))
            self.Main_Table.setHorizontalHeaderLabels(header)
            for i in range(len(bs)):
                for j in range(len(bs[0])):
                    self.Main_Table.setItem(i,j,QTableWidgetItem(str(bs[i][j])))
        else:
            bh = ReadSQL("pragma table_info('batchsequence')")
            if len(bh)>0:
                header = []
                for h in range(3,len(bh)):
                    header.append(bh[h][1])
            self.Main_Table.setColumnCount(0)
            self.Main_Table.setRowCount(0)
            for rr in range(self.Main_Table.rowCount()):
                self.Main_Table.removeRow(rr)

    #Main: Insert Record
    def on_InsertRec(self):
        if len(str(self.Main_CB_Tabs.currentText()))>0 and len(str(self.Main_CB_Buttons.currentText()))>0:
            tname = str(self.Main_CB_Tabs.currentText())
            bname = str(self.Main_CB_Buttons.currentText())
            bh = ReadSQL("pragma table_info('batchsequence')")
            bs = ReadSQL("select runsequence,folderpath,filename,type,source,target,"+
                         "databasepath,databasename,keypath,keyfile "+
                         "from batchsequence where formname = '"+self.title+"' and tab = '"+
                         tname+"' and buttonname = '"+bname+"'")
            bt = str(bs[0][3])
            if bt=="bat":
                self.files = QFileDialog.getOpenFileName(None,'Open file','',"BAT Files (*.bat)")
                if self.files!='':
                    temp = str(self.files[0])
                    idir = os.path.dirname(temp)
                    idir = idir.replace("/","\\")
                    ibase = os.path.basename(temp)
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+
                                  self.title+"' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,"+
                             "folderpath,filename,type) values('"+self.title+"','"+tname+"','"+
                             bname+"',"+str(int(cnt[0][0])+1)+",'"+idir+"\\','"+ibase+"','bat')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()
            if bt=="sql":
                self.files = QFileDialog.getOpenFileName(None,'Choose SQL File','',"SQL Files (*.sql)")
                if self.files!='':
                    temp1 = str(self.files[0])
                    idir1 = os.path.dirname(temp1)
                    idir1 = idir1.replace("/","\\")
                    ibase1 = os.path.basename(temp1)
                    self.dbs = QFileDialog.getOpenFileName(None,'Choose Database File','',"DB Files (*.db)")
                    if self.dbs!='':
                        temp2 = str(self.dbs[0])
                        idir2 = os.path.dirname(temp2)
                        idir2 = idir2.replace("/","\\")
                        ibase2 = os.path.basename(temp2)
                        WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,folderpath,"+
                                 "filename,type,databasepath,databasename) values('"+self.title+
                                 "','"+tname+"','"+bname+"',"+str(int(cnt[0][0])+1)+",'"+idir1+"\\','"+ibase1+
                                 "','sql','"+idir2+"\\','"+ibase2+"')")
                        self.Main_Table.deleteLater()
                        self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                        self.Main_Table.setObjectName("tableMain")
                        self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                        self.on_button_change()
            if bt=="srtg":
                self.files = QFileDialog.getOpenFileName(None,'Open file','',"BAT Files (*.bat)")
                if self.files!='':
                    temp = str(self.files[0])
                    idir1 = os.path.dirname(temp)
                    idir1 = idir1.replace("/","\\")
                    ibase1 = os.path.basename(temp)
                    self.src = str(QFileDialog.getExistingDirectory(None, "Select Source Path"))
                    if self.src!='':
                        self.src = self.src.replace("/","\\")
                        self.trg = str(QFileDialog.getExistingDirectory(None, "Select Target Path"))
                        if self.trg!='':
                            self.trg = self.trg.replace("/","\\")
                            cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                          "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                            WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,folderpath,"+
                                     "filename,type,source,target) values('"+self.title+"','"+tname+"','"+bname+"',"+
                                     str(int(cnt[0][0])+1)+",'"+idir1+"\\','"+ibase1+"','srtg','"+str(self.src)+
                                     "','"+str(self.trg)+"')")
                            self.Main_Table.deleteLater()
                            self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                            self.Main_Table.setObjectName("tableMain")
                            self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                            self.on_button_change()
            if bt=="RARa":
                self.src = str(QFileDialog.getExistingDirectory(None, "Select Source Path"))
                if self.src!='':
                    self.src = self.src.replace("/","\\")
                    self.trg = str(QFileDialog.getExistingDirectory(None, "Select Target Path"))
                    if self.trg!='':
                        self.trg = self.trg.replace("/","\\")
                        self.filename = str(QInputDialog.getText(None,"Add RAR target","Name:",QLineEdit.Normal,"")[0])+".rar"
                        if self.filename!='':
                            self.files = QFileDialog.getOpenFileName(None,'Choose key file','',"TXT Files (*.txt)")
                            if len(self.files[0])>0:
                                temp = str(self.files[0])
                                idir = os.path.dirname(temp)
                                idir = idir.replace("/","\\")
                                ibase = os.path.basename(temp)
                                cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                              "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                                WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,filename,"+
                                         "type,source,target,keypath,keyfile) values('"+self.title+"','"+tname+"','"+
                                         bname+"',"+str(int(cnt[0][0])+1)+",'"+str(self.filename)+"','RARa','"+
                                         str(self.src)+"','"+str(self.trg)+"','"+idir+"\\','"+ibase+"')")
                            else:
                                cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                              "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                                WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,filename,"+
                                         "type,source,target,keypath,keyfile) values('"+self.title+"','"+tname+"','"+bname+
                                         "',"+str(int(cnt[0][0])+1)+",'"+str(self.filename)+"','RARa','"+
                                         str(self.src)+"','"+str(self.trg)+"','"+winshell.desktop()+"\\','key.txt')")
                            self.Main_Table.deleteLater()
                            self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                            self.Main_Table.setObjectName("tableMain")
                            self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                            self.on_button_change()
            if bt=="RARe":
                self.files = QFileDialog.getOpenFileName(None,'Choose RAR File','',"RAR Files (*.rar)")
                if self.files!='':
                    temp = str(self.files[0])
                    idir1 = os.path.dirname(temp)
                    idir1 = idir1.replace("/","\\")
                    ibase1 = os.path.basename(temp)
                    self.trg = str(QFileDialog.getExistingDirectory(None, "Select Target Path"))
                    if self.trg!='':
                        self.trg = self.trg.replace("/","\\")
                        self.keyfiles = QFileDialog.getOpenFileName(None,'Choose key file','',"TXT Files (*.txt)")
                        if len(self.keyfiles[0])>0:
                            temp2 = str(self.keyfiles[0])
                            idir2 = os.path.dirname(temp2)
                            idir2 = idir2.replace("/","\\")
                            ibase2 = os.path.basename(temp2)
                            cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                          "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                            WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,filename,"+
                                     "type,source,target,keypath,keyfile) values('"+self.title+"','"+tname+"','"+
                                     bname+"',"+str(int(cnt[0][0])+1)+",'"+ibase1+"','RARe','"+idir1+"','"+
                                     str(self.trg)+"','"+idir2+"\\','"+ibase2+"')")
                        else:
                            cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                          "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                            WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,filename,"+
                                     "type,source,target,keypath,keyfile) values('"+self.title+"','"+tname+"','"+
                                     bname+"',"+str(int(cnt[0][0])+1)+",'"+ibase1+"','RARe','"+idir1+"','"+
                                     str(self.trg)+"','"+winshell.desktop()+"\\','key.txt')")
                        self.Main_Table.deleteLater()
                        self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                        self.Main_Table.setObjectName("tableMain")
                        self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                        self.on_button_change()
            if bt=="mtini":
                self.folders = str(QFileDialog.getExistingDirectory(None, "Select MT4 Instance Path"))
                if self.folders!='':
                    self.folders = self.folders.replace("/","\\")
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                  "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,folderpath,"+
                             "type) values('"+self.title+"','"+tname+"','"+bname+"',"+str(int(cnt[0][0])+1)+
                             ",'"+str(self.folders)+"\\','mtini')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()
            if bt=="mtcls":
                self.folders = str(QFileDialog.getExistingDirectory(None, "Select MT4 Instance Path"))
                if self.folders!='':
                    self.folders = self.folders.replace("/","\\")
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                  "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,folderpath,"+
                             "type) values('"+self.title+"','"+tname+"','"+bname+"',"+str(int(cnt[0][0])+1)+
                             ",'"+str(self.folders)+"\\','mtcls')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()
            if bt=="exe":
                self.files = QFileDialog.getOpenFileName(None,'Open file','',"EXE Files (*.exe)")
                if self.files!='':
                    temp = str(self.files[0])
                    idir = os.path.dirname(temp)
                    idir = idir.replace("/","\\")
                    ibase = os.path.basename(temp)
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+
                                  self.title+"' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,"+
                             "folderpath,filename,type) values('"+self.title+"','"+tname+"','"+
                             bname+"',"+str(int(cnt[0][0])+1)+",'"+idir+"\\','"+ibase+"','exe')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()
            if bt=="cli_exe": ##type works for exes that run through command line. prevents main application being overridden
                self.files = QFileDialog.getOpenFileName(None,'Open file','',"EXE Files (*.exe)")
                if self.files!='':
                    temp = str(self.files[0])
                    idir = os.path.dirname(temp)
                    idir = idir.replace("/","\\")
                    ibase = os.path.basename(temp)
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+
                                  self.title+"' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,"+
                             "folderpath,filename,type) values('"+self.title+"','"+tname+"','"+
                             bname+"',"+str(int(cnt[0][0])+1)+",'"+idir+"\\','"+ibase+"','exe')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()           
            if bt=="py":
                self.files = QFileDialog.getOpenFileName(None,'Open file','',"PY Files (*.py)")
                if self.files!='':
                    temp = str(self.files[0])
                    idir = os.path.dirname(temp)
                    idir = idir.replace("/","\\")
                    ibase = os.path.basename(temp)
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+
                                  self.title+"' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,"+
                             "folderpath,filename,type) values('"+self.title+"','"+tname+"','"+
                             bname+"',"+str(int(cnt[0][0])+1)+",'"+idir+"\\','"+ibase+"','py')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()
            if bt=="dellocdb":
                self.src = str(QFileDialog.getExistingDirectory(None, "Select Instance Path"))
                if self.src!='':
                    self.src = self.src.replace("/","\\")
                    cnt = ReadSQL("select max(runsequence) from batchsequence where formname = '"+self.title+
                                  "' and tab = '"+tname+"' and buttonname = '"+bname+"'")
                    WriteSQL("insert into batchsequence (formname,tab,buttonname,runsequence,folderpath,"+
                             "type) values('"+self.title+"','"+tname+"','"+bname+"',"+
                             str(int(cnt[0][0])+1)+",'"+str(self.src)+"\\','dellocdb')")
                    self.Main_Table.deleteLater()
                    self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
                    self.Main_Table.setObjectName("tableMain")
                    self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
                    self.on_button_change()

    #Main: Delete record
    def on_DeleteRec(self):
        if len(str(self.Main_CB_Tabs.currentText()))>0 and len(str(self.Main_CB_Buttons.currentText()))>0:
            tname = str(self.Main_CB_Tabs.currentText())
            bname = str(self.Main_CB_Buttons.currentText())
            bh = ReadSQL("pragma table_info('batchsequence')")
            bs = ReadSQL("select runsequence,folderpath,filename,type,source,target,"+
                         "databasepath,databasename,keypath,keyfile "+
                         "from batchsequence where formname = '"+self.title+"' and tab = '"+
                         tname+"' and buttonname = '"+bname+"'")
            bt = str(bs[0][3])
            for i in self.Main_Table.selectionModel().selectedRows():
                n = str(self.Main_Table.item(i.row(),0).text())
                WriteSQL("delete from batchsequence where formname = '"+self.title+"' and tab = '"+tname+
                         "' and buttonname = '"+bname+"' and runsequence = "+n+" and type = '"+bt+"'")
            self.Main_Table.deleteLater()
            self.Main_Table = QtWidgets.QTableWidget(self.scrollAreaWidgetContents_2)
            self.Main_Table.setObjectName("tableMain")
            self.gridLayout_4.addWidget(self.Main_Table, 0, 0, 1, 1)
            self.on_button_change()

    #Main: Change Cell
    def on_CellChange(self):
        if len(str(self.Main_CB_Tabs.currentText()))>0 and len(str(self.Main_CB_Buttons.currentText()))>0:
            tname = str(self.Main_CB_Tabs.currentText())
            bname = str(self.Main_CB_Buttons.currentText())
            h = ReadSQL("pragma table_info('batchsequence')")
            r = self.Main_Table.currentRow()
            c = self.Main_Table.currentColumn()+3
            if r>=0 and c>2:
                txt = self.Main_Table.item(self.Main_Table.currentRow(),self.Main_Table.currentColumn()).text()
                if txt=='':
                    txt = "null"
                    self.Main_Table.setItem(self.Main_Table.currentRow(),self.Main_Table.currentColumn(),QTableWidgetItem("None"))
                else:
                    tmp = h[c][2]
                    if tmp.find("char")>=0 or tmp.find("CHAR")>=0 or tmp.find("datetime")>=0 or tmp.find("DATETIME")>=0 or tmp.find("text")>=0 or tmp.find("TEXT")>=0:
                        txt = "'"+txt+"'"
                WriteSQL("update batchsequence set "+h[c][1]+" = "+txt+" where formname = '"+self.title+"' and tab = '"+
                         tname+"' and buttonname = '"+bname+"' and runsequence = "+str(self.Main_Table.item(r,0).text()))
    
    def on_click_button(self,pname,tname,bname,objn,mode=0): ##pname = SGX, ##
        result = 1 

        if(result==1):
            #self.SM_Grid.findChild(QPushButton,objn).setStyleSheet("background: tomato")
            bseq = ReadSQL("select folderpath,filename,type,source,target,databasepath,"+
                                   "databasename,keypath,keyfile,runsequence,treepath from batchsequence where "+
                                   "formname = '"+pname+"' and tab = '"+tname+"' and buttonname = '"+
                                   bname+"' order by runsequence asc")
            if len(bseq) == 0:
                #error popup batchsequence not found for button x

                ctypes.windll.user32.MessageBoxW(0, "Batchsequence not found for button \""+bname+"\" in tab \""+tname+"\" in form \""+pname+"\"", "Error", 0)
                return
            if self.edit_mode:
                self.edit_button(bseq,pname,tname,bname,objn,mode)
                return
            if self.cur_seq:
                self.cur_seq.add_button(bseq,pname,tname,bname,objn,mode)
                return
            for pf in range(len(bseq)):
                if str(bseq[pf][2])=="bat":
                    if os.path.exists(str(bseq[pf][0])+"\\"+str(bseq[pf][1]))==True:
                        os.system("\""+str(bseq[pf][0])+"\\"+str(bseq[pf][1])+"\"")
                    else:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" does not exist?","Failed bat: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="sql":
                    eng = ReadSQL("select * from sqlengine")
                    if len(eng)<=0:
                        ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        bseq[pf][0] = addSlash(str(bseq[pf][0]))
                        path = str(eng[0][0])+" "+str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" \".read '"+str(bseq[pf][0])+"\\\\"+str(bseq[pf][1])+"'\""
                        if os.path.exists(str(eng[0][0]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(eng[0][0])+" does not exist?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][0])+"\\"+str(bseq[pf][1]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" does not exist?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            try:
                                #os.rename(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][5])+"\\"+str(bseq[pf][6]))
                                os.system(path)
                            except OSError as e:
                                ctypes.windll.user32.MessageBoxW(0,path+" error running .sql file, permission issue? \\"+str(bseq[pf][9]),"",1)
                elif str(bseq[pf][2])=="sqlinput":
                    eng = ReadSQL("select * from sqlengine")
                    if len(eng)<=0:
                        ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        bseq[pf][0] = addSlash(str(bseq[pf][0]))
                        path = str(eng[0][0])+" "+str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" \".read '"+str(bseq[pf][0])+"\\\\"+str(bseq[pf][1])+"'\""
                        if os.path.exists(str(eng[0][0]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(eng[0][0])+" does not exist?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][0])+"\\"+str(bseq[pf][1]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" does not exist?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            variables = bseq[pf][3]
                            if variables==None or variables=="":
                                ctypes.windll.user32.MessageBoxW(0,"No variables specified in source?","Failed sql: "+bname+"! \\"+str(bseq[pf][9]),0)
                            else:
                                variables = variables.split('|')
                                txt = open(str(bseq[pf][0])+"\\"+str(bseq[pf][1])).read()
                                for i in range(len(variables)):
                                    txt = txt.replace("%var"+str(i+1)+"%",variables[i])
                                os.system("type NUL > "+str(bseq[pf][0])+"\\%%%"+str(bseq[pf][1]))
                                if os.path.exists(str(bseq[pf][0])+"\\%%%"+str(bseq[pf][1]))==True:
                                    newtxt = open(str(bseq[pf][0])+"\\%%%"+str(bseq[pf][1]),'w+')
                                    newtxt.write(txt)
                                    newtxt.close()
                                try:
                                    path = str(eng[0][0])+" "+str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" \".read '"+str(bseq[pf][0])+"\\\\%%%"+str(bseq[pf][1])+"'\""
                                    os.system(path)
                                    os.remove(str(bseq[pf][0])+"\\%%%"+str(bseq[pf][1]))
                                except OSError as e:
                                    ctypes.windll.user32.MessageBoxW(0,path+" error running .sql file, permission issue? \\"+str(bseq[pf][9]),"",1)
                elif str(bseq[pf][2])=="pg":
                    if str(bseq[pf][4])=="" or str(bseq[pf][4])=="None":
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials provided?","Failed pg: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(str(bseq[pf][0])+"\\"+str(bseq[pf][1]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" does not exist?","Failed pg: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        try:
                            self.PG_Function(str(bseq[pf][0])+"\\"+str(bseq[pf][1]),str(bseq[pf][4]),"",str(bseq[pf][9]),0)
                        except OSError as e:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" error running .sql file, permission issue?"," \\"+str(bseq[pf][9]),1)
                # Run a faux .sql file that allows for customizable inputs
                elif str(bseq[pf][2])=="pginput":
                    if str(bseq[pf][4])=="" or str(bseq[pf][4])=="None":
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials provided?","Failed pginput: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(str(bseq[pf][0])+"\\"+str(bseq[pf][1]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" does not exist?","Failed pginput: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.PG_Function(str(bseq[pf][0])+"\\"+str(bseq[pf][1]),str(bseq[pf][4]),str(bseq[pf][3]),str(bseq[pf][9]),1)
                elif str(bseq[pf][2])=="srtg":
                    if os.path.exists(str(bseq[pf][0])+"\\$Srtg.bat")==True:
                        os.system("rm "+str(bseq[pf][0])+"\\$Srtg.bat")
                    if os.path.exists(str(bseq[pf][0])+"\\"+str(bseq[pf][1]))==True:
                        txt = open(str(bseq[pf][0])+"\\"+str(bseq[pf][1])).read()
                        txt = txt.replace("%Source%",str(bseq[pf][3]))
                        txt = txt.replace("%Target%",str(bseq[pf][4]))
                        os.system("type NUL > "+str(bseq[pf][0])+"\\$Srtg.bat")
                        if os.path.exists(str(bseq[pf][0])+"\\$Srtg.bat")==True:
                            newtxt = open(str(bseq[pf][0])+"\\$Srtg.bat",'w')
                            newtxt.write(txt)
                            newtxt.close()
                            os.system(str(bseq[pf][0])+"\\$Srtg.bat")
                        else:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\$Srtg.bat did not create?","Failed srtg: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1])+" does not exist?","Failed srtg: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="RARa":
                    key = ''
                    if os.path.exists(str(bseq[pf][7])+"\\"+str(bseq[pf][8]))==True:
                        key = open(str(bseq[pf][7])+"\\"+str(bseq[pf][8])).read()
                        if key=='':
                            ctypes.windll.user32.MessageBoxW(0,"Key not provided within "+str(bseq[pf][7])+"\\"+str(bseq[pf][8])+".","Failed RARa: "+bname+"! \\"+str(bseq[pf][9]),1)
                        else:
                            src = str(bseq[pf][3])
                            trg = str(bseq[pf][4])
                            if os.path.exists(trg)==False:
                                os.system("mkdir \""+trg+"\"")
                            if os.path.exists(src)==False:
                                if src==None or src=="None":
                                    src = "source"
                                ctypes.windll.user32.MessageBoxW(0,src+" does not exist?","Failed RARa: "+bname+"! \\"+str(bseq[pf][9]),0)
                            elif os.path.exists(trg)==False:
                                if trg==None or trg=="None":
                                    trg = "target"
                                ctypes.windll.user32.MessageBoxW(0,trg+" does not exist?","Failed RARa: "+bname+"! \\"+str(bseq[pf][9]),0)
                            else:
                                ibase = os.path.basename(src)
                                shutil.copytree(src,"C:\\"+ibase)
                                os.system("WinRAR a -u -r -p"+key+" \""+str(bseq[pf][4])+"\\"+str(bseq[pf][1])+"\" \"C:\\"+ibase+"\*.*\"")
                                shutil.rmtree("C:\\"+ibase)
                    else:
                        keypath = bseq[pf][7]
                        keyfile = bseq[pf][8]
                        if keypath==None or keypath=="None":
                            keypath = "keypath "
                        if keyfile==None or keyfile=="None":
                            keyfile = "keyfile"
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][7])+"\\"+str(bseq[pf][8])+" does not exist?","Failed RARa: "+bname+"! \\"+str(bseq[pf][9]),0)
                if str(bseq[pf][2])=="RARe":
                    key = ''
                    if os.path.exists(str(bseq[pf][7])+"\\"+str(bseq[pf][8]))==True:
                        key = open(str(bseq[pf][7])+"\\"+str(bseq[pf][8])).read()
                        if key=='':
                            ctypes.windll.user32.MessageBoxW(0,"Key not provided within "+str(bseq[pf][7])+"\\"+str(bseq[pf][8])+".","Failed RARa: "+bname+"! \\"+str(bseq[pf][9]),1)
                        else:
                            val = str(bseq[pf][3])+"\\"+str(bseq[pf][1])
                            trg = str(bseq[pf][4])
                            if os.path.exists(trg)==False:
                                if trg==None or trg=="None":
                                    trg = "target"
                                os.system("mkdir \""+trg+"\"")
                            if os.path.exists(trg)==False:
                                if trg==None or trg=="None":
                                    trg = "target"
                                ctypes.windll.user32.MessageBoxW(0,trg+" does not exist?","Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                            if os.path.exists(val)==True:
                                temp_dir = os.path.dirname(trg)+"\\"+os.path.basename(trg)+"\\$$temp$$"
                                if os.path.exists(temp_dir)==True:
                                    try:
                                        shutil.rmtree(temp_dir)
                                    except:
                                        ctypes.windll.user32.MessageBoxW(0,"Failed to delete "+temp_dir+", permission issue?","Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                                if os.path.exists(temp_dir)==False:
                                    try:
                                        os.system("mkdir \""+temp_dir+"\"")
                                    except Exception as e:
                                        ctypes.windll.user32.MessageBoxW(0,"Failed to recreate "+temp_dir+", "+str(e),"Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                                try:
                                    os.system("WinRAR x -ep1 -r -o+ \""+val+"\" \""+temp_dir+"\" -p"+key)
                                except Exception as e:
                                    ctypes.windll.user32.MessageBoxW(0,"Failed to run WinRAR x -ep1 -r -o+ \""+val+"\" \""+temp_dir+"\" -p{password}, "+str(e),"Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                                #os.system("WinRAR x -ep1 -r -o+ \""+val+"\" \""+trg+"\" -p"+key)
                                for d in os.listdir(temp_dir):
                                    try:
                                        copyDir(os.path.join(temp_dir,str(d)),trg)##########
                                    except Exception as e:
                                        ctypes.windll.user32.MessageBoxW(0,"Failed to copy "+os.path.join(temp_dir,str(d))+" to "+trg+", "+str(e),"Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                                try:
                                    shutil.rmtree(temp_dir)
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Failed to delete "+temp_dir+", permission issue?","Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                            else:
                                source = str(bseq[pf][3])
                                rarfile = str(bseq[pf][1])
                                val = ""
                                if source==None or source=="None":
                                    source = "source "
                                    val = source
                                if rarfile==None or rarfile=="None":
                                    val += "filename"
                                ctypes.windll.user32.MessageBoxW(0,val+" does not exist?","Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        keypath = str(bseq[pf][7])
                        keyfile = str(bseq[pf][8])
                        if keypath==None or keypath=="None":
                            keypath = "keypath "
                        if keyfile==None or keyfile=="None":
                            keyfile = "keyfile"
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][7])+"\\"+str(bseq[pf][8])+" does not exist?","Failed RARe: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="RARa1":
                    if os.path.exists(str(bseq[pf][7])+"\\"+str(bseq[pf][8]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][7])+"\\"+str(bseq[pf][8])+" does not exist?","Failed RARa1: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        key = open(str(bseq[pf][7])+"\\"+str(bseq[pf][8])).read()
                        if key=='':
                            ctypes.windll.user32.MessageBoxW(0,"Key not provided within "+str(bseq[pf][7])+"\\"+str(bseq[pf][8])+"?","Failed RARa1: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            key = str(key)
                            ibase = os.path.basename(str(bseq[pf][3]))
                            if os.path.exists(str(bseq[pf][3]))==False:
                                ctypes.windll.user32.MessageBoxW(0,"Source "+str(bseq[pf][3])+" does not exist?","Failed RARa1: "+bname+"! \\"+str(bseq[pf][9]),0)
                            elif os.path.exists(str(bseq[pf][4]))==False:
                                ctypes.windll.user32.MessageBoxW(0,"Target "+str(bseq[pf][4])+" does not exist?","Failed RARa1: "+bname+"! \\"+str(bseq[pf][9]),0)
                            else:
                                os.system("WinRAR a -ep1 -u -r -p"+key+" \""+str(bseq[pf][4])+"\\"+str(bseq[pf][1])+"\" \""+str(bseq[pf][3])+"\*.*\"")
                                for root, dirs, files in os.walk(str(bseq[pf][3])):
                                    for f in files:
                                        os.remove(str(bseq[pf][3])+"\\"+f)
                                    for d in dirs:
                                        shutil.rmtree(str(bseq[pf][3])+"\\"+d)
                elif str(bseq[pf][2])=="RARe1":
                    if os.path.exists(str(bseq[pf][7])+"\\"+str(bseq[pf][8]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][7])+"\\"+str(bseq[pf][8])+" does not exist?","Failed RARe1: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        key = open(str(bseq[pf][7])+"\\"+str(bseq[pf][8])).read()
                        if key=='':
                            ctypes.windll.user32.MessageBoxW(0,"Key not provided within "+str(bseq[pf][7])+"\\"+str(bseq[pf][8])+"?","Failed RARe1: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            key = str(key)
                            if os.path.exists(str(bseq[pf][4]))==False:
                                os.system("mkdir \""+str(bseq[pf][4])+"\"")
                            if os.path.exists(str(bseq[pf][3])+"\\"+str(bseq[pf][1]))==True:
                                temp_dir = str(bseq[pf][4])+"\\$$temp$$"
                                if os.path.exists(temp_dir)==False:
                                    os.system("mkdir \""+temp_dir+"\"")
                                    os.system("WinRAR x -ep1 -r -o+ \""+str(bseq[pf][3])+"\\"+str(bseq[pf][1])+"\" \""+temp_dir+"\" -p"+key)
                                    for d in os.listdir(temp_dir):
                                        copyDir(os.path.join(temp_dir,str(d)),str(bseq[pf][4])+"\\"+str(d))
                                    try:
                                        shutil.rmtree(temp_dir)
                                    except:
                                        ctypes.windll.user32.MessageBoxW(0,"Failed to delete "+temp_dir+", permission issue?","Failed RARe1: "+bname+"! \\"+str(bseq[pf][9]),0)
                                    try:
                                        os.remove(str(bseq[pf][3])+"\\"+str(bseq[pf][1]))
                                    except OSError as e:
                                        ctypes.windll.user32.MessageBoxW(0,"Failed to delete "+str(bseq[pf][3])+"\\"+str(bseq[pf][1])+", permission issue?","Failed RARe1: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="rdstini":
                    if os.path.exists(bseq[pf][4])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" target does not exist?","Failed rdstini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        try:
                            shutil.rmtree(bseq[pf][4])
                        except:
                            ctypes.windll.user32.MessageBoxW(0,"Problem removing "+bseq[pf][4]+"?","Failed rdstini: "+bname+"! \\"+str(bseq[pf][9]),0)
                        spath = os.path.realpath(sys.argv[0])
                        spath = os.path.dirname(spath)+"\\"
                        spath = addSlash(spath)
                        if bseq[pf][4]!=None and bseq[pf][4]!="":
                            WriteSQL("attach database '"+spath+"PICAT_SM_connect.db' as PCONNECT")
                            WriteSQL("delete from PCONNECT.buttons where treepath = '"+bseq[pf][4]+"'")
                            WriteSQL("update buttons set active = null where treepath = '"+bseq[pf][4]+"'")
                            WriteSQL("detach database PCONNECT")
                elif str(bseq[pf][2])=="stini":
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed stini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.Open_Ini_Dialog("MT4",pname,tname,bname,str(bseq[pf][0]),str(bseq[pf][9]),str(bseq[pf][4]),0)
                elif str(bseq[pf][2])=="stinimt5":
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed stini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.Open_Ini_Dialog("MT5",pname,tname,bname,str(bseq[pf][0]),str(bseq[pf][9]),str(bseq[pf][4]),0)
                # Opens an MT4 instance
                elif str(bseq[pf][2])=="mtini":
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed mtini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.Open_Ini_Dialog("MT4",pname,tname,bname,str(bseq[pf][0]),str(bseq[pf][9]),str(bseq[pf][4]),1)
                elif str(bseq[pf][2])=="mtinimt5":
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed mtini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.Open_Ini_Dialog("MT5",pname,tname,bname,str(bseq[pf][0]),str(bseq[pf][9]),str(bseq[pf][4]),1)
                # Creates a preset MT4 profile for initial two charts from an existing profile
                elif str(bseq[pf][2])=="prochr":
                    if bseq[pf][3]==None or bseq[pf][3]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Source not specified?","Failed prochr: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        dbyes = ""
                        if type(bseq[pf][5])!=None and type(bseq[pf][6])!=None:
                            dbyes = str(bseq[pf][5])+"\\"+str(bseq[pf][6])
                        self.Open_Ini_Dialog("MT4",pname,tname,bname,str(bseq[pf][3]),str(bseq[pf][9]),dbyes,3,str(bseq[pf][1]),str(bseq[pf][0]))
                elif str(bseq[pf][2])=="prochrmt5":
                    if bseq[pf][3]==None or bseq[pf][3]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Source not specified?","Failed prochr: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        dbyes = ""
                        if type(bseq[pf][5])!=None and type(bseq[pf][6])!=None:
                            dbyes = str(bseq[pf][5])+"\\"+str(bseq[pf][6])
                        self.Open_Ini_Dialog("MT5",pname,tname,bname,str(bseq[pf][3]),str(bseq[pf][9]),dbyes,3,str(bseq[pf][1]),str(bseq[pf][0]))
                # Updates another MT4 profile with the preset profile generated by prochr
                elif str(bseq[pf][2])=="proini":
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed profileini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.Open_Ini_Dialog("MT4",pname,tname,bname,str(bseq[pf][3]),str(bseq[pf][9]),str(bseq[pf][4]),4,str(bseq[pf][1]),str(bseq[pf][0]))
                elif str(bseq[pf][2])=="proinimt5":
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed profileini: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.Open_Ini_Dialog("MT5",pname,tname,bname,str(bseq[pf][3]),str(bseq[pf][9]),str(bseq[pf][4]),4,str(bseq[pf][1]),str(bseq[pf][0]))
                elif str(bseq[pf][2])=="mtfinish" or str(bseq[pf][2])=="mt5finish":
                    mt = "MQL4"
                    if str(bseq[pf][2])=="mt5finish":
                        mt = "MQL5"
                    if bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"Target not specified?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            path = str(bseq[pf][5])+"\\"+str(bseq[pf][6])
                            temp_path = path
                            if temp_path[:1]=="\\":
                                temp_path = temp_path[1:]
                            attach_path = temp_path.replace("\\","\\\\")
                            temp00 = str(bseq[pf][4])
                            if temp00[-1:]=="\\":
                                temp00 = temp00[:-1]
                            if os.path.exists(temp00)==False:
                                ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                            else:
                                #print("attach database '"+attach_path+"' as COMBOT")
                                WriteSQL("attach database '"+attach_path+"' as COMBOT")
                                #os.system("pause")
                                try:
                                    get_data = ReadSQL("select brokerid,acctnum from COMBOT.brokeraccountserver where running = 1 and serverpath = '"+temp00+"'")
                                    if len(get_data)<=0:
                                        ctypes.windll.user32.MessageBoxW(0,"Could not find record in brokeraccountserver for "+temp00+"?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                                    elif get_data[0][0]==None or get_data[0][0]=="":
                                        ctypes.windll.user32.MessageBoxW(0,"Invalid brokerid in brokeraccountserver for "+temp00+"?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                                    elif get_data[0][1]==None or get_data[0][1]=="" or get_data[0][1]==0:
                                        ctypes.windll.user32.MessageBoxW(0,"Invalid acctnum in brokeraccountserver for "+temp00+"?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                                    
                                    else:
                                        #print("\""+temp00+"\\"+mt+"\\Files\\PICAT_Close.exe\" "+get_data[0][0]+" "+str(get_data[0][1])+" \""+temp00+"\" \""+path+"\"")
                                        win32api.WinExec("\""+temp00+"\\"+mt+"\\Files\\PICAT_Close.exe\" "+get_data[0][0]+" "+str(get_data[0][1])+" \""+temp00+"\" \""+path+"\"")
                                    #os.system("pause")
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Problem running \""+temp00+"\\"+mt+"\\Files\\PICAT_Close.exe\"?","Failed mtfinish: "+bname+"! \\"+str(bseq[pf][9]),0)
                                #print("detach database COMBOT")
                                WriteSQL("detach database COMBOT")
                                #os.system("pause")
                elif str(bseq[pf][2])=="mtcls":
                    temp00 = str(bseq[pf][0])
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed mtcls: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists("C:\\Windows\\nircmd.exe")==False:
                        ctypes.windll.user32.MessageBoxW(0,"C:\\Windows\\nircmd.exe does not exist?","Failed mtcls: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        cmd = "C:\\Windows\\nircmd.exe closeprocess "+str(bseq[pf][0])+"\\terminal.exe"
                        os.system(cmd)
                elif str(bseq[pf][2])=="mt5cls":
                    temp00 = str(bseq[pf][0])
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed mt5cls: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists("C:\\Windows\\nircmd.exe")==False:
                        ctypes.windll.user32.MessageBoxW(0,"C:\\Windows\\nircmd.exe does not exist?","Failed mt5cls: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        cmd = "C:\\Windows\\nircmd.exe closeprocess "+str(bseq[pf][0])+"\\terminal64.exe"
                        os.system(cmd)
                elif str(bseq[pf][2])=="mtcla":
                    if os.path.exists("C:\\Windows\\nircmd.exe")==False:
                        ctypes.windll.user32.MessageBoxW(0,"C:\\Windows\\nircmd.exe does not exist?","Failed mtcla: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        for pid in (process.pid for process in psutil.process_iter() if process.name()=="terminal.exe"):
                            os.system("C:\\Windows\\nircmd.exe closeprocess "+psutil.Process(pid).cmdline()[0])
                elif str(bseq[pf][2])=="mt5cla":
                    if os.path.exists("C:\\Windows\\nircmd.exe")==False:
                        ctypes.windll.user32.MessageBoxW(0,"C:\\Windows\\nircmd.exe does not exist?","Failed mt5cla: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        for pid in (process.pid for process in psutil.process_iter() if process.name()=="terminal64.exe"):
                            os.system("C:\\Windows\\nircmd.exe closeprocess "+psutil.Process(pid).cmdline()[0])
                elif str(bseq[pf][2])=="shcutport" or str(bseq[pf][2])=="shcutportmt5":
                    target = ""
                    if str(bseq[pf][2])=="shcutportmt5":
                        target = "64"
                    temp00 = bseq[pf][0]
                    if os.path.exists(bseq[pf][0]+"\\terminal"+target+".exe")==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][0]+"\\terminal"+target+".exe does not exist?","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(bseq[pf][0]+"\\terminal"+target+".exe")==True:
                            target = bseq[pf][0]+"\\terminal"+target+".exe"
                        argue = ""
                        if str(bseq[pf][2])=="shcutport":
                            argue = "/portable"
                        name = os.path.basename(temp00)
                        desktop = winshell.desktop()
                        path = os.path.join(desktop,name+".lnk")
                        wDir = temp00
                        icon = bseq[pf][0]+"\\terminal.ico"
                        shell = Dispatch('WScript.Shell')
                        shortcut = shell.CreateShortCut(path)
                        shortcut.Targetpath = target
                        shortcut.WorkingDirectory = wDir
                        shortcut.Arguments = argue
                        shortcut.IconLocation = icon
                        shortcut.save()
                elif str(bseq[pf][2])=="shcutportstart" or str(bseq[pf][2])=="shcutportstartmt5":
                    target = ""
                    if str(bseq[pf][2])=="shcutportstartmt5":
                        target = "64"
                    temp00 = bseq[pf][0]
                    if os.path.exists(bseq[pf][0]+"\\terminal"+target+".exe")==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][0]+"\\terminal"+target+".exe does not exist?","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(bseq[pf][0]+"\\terminal"+target+".exe")==True:
                            target = bseq[pf][0]+"\\terminal"+target+".exe"
                        argue = "/portable config\start.ini"
                        name = os.path.basename(temp00)
                        desktop = winshell.desktop()
                        path = os.path.join(desktop,name+".lnk")
                        wDir = temp00
                        icon = bseq[pf][0]+"terminal.ico"
                        shell = Dispatch('WScript.Shell')
                        shortcut = shell.CreateShortCut(path)
                        shortcut.Targetpath = target
                        shortcut.WorkingDirectory = wDir
                        shortcut.Arguments = argue
                        shortcut.IconLocation = icon
                        shortcut.save()
                elif str(bseq[pf][2])=="shcutportcpu0" or str(bseq[pf][2])=="shcutportcpu0start" or str(bseq[pf][2])=="shcutportcpu0mt5" or str(bseq[pf][2])=="shcutportcpu0startmt5":
                    target = ""
                    if str(bseq[pf][2])=="shcutportcpu0mt5" or str(bseq[pf][2])=="shcutportcpu0startmt5":
                        target = "64"
                    temp00 = bseq[pf][0]
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(bseq[pf][0]+"\\terminal"+target+".exe")==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][0]+"\\terminal"+target+".exe does not exist?","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            name = os.path.basename(temp00)
                            if os.path.exists(bseq[pf][0]+"\\"+name+".bat")==True:
                                os.remove(bseq[pf][0]+"\\"+name+".bat")
                            f = open(bseq[pf][0]+"\\"+name+".bat","w+")
                            append = ""
                            if str(bseq[pf][2])=="shcutportcpu0start":
                                append = " config\start.ini"
                            f.write("C:\Windows\system32\cmd.exe /c start \"Process Monitor\" /affinity 1 \""+bseq[pf][0]+"\\terminal"+target+".exe\""+append+" /portable")
                            f.close()
                            desktop = winshell.desktop()
                            path = os.path.join(desktop,name+".lnk")
                            target = bseq[pf][0]+"\\"+name+".bat"
                            wDir = temp00
                            argue = "/portable"
                            icon = bseq[pf][0]+"\\terminal.ico"
                            shell = Dispatch('WScript.Shell')
                            shortcut = shell.CreateShortCut(path)
                            shortcut.Targetpath = target
                            shortcut.WorkingDirectory = wDir
                            shortcut.Arguments = argue
                            shortcut.IconLocation = icon
                            shortcut.save()
                elif str(bseq[pf][2])=="pexce":
                    eng = ReadSQL("select * from sqlengine")
                    if len(eng)<=0:
                        ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed pexce: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        temp00 = str(bseq[pf][0])
                        if os.path.exists(temp00)==False:
                            ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed pexce: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed pexce: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            runExcel(1,str(bseq[pf][0])+"\\"+str(bseq[pf][1]),str(eng[0][0]),str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][0]),str(bseq[pf][3]),0,1)
                elif str(bseq[pf][2])=="pexim":
                    eng = ReadSQL("select * from sqlengine")
                    if len(eng)<=0:
                        ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed pexim: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        temp00 = str(bseq[pf][0])
                        if os.path.exists(temp00)==False:
                            ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed pexim: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(eng[0][0]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(eng[0][0])+" does not exist?","Failed pexim: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed pexim: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            runExcel(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1]),str(eng[0][0]),str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][0]),str(bseq[pf][3]),0,0)
                elif str(bseq[pf][2])=="pexir":
                    eng = ReadSQL("select * from sqlengine")
                    if len(eng)<=0:
                        ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed pexir: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        temp00 = str(bseq[pf][0])
                        if os.path.exists(temp00)==False:
                            ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed pexir: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(eng[0][0]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(eng[0][0])+" does not exist?","Failed pexir: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed pexir: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            runRange(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][3]),str(bseq[pf][0])+"\\"+str(bseq[pf][1]))
                            #runExcel(0,str(bseq[pf][0])+"\\"+str(bseq[pf][1]),str(eng[0][0]),str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][0]),str(bseq[pf][3]),1,0)
                elif str(bseq[pf][2])=="pexqu":
                    eng = ReadSQL("select * from sqlengine")
                    if len(eng)<=0:
                        ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed pexqu: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        temp00 = str(bseq[pf][0])
                        if os.path.exists(temp00)==False:
                            ctypes.windll.user32.MessageBoxW(0,temp00+" folderpath does not exist?","Failed pexqu: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" does not exist?","Failed pexqu: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif os.path.exists(str(bseq[pf][3]))==False:
                            ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][3])+" does not exist?","Failed pexqu: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif bseq[pf][1]==None or bseq[pf][1]=="":
                            ctypes.windll.user32.MessageBoxW(0,"No file specified in filename?","Failed pexqu: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            queryname = os.path.basename(str(bseq[pf][3]))[:-4]
                            runQuery(temp00+"\\"+str(bseq[pf][1]),str(eng[0][0]),str(bseq[pf][3]),str(bseq[pf][5])+"\\"+str(bseq[pf][6]),queryname)
                elif str(bseq[pf][2])=="exe" or str(bseq[pf][2])=="py":
                    if bseq[pf][4]!=None and str(bseq[pf][4])!="":
                        val = str(bseq[pf][3])+"~"+str(bseq[pf][4])
                    else:
                        val = str(bseq[pf][3])
                    self.Open_App(str(bseq[pf][2]),str(bseq[pf][0]),str(bseq[pf][1]),val)
                elif str(bseq[pf][2])=="dellocdb":
                    temp00 = str(bseq[pf][0])
                    #temp00 = temp00[:-1]
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed dellocdb: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        pathinst = str(bseq[pf][0])+"\\MQL4\\Files"
                        if os.path.exists(pathinst)==False:
                            ctypes.windll.user32.MessageBoxW(0,pathinst+" contains no MQL4\\Files directory?","Failed dellocdb: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            filelist = glob.glob(pathinst+"\\*.db")
                            for file in filelist:
                                try:
                                    if os.path.basename(file)!="Error.db":
                                        os.remove(file)
                                except OSError as e:
                                    logger.error("Failed to delete "+file+", permission issue?")
                elif str(bseq[pf][2])=="dellocdbmt5":
                    temp00 = str(bseq[pf][0])
                    #temp00 = temp00[:-1]
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed dellocdb: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        pathinst = str(bseq[pf][0])+"\\MQL5\\Files"
                        if os.path.exists(pathinst)==False:
                            ctypes.windll.user32.MessageBoxW(0,pathinst+" contains no MQL5\\Files directory?","Failed dellocdbmt5: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            filelist = glob.glob(pathinst+"\\*.db")
                            for file in filelist:
                                try:
                                    if os.path.basename(file)!="Error.db":
                                        os.remove(file)
                                except OSError as e:
                                    logger.debug("Failed to delete "+file+", permission issue?")
                elif str(bseq[pf][2])=="createdb" or str(bseq[pf][2])=="alterdb":
                    temp00 = str(bseq[pf][5])
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        copysuccess = True
                        if str(bseq[pf][2])=="alterdb":
                            if os.path.exists(str(bseq[pf][5])+"\\"+bseq[pf][6])==False:
                                #ctypes.windll.user32.MessageBoxW(0,path+bseq[pf][6]+" does not exist?","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                                    logger.warning(str(bseq[pf][5])+"\\"+bseq[pf][6]+" does not exist? Button: "+bname+"\\"+str(bseq[pf][9]))
                                    copysuccess = False
                            else:
                                try:
                                    logger.warning("copy \""+str(bseq[pf][5])+"\\"+bseq[pf][6]+"\" \""+str(bseq[pf][5])+"\\"+bseq[pf][6][0:-3]+"_bak.db\"")
                                    os.system("copy \""+str(bseq[pf][5])+"\\"+bseq[pf][6]+"\" \""+str(bseq[pf][5])+"\\"+bseq[pf][6][0:-3]+"_bak.db\"")
                                    #shutil.copy2(path+bseq[pf][6],path+bseq[pf][6][0:-3]+"_bak.db")
                                except Exception as e:
                                    logger.error(str(e))
                                    ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+bseq[pf][6]+" failed to backup.","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                                    copysuccess = False
                        if copysuccess==True:
                            statements = ReadSQL("select mastdbtbls.databasefile,mastdbtbls.tbl_name,masttbls.sql,masttbls.fields,masttbls.reqpopulate "+
                                                 "from mastdbtbls inner join masttbls on masttbls.tbl_name = mastdbtbls.tbl_name where "+
                                                 "mastdbtbls.databasefile = '"+bseq[pf][6]+"'")
                            #print(statements)
                            warning = ""
                            for i in range(len(statements)):
                                createtable = statements[i][2][0:13]
                                createtable = createtable.upper()
                                if createtable=="CREATE TABLE ":
                                    statements[i][2] = statements[i][2][0:13]+"IF NOT EXISTS NEWDB."+statements[i][2][13:]
                                path = bseq[pf][5]
                                dbpath = addSlash(path)
                                #print("attach database '"+dbpath+bseq[pf][6]+"' as NEWDB")
                                WriteSQL("attach database '"+dbpath+"\\\\"+bseq[pf][6]+"' as NEWDB")
                                WriteSQL("pragma foreign_keys = off")
                                WriteSQL("begin transaction")
                                WriteSQL(statements[i][2])
                                old_exist = ReadSQL("select distinct tbl_name from NEWDB.sqlite_master where tbl_name = '"+statements[i][1]+"'")
                                if str(bseq[pf][2])=="alterdb":
                                    WriteSQL("drop table if exists NEWDB._"+statements[i][1]+"_old")
                                    
                                    if len(old_exist)>0:
                                        WriteSQL("alter table NEWDB."+statements[i][1]+" rename to _"+statements[i][1]+"_old")
                                    #print(statements[i][2])
                                    WriteSQL(statements[i][2])
                                    if statements[i][3]!=None:
                                        if len(old_exist)>0:
                                            WriteSQL("insert into NEWDB."+statements[i][1]+"("+statements[i][3]+") select "+statements[i][3]+
                                                     " from _"+statements[i][1]+"_old")
                                    if statements[i][4]!=None and statements[i][4]!="" and statements[i][4]!="None":
                                        prefix = statements[i][4][0:6]
                                        prefix = prefix.upper()
                                        if prefix=="UPDATE":
                                            statements[i][4] = statements[i][4][0:7]+"NEWDB."+statements[i][4][7:]
                                            WriteSQL(statements[i][4])
                                        else:
                                            WriteSQL(statements[i][4])
                                    #WriteSQL("drop table if exists NEWDB._"+statements[i][1]+"_old")
                                WriteSQL("commit")
                                
                                if str(bseq[pf][2])=="alterdb" and len(old_exist)>0:
                                    old_cnt = ReadSQL("select count(*) from NEWDB._"+statements[i][1]+"_old")
                                    new_cnt = ReadSQL("select count(*) from NEWDB."+statements[i][1])
                                    if len(old_cnt)>0 and len(new_cnt)>0:
                                        if new_cnt[0][0]<=0 and old_cnt[0][0]>0:
                                            warning += path+bseq[pf][6]+": Failed to copy data from _"+statements[i][1]+"_old to "+statements[i][1]+"\n"
                                        else:
                                            WriteSQL("drop table if exists NEWDB._"+statements[i][1]+"_old")
                                    else:
                                        WriteSQL("drop table if exists NEWDB._"+statements[i][1]+"_old")
                                WriteSQL("pragma foreign_keys = on")
                                #if str(bseq[pf][2])=="alterdb":
                                #    WriteSQL("vacuum NEWDB")
                                WriteSQL("detach database NEWDB")
                            if str(bseq[pf][2])=="alterdb" and copysuccess==True:
                                path = bseq[pf][5]
                                dbpath = addSlash(path)
                                WriteSQL("attach database '"+dbpath+"\\\\"+bseq[pf][6]+"' as NEWDB")
                                WriteSQL("vacuum NEWDB")
                                WriteSQL("detach database NEWDB")
                            if warning!="":
                                ctypes.windll.user32.MessageBoxW(0,warning,"Warning: "+str(bseq[pf][2]),0)
                elif str(bseq[pf][2])=="createpg" or str(bseq[pf][2])=="alterpg":
                    if str(bseq[pf][3])=="" or bseq[pf][3]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in source field.","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in target field.","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        pg_data_src = str(bseq[pf][3])
                        pg_data_src = pg_data_src.split('|')
                        pg_data_trg = str(bseq[pf][4])
                        pg_data_trg = pg_data_trg.split('|')
                        self.conn = psycopg2.connect(dbname=pg_data_trg[0],user=pg_data_trg[1],password=PG_PASS,host=pg_data_trg[2],port=pg_data_trg[3])
                        self.cursor = self.conn.cursor()
                        statements = ReadSQL("select * from pgtables where hostip = '"+str(pg_data_src[2])+"' and port = "+str(pg_data_src[3])+
                                             " and username = '"+str(pg_data_src[1])+"' and password = '"+str(pg_data_src[4])+"' and dbname = '"+
                                             str(pg_data_src[0])+"'")
                        warning = ""
                        
                        if str(bseq[pf][2])=="alterpg":
                            self.WritePG("begin transaction")
                            all_tables = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public'")
                            all_table_names = []
                            for i in range(len(all_tables)):
                                all_table_names.append(str(all_tables[i][0]))
                                all_constraints = self.ReadPG("select constraint_name from information_schema.table_constraints where table_schema = 'public' and table_name = '"+
                                                              str(all_tables[i][0])+"' and (constraint_type = 'FOREIGN KEY' or constraint_type = 'PRIMARY KEY')")
                                for j in range(len(all_constraints)):
                                    self.WritePG("alter table "+str(all_tables[i][0])+" drop constraint if exists "+str(all_constraints[j][0])+" cascade")
                            self.WritePG("commit")
                            self.WritePG("begin transaction")
                            for i in range(len(statements)):
                                if statements[i][5][:1]=="\"" and statements[i][5][-1:]=="\"":
                                    temp = str(statements[i][5][1:-1])
                                else:
                                    temp = str(statements[i][5])
                                if all_table_names.count("_"+temp+"_old")<1:
                                    self.WritePG("alter table if exists "+str(statements[i][5])+" rename to _"+temp+"_old")
                            self.WritePG("commit")
                        creation = statements
                        self.WritePG("begin transaction")
                        while len(creation)>0:
                            references = creation[0][6].split("REFERENCES ")
                            if len(references)>1:
                                all_exist = True
                                references.pop(0)
                                for t in range(len(references)):
                                    table_ref = references[t].split('(')[0]
                                    does_exist = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+str(table_ref)+"'")
                                    if len(does_exist)<1:
                                        all_exist = False
                                        creation.append(creation[0])
                                        creation.pop(0)
                                        break
                                if all_exist==True:
                                    self.WritePG(creation[0][6])
                                    creation.pop(0)
                            else:
                                self.WritePG(creation[0][6])
                                creation.pop(0)
                        self.WritePG("commit")
                        statements = ReadSQL("select * from pgtables where hostip = '"+str(pg_data_src[2])+"' and port = "+str(pg_data_src[3])+
                                             " and username = '"+str(pg_data_src[1])+"' and password = '"+str(pg_data_src[4])+"' and dbname = '"+
                                             str(pg_data_src[0])+"'")
                        if str(bseq[pf][2])=="alterpg":
                            self.WritePG("begin transaction")
                            for i in range(len(statements)):
                                is_fk = statements[i][6].lower()
                                if is_fk.find('foreign key')==-1:
                                    if statements[i][5][:1]=="\"" and statements[i][5][-1:]=="\"":
                                        temp = str(statements[i][5][1:-1])
                                    else:
                                        temp = str(statements[i][5])
                                    if statements[i][7]!=None:
                                        self.WritePG("insert into "+str(statements[i][5])+"("+str(statements[i][7])+") select "+str(statements[i][7])+
                                                     " from _"+temp+"_old")
                                    if statements[i][8]!=None and statements[i][8]!="" and statements[i][8]!="None":
                                        statements[i][8] = str(statements[i][8])
                                        self.WritePG(str(statements[i][8]))
                                    old_cnt = self.ReadPG("select count(*) from _"+temp+"_old")
                                    new_cnt = self.ReadPG("select count(*) from "+str(statements[i][5]))
                                    if len(old_cnt)>0 and len(new_cnt)>0:
                                        if new_cnt[0][0]<=0 and old_cnt[0][0]>0:
                                            warning += str(bseq[pf][4])+": Failed to copy data from _"+temp+"_old to "+str(statements[i][5])+"\n"
                                        else:
                                            self.WritePG("drop table if exists _"+temp+"_old")
                                    else:
                                        self.WritePG("drop table if exists _"+temp+"_old")
                            for i in range(len(statements)):
                                is_fk = statements[i][6].lower()
                                if is_fk.find('foreign key')>=0:
                                    if statements[i][5][:1]=="\"" and statements[i][5][-1:]=="\"":
                                        temp = str(statements[i][5][1:-1])
                                    else:
                                        temp = str(statements[i][5])
                                    if statements[i][7]!=None:
                                        self.WritePG("insert into "+str(statements[i][5])+"("+str(statements[i][7])+") select "+str(statements[i][7])+
                                                     " from _"+temp+"_old")
                                    if statements[i][8]!=None and statements[i][8]!="" and statements[i][8]!="None":
                                        statements[i][8] = str(statements[i][8])
                                        self.WritePG(str(statements[i][8]))
                                    old_cnt = self.ReadPG("select count(*) from _"+temp+"_old")
                                    new_cnt = self.ReadPG("select count(*) from "+str(statements[i][5]))
                                    if len(old_cnt)>0 and len(new_cnt)>0:
                                        if new_cnt[0][0]<=0 and old_cnt[0][0]>0:
                                            warning += str(bseq[pf][4])+": Failed to copy data from _"+temp+"_old to "+str(statements[i][5])+"\n"
                                        else:
                                            self.WritePG("drop table if exists _"+temp+"_old")
                                    else:
                                        self.WritePG("drop table if exists _"+temp+"_old")
                            self.WritePG("commit")
                        '''
                        self.WritePG("begin transaction")
                        for i in range(len(statements)):
                            self.WritePG(statements[i][6])
                            old_exist = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+str(statements[i][5])+"'")
                            if statements[i][5][:1]=="\"" and statements[i][5][-1:]=="\"":
                                temp = str(statements[i][5][1:-1])
                            else:
                                temp = str(statements[i][5])
                            if str(bseq[pf][2])=="alterpg":
                                #self.WritePG("alter table if exists _"+temp+"_old disable trigger all")
                                if len(old_exist)>0:
                                    print(statements[i][6])
                                    self.WritePG(statements[i][6])
                                    if statements[i][7]!=None:
                                        if len(old_exist)>0:
                                            self.WritePG("insert into "+str(statements[i][5])+"("+str(statements[i][7])+") select "+str(statements[i][7])+
                                                         " from _"+temp+"_old")
                                    if statements[i][8]!=None and statements[i][8]!="" and statements[i][8]!="None":
                                        statements[i][8] = str(statements[i][8])
                                        self.WritePG(str(statements[i][8]))
                            
                            if str(bseq[pf][2])=="alterpg" and len(old_exist)>0:
                                old_cnt = self.ReadPG("select count(*) from _"+temp+"_old")
                                new_cnt = self.ReadPG("select count(*) from "+str(statements[i][5]))
                                if len(old_cnt)>0 and len(new_cnt)>0:
                                    if new_cnt[0][0]<=0 and old_cnt[0][0]>0:
                                        warning += str(bseq[pf][4])+": Failed to copy data from _"+temp+"_old to "+str(statements[i][5])+"\n"
                                    else:
                                        self.WritePG("drop table if exists _"+temp+"_old")
                                else:
                                    self.WritePG("drop table if exists _"+temp+"_old")
                        self.WritePG("commit")
                        if str(bseq[pf][2])=="alterpg":
                            self.WritePG("begin transaction")
                            for i in range(len(statements)):
                                if statements[i][5][:1]=="\"" and statements[i][5][-1:]=="\"":
                                    temp = str(statements[i][5][1:-1])
                                else:
                                    temp = str(statements[i][5])
                                self.WritePG("drop table if exists _"+temp+"_old")
                            self.WritePG("commit")
                        '''
                        #if str(bseq[pf][2])=="alterdb":
                        #    self.WritePG("vacuum")
                        self.conn.close()
                elif str(bseq[pf][2])=="deletedb":
                    temp00 = str(bseq[pf][5]+"\\"+bseq[pf][6])
                    if os.path.exists(temp00)==False:
                        #ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed deletedb: "+bname+"! \\"+str(bseq[pf][9]),1)
                        print(temp00+" does not exist? Button: "+bname+"\\"+str(bseq[pf][9]))
                    else:
                        try:
                            os.remove(temp00)
                        except OSError as e:
                            print("except")
                            msg = "Failed to delete "+temp00+", permission issue?"
                            ctypes.windll.user32.MessageBoxW(0,msg,"Failed to delete DB \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="rename":
                    if bseq[pf][1]=="" or bseq[pf][1]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No new name specified?","Failed rename: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(str(bseq[pf][3]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][3])+" does not exist?","Failed rename: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(str(bseq[pf][4]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][4])+" does not exist?","Failed rename: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        os.rename(str(bseq[pf][3]),str(bseq[pf][4])+"\\"+str(bseq[pf][1]))
                elif str(bseq[pf][2])=="postfield":
                    all_dbs = ReadSQL("select * from mastdbs")
                    for i in range(len(all_dbs)):
                        writeTable(all_dbs[i][0],all_dbs[i][1])
                elif str(bseq[pf][2])=="scrape":
                    self.Scrape(bseq[pf][3])
                elif str(bseq[pf][2])=="copy":
                    file_name  = bseq[pf][1] 
                    file_source = bseq[pf][3]
                    file_destination = bseq[pf][4]
                    run_sequence = bseq[pf][9]
                    new_name = bseq[pf][6] if bseq[pf][6] else file_name
                    copy_exec = self.type_copy(bname,file_name,file_source,file_destination, run_sequence, new_name)
                    if copy_exec:
                        logger.success(f"Copied {file_name} from {file_source} to {file_destination} as {new_name}")
                    else:
                        logger.error(f"Failed to copy {file_name} from {file_source} to {file_destination} as {new_name}")
                    
                elif str(bseq[pf][2])=="copyd":
                    if os.path.exists(bseq[pf][3]+"\\"+bseq[pf][1])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+"\\"+bseq[pf][1]+" source does not exist?","Failed copyd: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(bseq[pf][4])==False:
                        os.system("mkdir \""+bseq[pf][4]+"\"")
                        if os.path.exists(bseq[pf][4])==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" target does not exist?","Failed copyd: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            if os.path.exists(bseq[pf][4]+"\\"+bseq[pf][1])==False:
                                try:
                                    shutil.copy2(bseq[pf][3]+"\\"+bseq[pf][1],bseq[pf][4])
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+bseq[pf][1]+" to "+bseq[pf][4]+"?","Failed copyd: "+bname+"! \\"+str(bseq[pf][9]),0)
                            else:
                                if os.stat(bseq[pf][3]+"\\"+bseq[pf][1]).st_mtime-os.stat(bseq[pf][4]+"\\"+bseq[pf][1]).st_mtime>=1:
                                    try:
                                        shutil.copy2(bseq[pf][3]+"\\"+bseq[pf][1],bseq[pf][4]+"\\"+bseq[pf][1])
                                    except:
                                        ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+bseq[pf][1]+" to "+bseq[pf][4]+"\\"+bseq[pf][1]+"?","Failed copyd: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(bseq[pf][4]+"\\"+bseq[pf][1])==False:
                            try:
                                shutil.copy2(bseq[pf][3]+"\\"+bseq[pf][1],bseq[pf][4])
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+bseq[pf][1]+" to "+bseq[pf][4]+"?","Failed copyd: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            if os.stat(bseq[pf][3]+"\\"+bseq[pf][1]).st_mtime-os.stat(bseq[pf][4]+"\\"+bseq[pf][1]).st_mtime>=1:
                                try:
                                    shutil.copy2(bseq[pf][3]+"\\"+bseq[pf][1],bseq[pf][4]+"\\"+bseq[pf][1])
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+bseq[pf][1]+" to "+bseq[pf][4]+"\\"+bseq[pf][1]+"?","Failed copyd: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="copyext":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(bseq[pf][4])==False:
                        os.system("mkdir \""+bseq[pf][4]+"\"")
                        if os.path.exists(bseq[pf][4])==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" target does not exist?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            for basename in os.listdir(bseq[pf][3]):
                                curext = bseq[pf][1]
                                if curext==None or curext=="":
                                    curext = "*.*"
                                if getExt(basename,curext)==True:
                                    if os.path.isfile(bseq[pf][3]+"\\"+basename):
                                        if os.path.exists(bseq[pf][4]+"\\"+basename)==False:
                                            try:
                                                shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4])
                                            except:
                                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                                        else:
                                            try:
                                                shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4]+"\\"+basename)
                                            except:
                                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"\\"+basename+"?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        for basename in os.listdir(bseq[pf][3]):
                            curext = bseq[pf][1]
                            if curext==None or curext=="":
                                curext = "*.*"
                            if getExt(basename,curext)==True:
                                if os.path.isfile(bseq[pf][3]+"\\"+basename):
                                    if os.path.exists(bseq[pf][4]+"\\"+basename)==False:
                                        try:
                                            shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4])
                                        except:
                                            ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                                    else:
                                        try:
                                            shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4]+"\\"+basename)
                                        except:
                                            ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"\\"+basename+"?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="copyextd":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(bseq[pf][4])==False:
                        os.system("mkdir \""+bseq[pf][4]+"\"")
                        if os.path.exists(bseq[pf][4])==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" target does not exist?","Failed copyext: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            for basename in os.listdir(bseq[pf][3]):
                                curext = bseq[pf][1]
                                if curext==None or curext=="":
                                    curext = "*.*"
                                if getExt(basename,curext)==True:
                                    if os.path.isfile(bseq[pf][3]+"\\"+basename):
                                        if os.path.exists(bseq[pf][4]+"\\"+basename)==False:
                                            try:
                                                shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4])
                                            except:
                                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"?","Failed copyextd: "+bname+"! \\"+str(bseq[pf][9]),0)
                                        else:
                                            if os.stat(bseq[pf][3]+"\\"+basename).st_mtime-os.stat(bseq[pf][4]+"\\"+basename).st_mtime>=1:
                                                try:
                                                    shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4]+"\\"+basename)
                                                except:
                                                    ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"\\"+basename+"?","Failed copyextd: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        for basename in os.listdir(bseq[pf][3]):
                            curext = bseq[pf][1]
                            if curext==None or curext=="":
                                curext = "*.*"
                            if getExt(basename,curext)==True:
                                if os.path.isfile(bseq[pf][3]+"\\"+basename):
                                    if os.path.exists(bseq[pf][4]+"\\"+basename)==False:
                                        try:
                                            shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4])
                                        except:
                                            ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"?","Failed copyextd: "+bname+"! \\"+str(bseq[pf][9]),0)
                                    else:
                                        if os.stat(bseq[pf][3]+"\\"+basename).st_mtime-os.stat(bseq[pf][4]+"\\"+basename).st_mtime>=1:
                                            try:
                                                shutil.copy2(bseq[pf][3]+"\\"+basename,bseq[pf][4]+"\\"+basename)
                                            except:
                                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+basename+" to "+bseq[pf][4]+"\\"+basename+"?","Failed copyextd: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="copyfile":
                    if os.path.exists(bseq[pf][3]+"\\"+bseq[pf][1])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+"\\"+bseq[pf][1]+" source does not exist?","Failed copyfile: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(bseq[pf][4]+"\\"+bseq[pf][1])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+"\\"+bseq[pf][1]+" target does not exist?","Failed copyfile: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        try:
                            shutil.copyfile(bseq[pf][3]+"\\"+bseq[pf][1],bseq[pf][4]+"\\"+bseq[pf][1])
                        except:
                            ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+"\\"+bseq[pf][1]+" to "+bseq[pf][4]+"\\"+bseq[pf][1]+"?","Failed copyfile: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="copytree":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed copytree: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(bseq[pf][4])==False:
                            try:
                                shutil.copytree(bseq[pf][3],bseq[pf][4])
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+bseq[pf][3]+" to "+bseq[pf][4]+"?","Failed copytree: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" already exists. \"copytree\" cannot copy directory which already exists.","Failed copytree: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="rmtree":
                    if os.path.exists(bseq[pf][4])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" target does not exist?","Failed rmtree: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        try:
                            shutil.rmtree(bseq[pf][4])
                        except:
                            ctypes.windll.user32.MessageBoxW(0,"Problem removing "+bseq[pf][4]+"?","Failed rmtree: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="move":
                    # If doing a folder, directory can be created
                    if bseq[pf][1]==None or bseq[pf][1]=="" or bseq[pf][1]=="None":
                        if os.path.exists(bseq[pf][3])==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed move: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            try:
                                shutil.move(bseq[pf][3],bseq[pf][4])
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem moving "+bseq[pf][3]+" to "+bseq[pf][4]+"?","Failed move: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(bseq[pf][3]+"\\"+bseq[pf][1])==False:
                        if os.path.exists(bseq[pf][3])==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed move: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            try:
                                shutil.move(bseq[pf][3],bseq[pf][4])
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem moving "+bseq[pf][3]+" to "+bseq[pf][4]+"?","Failed move: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        # If doing a file, directory must already exist
                        if os.path.exists(bseq[pf][4])==False:
                            ctypes.windll.user32.MessageBoxW(0,bseq[pf][4]+" target does not exist?","Failed move: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            try:
                                shutil.move(bseq[pf][3]+"\\"+bseq[pf][1],bseq[pf][4]+"\\"+bseq[pf][1])
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem moving "+bseq[pf][3]+"\\"+bseq[pf][1]+" to "+bseq[pf][4]+"\\"+bseq[pf][1]+"?","Failed move: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="delfile":
                    if os.path.exists(bseq[pf][3]+"\\"+bseq[pf][1])==True:
                        try:
                            os.remove(bseq[pf][3]+"\\"+bseq[pf][1])
                        except:
                            ctypes.windll.user32.MessageBoxW(0,"Problem deleting "+bseq[pf][3]+"\\"+bseq[pf][1]+"?","Failed delfile: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="cleardir":
                    if os.path.exists(bseq[pf][3])==True:
                        for root, dirs, files in os.walk(bseq[pf][3]):
                            for f in files:
                                try:
                                    os.remove(bseq[pf][3]+"\\"+f)
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Problem removing "+bseq[pf][3]+"\\"+f,"Warning: cleardir: "+bname+"! \\"+str(bseq[pf][9]),0)
                                #os.unlink(os.path.join(root,f))
                            for d in dirs:
                                try:
                                    shutil.rmtree(bseq[pf][3]+"\\"+d)
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Problem removing "+bseq[pf][3]+"\\"+d,"Warning: cleardir: "+bname+"! \\"+str(bseq[pf][9]),0)
                                #shutil.rmtree(os.path.join(root,d))
                    else:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed cleardir: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="clearfiletx":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed cleardirext: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        ext = bseq[pf][1]
                        if ext==None or ext=="":
                            ext = "*.*"
                        for root, dirs, files in os.walk(bseq[pf][3]):
                            for f in files:
                                if getExt(bseq[pf][3]+"\\"+f,ext)==True:
                                    if os.path.exists(bseq[pf][3]+"\\"+f)==True:
                                        try:
                                            os.remove(bseq[pf][3]+"\\"+f)
                                        except:
                                            ctypes.windll.user32.MessageBoxW(0,"Problem removing "+bseq[pf][3]+"\\"+f,"Warning: clearfiletx: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="cleardirext":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed cleardirext: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        ext = bseq[pf][1]
                        if ext==None or ext=="":
                            ext = "*.*"
                        deleteExt(bseq[pf][3],ext,bname,bseq[pf][9])
                elif str(bseq[pf][2])=="copydir":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed copydir: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"target not specified?","Failed copydir: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        copyDir(bseq[pf][3],bseq[pf][4])####
                elif str(bseq[pf][2])=="copydird":
                    if os.path.exists(bseq[pf][3])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][3]+" source does not exist?","Failed copydird: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif bseq[pf][4]==None or bseq[pf][4]=="":
                        ctypes.windll.user32.MessageBoxW(0,"target not specified?","Failed copydird: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        copyDir(bseq[pf][3],bseq[pf][4],True)
                elif str(bseq[pf][2])=="sleep":
                    print("sleeping")
                    time.sleep(1)
                    print("awake")
                elif str(bseq[pf][2])=="vacuum":
                    if os.path.exists(bseq[pf][5]+"\\"+bseq[pf][6])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][5]+"\\"+bseq[pf][6]+" database path does not exist?","Failed vacuum: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        path = bseq[pf][5]+"\\"+bseq[pf][6]
                        if path[0:1]=="\\":
                            path = path[1:]
                        path = path.replace("\\","\\\\")
                        vc = apsw.Connection(path)
                        vcursor = vc.cursor()
                        aWriteSQL(vcursor,"vacuum")
                        vc.close(True)
                elif str(bseq[pf][2])=="ps1":
                    if os.path.exists(bseq[pf][0]+"\\"+bseq[pf][1])==False:
                        ctypes.windll.user32.MessageBoxW(0,bseq[pf][0]+"\\"+bseq[pf][1]+" database path does not exist?","Failed ps1: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.RunPS1(bseq[pf][0]+"\\"+bseq[pf][1])
                elif str(bseq[pf][2])=="url":
                    path = str(bseq[pf][0])+"\\"+str(bseq[pf][1])
                    if os.path.exists(path)==False:
                        try:
                            webbrowser.open(str(bseq[pf][3]))  # Go to example.com
                        except Exception as e:
                            ctypes.windll.user32.MessageBoxW(0,path+" does not exist?","Failed url: "+bname+"! \\"+str(bseq[pf][9]),0)
                            logger.error(e)
                    else:

                        subprocess.call([path,'-new-tab',str(bseq[pf][3])])
                    #webbrowser.open(str(bseq[pf][3]))
                elif str(bseq[pf][2])=="mt5event":
                    self.MT_Event("MT5",bname,bseq[pf][9],bseq[pf][0],bseq[pf][1])
                elif str(bseq[pf][2])=="mtevent":
                    self.MT_Event("MT4",bname,bseq[pf][9],bseq[pf][0],bseq[pf][1])
                elif str(bseq[pf][2])=="assignseries":
                    if str(bseq[pf][3])==None:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][3])+" button series name not assigned in source?","Failed assignseries: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        get_series = ReadSQL("select formname,tab,buttonname from buttonseries where assignname = '"+str(bseq[pf][3])+"' order by runsequence asc")
                        if len(get_series)<=0:
                            ctypes.windll.user32.MessageBoxW(0,"No assigned buttons found for "+str(bseq[pf][3])+" series assignment?","Failed assignseries: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            for i in range(len(get_series)):
                                butp = get_series[i][0]
                                butt = get_series[i][1]
                                butb = get_series[i][2]
                                if butp!=None and butp!="":
                                    self.on_click_button(butp,butt,butb,"",1)
                elif str(bseq[pf][2])=="archivedb":
                    if os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" database path does not exist?","Failed archivedb: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        temp00 = str(bseq[pf][0])
                        if os.path.exists(temp00)==False:
                            os.system("mkdir \""+temp00+"\"")
                        gettime = time.time()
                        newdb = bseq[pf][6][:-3]+"_"+str(datetime.datetime.fromtimestamp(gettime).strftime('%Y-%m-%d_%H-%M-%S'))+".db"
                        if os.path.exists(temp00+"\\"+newdb)==False:
                            try:
                                shutil.copy2(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),temp00+"\\"+newdb)
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem copying "+str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" to "+temp00+"\\"+newdb+"?","Failed archivedb: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            if os.stat(str(bseq[pf][5])+"\\"+str(bseq[pf][6])).st_mtime-os.stat(temp00+"\\"+newdb).st_mtime>=1:
                                try:
                                    shutil.copy2(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),temp00+"\\"+newdb)
                                except:
                                    ctypes.windll.user32.MessageBoxW(0,"Problem copying "+str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" to "+temp00+"\\"+newdb+"?","Failed archivedb: "+bname+"! \\"+str(bseq[pf][9]),0)
                elif str(bseq[pf][2])=="sqlpgfresh":
                    if os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" database path does not exist?","Failed sqlpgfresh: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in source field.","Failed sqlpgfresh: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.SQL_To_PG(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][4]),bseq[pf][1],0)
                elif str(bseq[pf][2])=="sqlpgignore":
                    if os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" database path does not exist?","Failed sqlpgignore: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in source field.","Failed sqlpgignore: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.SQL_To_PG(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][4]),bseq[pf][1],1)
                elif str(bseq[pf][2])=="sqlpgreplace":
                    if os.path.exists(str(bseq[pf][5])+"\\"+str(bseq[pf][6]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][5])+"\\"+str(bseq[pf][6])+" database path does not exist?","Failed sqlpgreplace: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in source field.","Failed sqlpgreplace: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        self.SQL_To_PG(str(bseq[pf][5])+"\\"+str(bseq[pf][6]),str(bseq[pf][4]),bseq[pf][1],2)
                elif str(bseq[pf][2])=="ordcreate":
                    if str(bseq[pf][3])=="" or bseq[pf][3]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in source field.","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in target field.","Failed "+str(bseq[pf][2])+": "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        pg_data_src = str(bseq[pf][3])
                        pg_data_src = pg_data_src.split('|')
                        pg_data_trg = str(bseq[pf][4])
                        pg_data_trg = pg_data_trg.split('|')
                        self.conn = psycopg2.connect(dbname=pg_data_trg[0],user=pg_data_trg[1],password=PG_PASS,host=pg_data_trg[2],port=pg_data_trg[3])
                        self.cursor = self.conn.cursor()
                        if str(bseq[pf][1])!="" and bseq[pf][1]!=None:
                            creation = ReadSQL("select * from pgtables where hostip = '"+str(pg_data_src[2])+"' and port = "+str(pg_data_src[3])+
                                               " and username = '"+str(pg_data_src[1])+"' and password = '"+str(pg_data_src[4])+"' and dbname = '"+
                                               str(pg_data_src[0])+"' and tbl_name = '"+str(bseq[pf][1])+"'")
                        else:
                            creation = ReadSQL("select * from pgtables where hostip = '"+str(pg_data_src[2])+"' and port = "+str(pg_data_src[3])+
                                               " and username = '"+str(pg_data_src[1])+"' and password = '"+str(pg_data_src[4])+"' and dbname = '"+
                                               str(pg_data_src[0])+"' and ("+self.createOrdCreate()+")")
                        self.WritePG("begin transaction")
                        while len(creation)>0:
                            references = creation[0][6].split("REFERENCES ")
                            if len(references)>1:
                                all_exist = True
                                references.pop(0)
                                for t in range(len(references)):
                                    table_ref = references[t].split('(')[0]
                                    does_exist = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+str(table_ref)+"'")
                                    if len(does_exist)<1:
                                        all_exist = False
                                        creation.append(creation[0])
                                        creation.pop(0)
                                        break
                                if all_exist==True:
                                    self.WritePG(creation[0][6])
                                    creation.pop(0)
                            else:
                                #print(creation[0][6]+"\n\n")
                                self.WritePG(creation[0][6])
                                creation.pop(0)
                        self.WritePG("commit")
                        self.conn.close()
                elif str(bseq[pf][2])=="ordbackup":
                    if str(bseq[pf][3])=="" or bseq[pf][3]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in source field.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in target field.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][0])=="":
                        ctypes.windll.user32.MessageBoxW(0,"Folder path is seen as blank.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif bseq[pf][0]==None:
                        ctypes.windll.user32.MessageBoxW(0,"Folder path is seen as 'null'.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif os.path.exists(str(bseq[pf][0]))==False:
                        ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+" does not exist or is not seen.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][1])=="" or bseq[pf][1]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No filename provided.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        filename = str(bseq[pf][1])
                        fname = filename.split('|')
                        pg_data = str(bseq[pf][3])
                        pg_data = pg_data.split('|')
                        pg_data2 = str(bseq[pf][4])
                        pg_data2 = pg_data2.split('|')
                        if len(pg_data)<5:
                            ctypes.windll.user32.MessageBoxW(0,"Insufficient Postgres source credentials.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                        elif len(pg_data2)<5:
                            ctypes.windll.user32.MessageBoxW(0,"Insufficient Postgres target credentials.","Failed ordbackup: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            curtime = time.strftime("%Y.%m.%d-%H.%M.%S")
                            if len(fname)>1:
                                #print("SET PGPASSWORD="+str(PG_PASS)+"& pg_dump -h "+str(pg_data[2])+" -p "+str(pg_data[3])+
                                #      " -U "+str(pg_data[1])+" --column-inserts --data-only -t "+fname[1]+" "+str(pg_data[0])+
                                #" > \""+str(bseq[pf][0])+"\\ord_bak_"+fname[0]+"_"+curtime+".sql\"")
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& pg_dump -h "+str(pg_data[2])+" -p "+str(pg_data[3])+" -U "+str(pg_data[1])+
                                          " --column-inserts --data-only -t "+fname[1]+" "+str(pg_data[0])+" > \""+str(bseq[pf][0])+"\\ord_bak_"+fname[0]+
                                          "_"+curtime+".sql\"")
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& pg_dump -h "+str(pg_data[2])+" -p "+str(pg_data[3])+" -U "+str(pg_data[1])+
                                          " --column-inserts --data-only -t "+fname[1]+" "+str(pg_data[0])+" > \""+str(bseq[pf][0])+"\\ord_bak_"+fname[1]+
                                          "_temp.sql\"")
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& psql -h "+str(pg_data2[2])+" -p "+str(pg_data2[3])+" -U "+str(pg_data2[1])+" "+
                                          str(pg_data2[0])+" < \""+str(bseq[pf][0])+"\\ord_bak_"+fname[0]+"_"+curtime+".sql\"")
                            else:
                                query = self.createOrdBackup()
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& pg_dump -h "+str(pg_data[2])+" -p "+str(pg_data[3])+
                                          " -U "+str(pg_data[1])+" --column-inserts --data-only"+query+" "+
                                          str(pg_data[0])+" > \""+str(bseq[pf][0])+"\\ord_bak_"+fname[0]+"_"+curtime+".sql\"")
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& pg_dump -h "+str(pg_data[2])+" -p "+str(pg_data[3])+
                                          " -U "+str(pg_data[1])+" --column-inserts --data-only"+query+" "+
                                          str(pg_data[0])+" > \""+str(bseq[pf][0])+"\\ord_bak_temp.sql\"")
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& psql -h "+str(pg_data2[2])+" -p "+str(pg_data2[3])+" -U "+str(pg_data2[1])+" "+
                                          str(pg_data2[0])+" < \""+str(bseq[pf][0])+"\\ord_bak_"+fname[0]+"_"+curtime+".sql\"")
                elif str(bseq[pf][2])=="ordrestore":
                    if str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        ctypes.windll.user32.MessageBoxW(0,"No Postgres credentials specified in target field.","Failed ordrestore: "+bname+"! \\"+str(bseq[pf][9]),0)
                    elif str(bseq[pf][0])=="" or bseq[pf][0]==None or os.path.exists(str(bseq[pf][0]))==False:
                        ctypes.windll.user32.MessageBoxW(0,"No folderpath specified or folderpath provided does not exist.","Failed ordrestore: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        pg_data2 = str(bseq[pf][4])
                        pg_data2 = pg_data2.split('|')
                        if len(pg_data2)<5:
                            ctypes.windll.user32.MessageBoxW(0,"Insufficient Postgres target credentials.","Failed ordrestore: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            if str(bseq[pf][1])!="" and bseq[pf][1]!=None:
                                fname = str(bseq[pf][1])
                                if os.path.exists(str(bseq[pf][0])+"\\"+fname)==False:
                                    ctypes.windll.user32.MessageBoxW(0,str(bseq[pf][0])+"\\"+fname+" does not exist.","Failed ordrestore: "+bname+"! \\"+str(bseq[pf][9]),0)
                                else:
                                    os.system("SET PGPASSWORD="+str(PG_PASS)+"& psql -h "+str(pg_data2[2])+" -p "+str(pg_data2[3])+" -U "+str(pg_data2[1])+" "+
                                              str(pg_data2[0])+" < \""+str(bseq[pf][0])+"\\ord_bak_"+fname+"_temp.sql\"")
                                    os.remove(str(bseq[pf][0])+"\\ord_bak_"+fname+"_temp.sql")
                            else:
                                os.system("SET PGPASSWORD="+str(PG_PASS)+"& psql -h "+str(pg_data2[2])+" -p "+str(pg_data2[3])+" -U "+str(pg_data2[1])+" "+
                                          str(pg_data2[0])+" < \""+str(bseq[pf][0])+"\\ord_bak_temp.sql\"")
                                os.remove(str(bseq[pf][0])+"\\ord_bak_temp.sql")
                elif str(bseq[pf][2])=="emojiseek":
                    if str(bseq[pf][4])=="" or bseq[pf][4]==None:
                        n = QInputDialog.getText(self,"emoji","Enter Unicode for desired emoji: ",text='')
                        if n[1]==True:
                            try:
                                val = ""
                                string = str(n[0])
                                string = string.split('|')
                                for i in range(len(string)):
                                    emoji = int(string[i],16)
                                    val += self.getEmoji(emoji)
                                if val=="":
                                    ctypes.windll.user32.MessageBoxW(0,"Cannot convert value "+str(n[0]),"Failed emojiseek: "+bname+"! \\"+str(bseq[pf][9]),0)
                                else:
                                    QInputDialog.getText(self,"emoji","Outcome: ",text=str(val))
                            except Exception as e1:
                                print(str(e1))
                                ctypes.windll.user32.MessageBoxW(0,"Cannot convert value "+str(n[0]),"Failed emojiseek: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        pg_data = str(bseq[pf][4])
                        if len(pg_data)<5:
                            ctypes.windll.user32.MessageBoxW(0,"Insufficient Postgres target credentials.","Failed ordrestore: "+bname+"! \\"+str(bseq[pf][9]),0)
                        else:
                            n = QInputDialog.getText(self,"signaltype","Enter signaltype for desired emoji: ",text='')
                            if n[0]!="" and n[1]==True:
                                signaltype = n[0]
                                n = QInputDialog.getText(self,"association","Enter association for desired emoji: ",text='')
                                if n[0]!="" and n[1]==True:
                                    association = n[0]
                                    n = QInputDialog.getText(self,"emoji","Enter Unicode for desired emoji: ",text='')
                                    if n[0]!="" and n[1]==True:
                                        val = ""
                                        string = str(n[0])
                                        string = string.split('|')
                                        for i in range(len(string)):
                                            emoji = int(string[i],16)
                                            val += self.getEmoji(emoji)
                                        if val=="":
                                            ctypes.windll.user32.MessageBoxW(0,"Cannot convert value "+str(n[0]),"Failed emojiseek: "+bname+"! \\"+str(bseq[pf][9]),0)
                                        else:
                                            pg_data = pg_data.split('|')
                                            self.conn = psycopg2.connect(dbname=pg_data[0],user=pg_data[1],password=PG_PASS,host=pg_data[2],port=pg_data[3])
                                            self.cursor = self.conn.cursor()
                                            self.WritePG("insert into telegram_emojis values('"+str(signaltype)+"','"+str(association)+"','"+
                                                         str(val)+"',null) on conflict(signaltype,association) do update set signaltype = '"+
                                                         str(signaltype)+"', association = '"+str(association)+"', emoji = '"+str(val)+"';")
                                            self.conn.close()
                elif str(bseq[pf][2])=="tabto":
                    if str(bseq[pf][1])!="" and bseq[pf][1]!=None:
                        self.lasttab = self.SM_Tabs.tabText(self.SM_Tabs.currentIndex())
                        print(str(bseq[pf][1]))
                        index = [index for index in range(self.SM_Tabs.count()) if str(bseq[pf][1]) == self.SM_Tabs.tabText(index)]
                        if len(index)>0:
                            print(str(index[0]))
                            self.SM_Tabs.setCurrentIndex(index[0])
                elif str(bseq[pf][2])=="tablast":
                    if self.lasttab!="":
                        index = [index for index in range(self.SM_Tabs.count()) if str(self.lasttab) == self.SM_Tabs.tabText(index)]
                        if len(index)>0:
                            self.SM_Tabs.setCurrentIndex(index[0])
                elif str(bseq[pf][2])=="buttondialog":
                    self.Create_Button_Dialog(pname,tname,bname,bseq[pf][0],bseq[pf][1],1)
                elif str(bseq[pf][2])=="buttontask":
                    self.Create_Button_Dialog(pname,tname,bname,bseq[pf][0],bseq[pf][1],2)
        if mode==0:
            self.findChild(QPushButton,objn).setStyleSheet("QPushButton { background-color: none }")
        self.Refresh()
        
    def getEmoji(self,e):
        try:
            e = chr(e)
            return(urllib.parse.quote_from_bytes(e.encode('utf-8')))
        except:
            return("")

    def createOrdBackup(self):
        val = ""
        string = ReadSQL("select pgschemaname,pgtablename from pgtablebak")
        if len(string)>0:
            for i in range(len(string)):
                schema = ""
                if string[i][0]!=None and string[i][0]!="":
                    schema = str(string[i][0])+"."
                val += " -t "+schema+str(string[i][1])
        return(val)
    
    def createOrdCreate(self):
        val = ""
        string = ReadSQL("select pgschemaname,pgtablename from pgtablebak")
        if len(string)>0:
            for i in range(len(string)):
                schema = ""
                if string[i][0]!=None and string[i][0]!="":
                    schema = str(string[i][0])+"."
                val += "tbl_name = '"+schema+str(string[i][1])+"' or "
            val = val[:-4]
        return(val)

    def ExportExcel(self):
        eng = ReadSQL("select * from sqlengine")
        print("sqlengine data?")
        print(eng)
        if len(eng)<=0:
            ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed sql!",0)
        else:
            eseq = ReadSQL("select * from pyexcelmenu where (type = 'Export' or type = 'export') order by runsequence asc")
            for exf in range(len(eseq)):
                temp00 = str(eseq[exf][2])
                if os.path.exists(temp00)==False:
                    ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed Tools/Export to Excel!",0)
                elif os.path.exists(str(eseq[exf][4])+"\\"+str(eseq[exf][5]))==False:
                    ctypes.windll.user32.MessageBoxW(0,str(eseq[exf][4])+"\\"+str(eseq[exf][5])+" does not exist?","Failed Tools/Export to Excel!",0)
                else:
                    runExcel(1,str(eseq[exf][2])+"\\"+str(eseq[exf][3]),str(eng[0][0]),str(eseq[exf][4])+"\\"+str(eseq[exf][5]),str(eseq[exf][2]),"",0,1)

    def ImportExcel(self):
        new_button = str(QInputDialog.getText(None,"Confirm Import",'Type "yes" to confirm import',QLineEdit.Normal,"")[0])
        if new_button.lower() != "yes":
            return
        eng = ReadSQL("select * from sqlengine")
        print("sqlengine data?")
        print(eng)
        if len(eng)<=0:
            ctypes.windll.user32.MessageBoxW(0,"No record exists in sqlengine table?","Failed sql!",0)
        else:
            iseq = ReadSQL("select * from pyexcelmenu where (type = 'Import' or type = 'import') order by runsequence asc")
            for imf in range(len(iseq)):
                temp00 = str(iseq[imf][2])
                if os.path.exists(temp00)==False:
                    ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed Tools/Import from Excel!",0)
                elif os.path.exists(str(eng[0][0]))==False:
                    ctypes.windll.user32.MessageBoxW(0,str(eng[0][0])+" does not exist?","Failed Tools/Import from Excel!",0)
                elif os.path.exists(str(iseq[imf][4])+"\\"+str(iseq[imf][5]))==False:
                    ctypes.windll.user32.MessageBoxW(0,str(iseq[imf][4])+"\\"+str(iseq[imf][5])+" does not exist?","Failed Tools/Import from Excel!",0)
                else:
                    runExcel(0,str(iseq[imf][2])+"\\"+str(iseq[imf][3]),str(eng[0][0]),str(iseq[imf][4])+"\\"+str(iseq[imf][5]),str(iseq[imf][2]),"",0,0)
        self.Refresh()
                        
    def isNumeric(self,n):
        try:
            float(n)
        except:
            return(False)
        else:
            return(True)
    
    # Open .exe program
    def Open_App(self,ftype,path,file,specfile):
        if os.path.exists(path+"\\"+file)==True:
            if ftype=="exe":
                if specfile!=None and specfile!="" and specfile!="None":
                    param = specfile
                    specfile = specfile.split('~')
                    if self.isNumeric(specfile[0])==True:
                        try:
                            win32api.WinExec("\""+path+"\\"+file+"\" \""+param+"\"")
                        except:
                            ctypes.windll.user32.MessageBoxW(0,"Problem running \""+path+"\\"+file+"\" \""+specfile[0]+"\"?","Failed exe: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(specfile[0])==False:
                            ctypes.windll.user32.MessageBoxW(0,str(specfile[0])+" does not exist?","Failed exe!",0)
                        else:
                            try:
                                win32api.WinExec("\""+path+"\\"+file+"\" \""+specfile[0]+"\"")
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem running \""+path+"\\"+file+"\" \""+specfile[0]+"\"?","Failed exe: "+bname+"! \\"+str(bseq[pf][9]),0)
                else:
                    try:
                        os.startfile("\""+path+"\\"+file+"\"")
                    except:
                        ctypes.windll.user32.MessageBoxW(0,"Problem running \""+path+"\\"+file+"\"?","Failed exe: "+bname+"! \\"+str(bseq[pf][9]),0)
            elif ftype=="py":
                if specfile!=None and specfile!="" and specfile!="None":
                    param = specfile
                    specfile = specfile.split('~')
                    if self.isNumeric(specfile[0])==True:
                        try:
                            win32api.WinExec("python \""+path+"\\"+file+"\" \""+param+"\"")
                        except:
                            ctypes.windll.user32.MessageBoxW(0,"Problem running \""+path+"\\"+file+"\" \""+specfile[0]+"\"?","Failed py: "+bname+"! \\"+str(bseq[pf][9]),0)
                    else:
                        if os.path.exists(specfile[0])==False:
                            ctypes.windll.user32.MessageBoxW(0,str(specfile[0])+" does not exist?","Failed py!",0)
                        else:
                            try:
                                win32api.WinExec("python \""+path+"\\"+file+"\" \""+specfile[0]+"\"")
                            except:
                                ctypes.windll.user32.MessageBoxW(0,"Problem running \""+path+"\\"+file+"\" \""+specfile[0]+"\"?","Failed py: "+bname+"! \\"+str(bseq[pf][9]),0)
                else:
                    try:
                        win32api.WinExec("python \""+path+"\\"+file+"\"")
                    except:
                        ctypes.windll.user32.MessageBoxW(0,"Problem running \""+path+"\\"+file+"\"?","Failed py: "+bname+"! \\"+str(bseq[pf][9]),0)
        else:
            actualpath = path+"\\"+file
            print(f"Program not found at location {actualpath}")
    def Scrape(self,link):
        def Scrape_Main_Table(link):
            get_scrape = requests.get(link).text
            soup = BeautifulSoup(get_scrape,'lxml')
            
            WriteSQL("create table if not exists CryptoRank(address char(1023) "+
                     "primary key,crypto char(63),rank integer,balance double); "+
                     "pragma foreign_keys = on")
            
            WriteSQL("delete from CryptoRank")
            
            soup_data = soup.find('table',{'class':'t99btc-rich-list'})
            for d in soup_data:
                try:
                    r = d.find('td',{'class':'t99btc-rl-index'}).text
                    r = int(r)
                    
                    x = d.find('a',{'class':'external'})
                    if x.has_attr('href'):
                        adr = x['href']

                    bal = d.find('td',{'class':'t99btc-rl-balance'}).text
                    bal = bal.split(' ')
                    bal[0] = bal[0].replace(',','')

                    WriteSQL("insert into CryptoRank values('"+str(adr)+"','"+bal[1]+
                             "',"+str(r)+","+str(bal[0])+")")
                except:
                    pass
        def Scrape_Summary_Table():
            WriteSQL("create table if not exists CryptoSummary(address char(1023) "+
                     "references CryptoRank(address) on delete cascade on update "+
                     "cascade primary key,crypto char(63),rank integer,"+
                     "transaction_count integer,tot_received double,final_bal "+
                     "double); pragma foreign_keys = on")

            WriteSQL("delete from CryptoSummary")

            adr_list = ReadSQL("select address from CryptoRank")
            rank = 0
            for link in range(len(adr_list)):
                rank = rank+1
                url = adr_list[link][0]
                source = requests.get(url).text
                soup2 = BeautifulSoup(source,'lxml')

                soup_data2 = soup2.find('div',{'class':'row'})
                for d in soup_data2:
                    try:
                        transactions = d.find('td',{'id':'n_transactions'}).text

                        tot_rec = d.find('td',{'id':'total_received'}).text
                        fin_bal = d.find('td',{'id':'final_balance'}).text
                        tot_rec = tot_rec.split(' ')
                        tot_rec[0] = tot_rec[0].replace(',','')
                        fin_bal = fin_bal.split(' ')
                        fin_bal[0] = fin_bal[0].replace(',','')

                        WriteSQL("insert into CryptoSummary values('"+url+"','"+
                                 fin_bal[1]+"',"+str(rank)+","+transactions+","+
                                 FormatMR(tot_rec[0],2,False)+","+
                                 FormatMR(fin_bal[0],2,False)+")")
                    except:
                        pass
        def Scrape_Transactions_Table():
            WriteSQL("create table if not exists CryptoTransactions(transactions "+
                     "char(1023) primary key,address char(1023) references "+
                     "CryptoRank(address) on delete cascade on update cascade,crypto "+
                     "char(63),rank integer,transtimestamp datetime,exchanged double,"+
                     "usdexchanged double,pagelink char(1023)); pragma foreign_keys = on")

            WriteSQL("delete from CryptoTransactions")
            Crypto = "BTC"

            source2 = requests.get('http://markets.businessinsider.com/currencies/btc-usd').text
            soup3 = BeautifulSoup(source2,'lxml')
            soup_data3 = soup3.find_all('span',{'class':'aktien-big-font text-nowrap'})
            
            for i in soup_data3:
                conv = i.find('span',{'class':'push-data'}).text
                conv = float(conv.replace(',',''))
            
            adr_list = ReadSQL("select address from CryptoSummary")
            rank = 0
            for link in range(len(adr_list)):
                rank = rank+1
                url = adr_list[link][0]
                source = requests.get(url).text
                soup2 = BeautifulSoup(source,'lxml')

                num = float(soup2.find('td',{'id':'n_transactions'}).text)
                calc = math.ceil(num/50)
                if calc>2:
                    calc = 2
                else:
                    calc = calc
                
                for i in range(calc):
                    x = str(i*50)
                    url2 = url+'?offset='+x+'&filter=6'
                    source2 = requests.get(url2).text
                    soup3 = BeautifulSoup(source2,'lxml')

                    soup_data2 = soup3.find_all('div',{'class':'txdiv'})
                    for d in soup_data2:
                        try:
                            x = d.find('a',{'class':'hash-link'})
                            if x.has_attr('href'):
                                Trans_Link = 'https://blockchain.info'+x['href']

                            dates = d.find('span',{'class':'pull-right'}).text

                            Exchanged = str(d.find('button',{'class':'btn btn-danger cb'}).text)
                            Exchanged = Exchanged[0:-4]
                            Exchanged = float(Exchanged.replace(',',''))
                            
                            Cost_USD = Exchanged*conv
                            
                            WriteSQL("insert into CryptoTransactions values('"+Trans_Link+"','"+url+
                                     "','"+Crypto+"',"+str(rank)+",'"+dates+"',"+
                                     FormatMR(Exchanged,2,False)+","+FormatMR(Cost_USD,2,False)+
                                     ",'"+url2+"')")
                        except:
                            pass
            
        WriteSQL("begin")
        print("Main")
        Scrape_Main_Table(link)
        print("Summary")
        Scrape_Summary_Table()
        print("Transactions")
        Scrape_Transactions_Table()
        WriteSQL("commit")
        print("finished")

    def edit_button(self, bseq,pname,tname,bname,objn, mode = "edit"):
        #create popup window with textbox for each item in bseq array with another textbox for textbox for bname, tname, and pname
        #add or remove buttons
        #use QMessageBox to display popup


        edit_pop = Edit_Popup(self, bseq,pname,tname,bname,objn)
        edit_pop.show()





        

            



    #INI
    def Open_Ini_Dialog(self,mt,p,t,b,fp,r,tp,mode,path=""):
        dto = QtWidgets.QDialog()
        dto.ui = Ui_Dialog()
        dto.ui.setupUi(dto)

        self.mtini_broker = "* Broker"
        self.mtini_acctnum = "* AcctNum"

        get_broker = ReadSQL("select broker.brokerid,broker.acctcompany from broker;")
        
        dto.ui.Ini_Broker.clear()
        dto.ui.Ini_Broker.addItem("* Broker")
        for i in range(len(get_broker)):
            if get_broker[i][0]!=None and get_broker[i][0]!="" and get_broker[i][0]!="None":
                dto.ui.Ini_Broker.addItem(get_broker[i][0]+"~"+get_broker[i][1])
        dto.ui.Ini_Broker.setCurrentIndex(0)
        dto.ui.Ini_AcctNum.clear()
        dto.ui.Ini_AcctNum.addItem("* AcctNum")
        dto.ui.Ini_AcctNum.setCurrentIndex(0)
        dto.ui.Ini_ServerName.setText("")

        dto.ui.Ini_Broker.activated.connect(partial(self.Change_Ini_Broker,dto))
        dto.ui.Ini_AcctNum.activated.connect(partial(self.Change_Ini_AcctNum,dto))
            
        result = dto.exec_()
        if result==1:
            if self.mtini_broker!="* Broker" and self.mtini_acctnum!="* AcctNum":
                broketab = self.mtini_broker.split('~')
                accttab = self.mtini_acctnum.split('~')

                mt4config = ReadSQL("select affinity,profile,marketwatch,login,password,server,autoconfig,sym,period,template "+
                                    "from mt4config where formname = '"+p+"' and tab = '"+broketab[0]+"' and buttonname = '"+
                                    accttab[0]+"'")
                if len(mt4config)>0:
                    temp00 = str(tp)
                    if temp00[-1:]=="\\":
                        temp00 = temp00[:-1]
                    if os.path.exists(temp00)==False:
                        ctypes.windll.user32.MessageBoxW(0,temp00+" does not exist?","Failed stini/mtini: "+b+"! \\"+str(r),0)
                    else:
                        affi = ""
                        if str(mt4config[0][0])!='None' and str(mt4config[0][0])!='':
                            affi = "C:\\Windows\\system32\\cmd.exe /c start \"Process Monitor\" /affinity "+str(mt4config[0][0])+" "
                        f = open(str(tp)+"\\config\\start.ini","w+")
                        if str(mt4config[0][1])!='None' and str(mt4config[0][1])!='':
                            f.write("Profile="+str(mt4config[0][1])+"\n")
                        if str(mt4config[0][2])!='None' and str(mt4config[0][2])!='':
                            f.write("MarketWatch="+str(mt4config[0][2])+"\n")
                        if str(mt4config[0][3])!='None' and str(mt4config[0][3])!='':
                            f.write("Login="+str(mt4config[0][3])+"\n")
                        if str(mt4config[0][4])!='None' and str(mt4config[0][4])!='':
                            f.write("Password="+str(mt4config[0][4])+"\n")
                        if str(mt4config[0][5])!='None' and str(mt4config[0][5])!='':
                            f.write("Server="+str(mt4config[0][5])+"\n")
                        if str(mt4config[0][6])!='None' and str(mt4config[0][6])!='':
                            if int(mt4config[0][6])<=0:
                                f.write("AutoConfiguration=false\n")
                            else:
                                f.write("AutoConfiguration=true\n")
                        if str(mt4config[0][7])!='None' and str(mt4config[0][7])!='':
                            f.write("Symbol="+str(mt4config[0][7])+"\n")
                        if str(mt4config[0][8])!='None' and str(mt4config[0][8])!='':
                            f.write("Period="+str(mt4config[0][8])+"\n")
                        if str(mt4config[0][9])!='None' and str(mt4config[0][9])!='':
                            f.write("Template="+str(mt4config[0][9])+"\n")
                        f.write("ExpertsEnable=true\n")
                        f.write("ExpertsDllImport=true\n")
                        f.write("ExpertsExpImport=true\n")
                        f.write("ExpertsTrades=true\n")
                        f.close()
                        if mode==1:
                            if mt=="MT5":
                                cmd = affi+str(tp)+"\\terminal64.exe /portable "+str(tp)+"\\config\\start.ini"
                            else:
                                cmd = affi+str(tp)+"\\terminal.exe /portable "+str(tp)+"\\config\\start.ini"
                            if str(mt4config[0][0])!='None' and str(mt4config[0][0])!='':
                                os.system(cmd)
                            else:
                                win32api.WinExec(cmd)
                        else:
                            WriteSQL("update buttons set active = 'A' where formname = '"+p+"' and tab = '"+t+
                                    "' and buttonname = '"+b+"'")
                            spath = os.path.realpath(sys.argv[0])
                            spath = os.path.dirname(spath)+"\\"
                            spath = addSlash(spath)
                            WriteSQL("attach database '"+spath+"PICAT_SM_connect.db' as PCONNECT")
                            doesExist = ReadSQL("select * from PCONNECT.buttons where formname = '"+p+"' and tab = '"+t+
                                                "' and buttonname = '"+b+"'")
                            if len(doesExist)<=0:
                                WriteSQL("insert into PCONNECT.buttons select * from buttons where formname = '"+p+"' and tab = '"+t+
                                         "' and buttonname = '"+b+"'")
                            WriteSQL("detach database PCONNECT")
                            if mt=="MT5":
                                f = open(str(tp)+"\\MQL5\\Files\\serverpath.csv","w+")
                            else:
                                f = open(str(tp)+"\\MQL4\\Files\\serverpath.csv","w+")
                            if tp[-1:]=="\\":
                                f.write(str(tp[:-1]))
                            else:
                                f.write(str(tp))
                            f.close()
            else:
                if mt=="MT5":
                    f = open(str(tp)+"\\MQL5\\Files\\serverpath.csv","w+")
                else:
                    f = open(str(tp)+"\\MQL4\\Files\\serverpath.csv","w+")
                if tp[-1:]=="\\":
                    f.write(str(tp[:-1]))
                else:
                    f.write(str(tp))
                    f.close()
                    #############################

    def Change_Ini_Broker(self,d):
        self.mtini_broker = d.findChild(QComboBox,"Ini_Broker").currentText()
        d.ui.Ini_AcctNum.clear()
        d.ui.Ini_AcctNum.addItem("* AcctNum")
        if self.mtini_broker!="* Broker":
            selection = self.mtini_broker
            selection = selection.split('~')
            get_acctnum = ReadSQL("select mt4config.buttonname,broker.brokerid from mt4config inner join broker on "+
                                  "mt4config.tab = broker.brokerid where broker.brokerid = '"+selection[0]+"'")
            for i in range(len(get_acctnum)):
                if get_acctnum[i][0]!=None and get_acctnum[i][0]!="" and get_acctnum[i][0]!="None":
                    d.ui.Ini_AcctNum.addItem(str(get_acctnum[i][0]))
        d.ui.Ini_AcctNum.setCurrentIndex(0)
        d.ui.Ini_ServerName.setText("")

    def Change_Ini_AcctNum(self,d):
        self.mtini_acctnum = d.findChild(QComboBox,"Ini_AcctNum").currentText()
        
        if self.mtini_acctnum!="* AcctNum":
            broketab = self.mtini_broker.split('~')
            accttab = self.mtini_acctnum.split('~')
            
            mt4config = ReadSQL("select server from mt4config where tab = '"+broketab[0]+
                                "' and buttonname = '"+str(accttab[0])+"'")
            if len(mt4config)>0:
                if mt4config[0][0]!=None and mt4config[0][0]!="" and mt4config[0][0]!="None":
                    d.ui.Ini_ServerName.setText(mt4config[0][0])
                else:
                    d.ui.Ini_ServerName.setText("")
            else:
                d.ui.Ini_ServerName.setText("")
        else:
            d.ui.Ini_ServerName.setText("")
        #print("broker = "+self.mtini_broker+"; buttonname = "+str(txt[0]))

    def RunPS1(self,path):
        os.system("Powershell.exe -Command \"& {Start-Process Powershell.exe -ArgumentList '-ExecutionPolicy Bypass -File \"\""+path+"\"\"' -Verb RunAs}\"")

    def Open_File_Tree(self):
        curtab = self.SM_Tabs.currentIndex()
        if curtab>0:
            tabpath = ReadSQL("select treepath from tabs where tab = '"+self.SM_Tabs.tabText(self.SM_Tabs.currentIndex())+"'")
            if len(tabpath)>0:
                tabpath = tabpath[0][0]
                if tabpath==None:
                    tabpath = ""
                if tabpath[-1:]=="\\":
                    tabpath = tabpath[:-1]
                if tabpath!="":
                    if os.path.exists(tabpath)==True:
                        win32api.WinExec("explorer.exe \""+tabpath+"\"")
                        '''self.filetree = QtWidgets.QMainWindow()
                        self.ui = Ui_SM_File_Tree()
                        self.ui.setupUi(self.filetree)
                            
                        
                        self.File_Tree_Refresh(tabpath)

                        
                        self.filetree.show()'''
        else:
            subprocess.Popen(f'explorer "{os.getcwd()}"')

    def File_Tree_Refresh(self,path):
        print("open "+path)

    def Add_EXE(self):
        curtab = self.SM_Tabs.currentIndex()
        curtab = str(self.SM_Tabs.tabText(curtab))
        new_file = QFileDialog.getOpenFileName(None,'Open file','',"EXE Files (*.exe);;Excel Files (*.xlsx *.xlsm *.xlsb *.xls);;SQLite DB Files (*.db);;All Files (*.*)")
        if len(new_file[0])<1:
            logger.error("No file selected")
            return
        new_file = new_file[0].replace("/","\\")
        maxbut = ReadSQL("select max(buttonsequence) from buttons where tab = '"+curtab+"'")
        if len(maxbut) < 1:
            logger.error("Max button is less then 1")
        print(new_file)
        idir = os.path.dirname(new_file)
        if idir[-1:]!="\\": #this is redundant as he is using os.path.dirname.
            idir = idir+"\\"
        ibase = os.path.basename(new_file)
        logger.success(ibase + " " + idir)
        new_button = str(QInputDialog.getText(None,"New EXE ButtonName","Name:",QLineEdit.Normal,"run " + ibase)[0])
        if not new_button:
            logger.error("ButtonName is None")
            return
        does_exist = ReadSQL("select buttonname from buttons where tab = '"+curtab+"' and buttonname = '"+new_button+"'")
        while len(does_exist)>0:
            new_button = str(QInputDialog.getText(None,"New EXE ButtonName","Name:",QLineEdit.Normal,"")[0])
            if new_button!=None and new_button!="":
                does_exist = ReadSQL("select buttonname from buttons where tab = '"+curtab+"' and buttonname = '"+new_button+"'")
        maxbut = maxbut[0][0] + 1 if maxbut[0][0] else 1
        maxcolumn = ReadSQL("select max(columnnum) from buttons where tab = '"+curtab+"'")
        if not maxcolumn:
            maxcolumn = "null"
        elif not maxcolumn[0][0]:
            maxcolumn = "null"
        else:
            maxcolumn = "1" #putting button automatically in column 1
        formtitle = self.title
        WriteSQL(f"""insert into buttons(formname,tab,buttonname,buttonsequence,columnnum) values('{formtitle}', '{curtab}','{new_button}',{maxbut},{maxcolumn}) """)
        WriteSQL("insert into batchsequence(formname,tab,buttonname,runsequence,folderpath,filename,type) values('"+formtitle+"','"+curtab+"','"+str(new_button)+"',1,'"+idir+"','"+ibase+"','exe')")    
        self.Refresh()
    
    

    def MT_Event(self,mt,b,r,instance,commands):
        if mt=="MT5":
            if os.path.exists(instance+"\\MQL5\\Files")==False:
                ctypes.windll.user32.MessageBoxW(0,instance+"\\MQL5\\Files does not exist?","Failed mtevent: "+b+"! \\"+str(r),0)
            else:
                if os.path.exists(instance+"\\MQL5\\Files\\pythonCom.txt")==True:
                    ctypes.windll.user32.MessageBoxW(0,instance+"\\MQL5\\Files\\pythonCom.txt already exists/processing. Try again after a few seconds.","Failed mtevent: "+b+"! \\"+str(r),0)
                else:
                    if commands==None or commands=="" or commands=="None":
                        ctypes.windll.user32.MessageBoxW(0,"No commands provided for "+instance+"\\MQL5\\Files\\pythonCom.txt?","Failed mtevent: "+b+"! \\"+str(r),0)
                    else:
                        commands = commands.split(';')
                        f = open(instance+"\\MQL5\\Files\\pythonCom.txt","w+")
                        for i in range(len(commands)):
                            print(commands[i])
                            f.write(str(commands[i])+"\n")
                        f.close()
        else:
            if os.path.exists(instance+"\\MQL4\\Files")==False:
                ctypes.windll.user32.MessageBoxW(0,instance+"\\MQL4\\Files does not exist?","Failed mtevent: "+b+"! \\"+str(r),0)
            else:
                if os.path.exists(instance+"\\MQL4\\Files\\pythonCom.txt")==True:
                    ctypes.windll.user32.MessageBoxW(0,instance+"\\MQL4\\Files\\pythonCom.txt already exists/processing. Try again after a few seconds.","Failed mtevent: "+b+"! \\"+str(r),0)
                else:
                    if commands==None or commands=="" or commands=="None":
                        ctypes.windll.user32.MessageBoxW(0,"No commands provided for "+instance+"\\MQL4\\Files\\pythonCom.txt?","Failed mtevent: "+b+"! \\"+str(r),0)
                    else:
                        commands = commands.split(';')
                        f = open(instance+"\\MQL4\\Files\\pythonCom.txt","w+")
                        for i in range(len(commands)):
                            print(commands[i])
                            f.write(str(commands[i])+"\n")
                        f.close()
    
    def SQL_To_PG(self,sqldir,pg_data,tblfilter,mode):
        sqldir = addSlash(sqldir)
        pg_data = pg_data.split('|')
        attachedconnection = apsw.Connection(sqldir)
        attachedcursor = attachedconnection.cursor()
        if tblfilter!=None and tblfilter!="":
            tblfilter = " where tbl_name = '"+str(tblfilter)+"'"
            table_names = aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master"+tblfilter+" order by tbl_name asc",12,0.2,True)
            if len(table_names)<1 and tblfilter!=None and tblfilter!="":
                istable = ReadSQL("select sql from masttbls"+tblfilter+";")
                if len(istable)>0:
                    aWriteSQL(attachedcursor,istable[0][0]+";")
                    table_names = aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master"+tblfilter+" order by tbl_name asc",12,0.2,True)
        else:
            tblfilter = ""
            table_names = aReadSQL(attachedcursor,"select distinct tbl_name from sqlite_master"+tblfilter+" order by tbl_name asc",12,0.2,True)

        for etd1 in range(len(table_names)):
            pragma_info = aReadSQL(attachedcursor,"pragma table_info('"+str(table_names[etd1][0])+"')",12,0.2,True)
            data = aReadSQL(attachedcursor,"select * from "+str(table_names[etd1][0]),12,0.2,True)
            self.conn = psycopg2.connect(dbname=pg_data[0],user=pg_data[1],password=PG_PASS,host=pg_data[2],port=pg_data[3])
            self.cursor = self.conn.cursor()
            if mode==0:
                doesTable = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+
                                        str(table_names[etd1][0])+"'")
                if len(doesTable)<1:
                    self.WritePG(self.makeCreatePG(str(table_names[etd1][0]),pragma_info))
                    alist = []
                    alist.append(str(table_names[etd1][0]))
                    doesTable.append(alist)
                '''
                if len(doesTable)<1 and tblfilter!=None and tblfilter!="":
                    istable = ReadSQL("select sql from pgtables"+tblfilter+" and hostip = '"+str(pg_data[2])+
                                      "' and port = "+str(pg_data[3])+" and username = '"+str(pg_data[1])+"' and dbname = '"+
                                      str(pg_data[0])+"' and password = '"+str(pg_data[4])+"';")
                    if len(istable)>0:
                        self.WritePG(istable[0][0])
                        doesTable = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+
                                                str(table_names[etd1][0])+"'")
                '''
                if len(doesTable)>0:
                    if str(table_names[etd1][0])=="user":
                        table_names[etd1][0] = "\""+str(table_names[etd1][0])+"\""
                    self.WritePG("delete from "+str(table_names[etd1][0]))
                    self.WritePG("begin")
                    for i in range(len(data)):
                        statement = ("insert into "+str(table_names[etd1][0])+"("+self.ConglomerateFields(pragma_info)+") values("+
                                     self.ConglomerateStatements(pragma_info,data[i])+")")
                        #print(statement)
                        self.WritePG(statement)
                    self.WritePG("commit")
            # insert or ignore into
            elif mode==1:
                doesTable = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+
                                        str(table_names[etd1][0])+"'")
                if len(doesTable)<1:
                    self.WritePG(self.makeCreatePG(str(table_names[etd1][0]),pragma_info))
                    alist = []
                    alist.append(str(table_names[etd1][0]))
                    doesTable.append(alist)
                '''
                if len(doesTable)<1 and tblfilter!=None and tblfilter!="":
                    istable = ReadSQL("select sql from pgtables"+tblfilter+" and hostip = '"+str(pg_data[2])+
                                      "' and port = "+str(pg_data[3])+" and username = '"+str(pg_data[1])+"' and dbname = '"+
                                      str(pg_data[0])+"' and password = '"+str(pg_data[4])+"';")
                    if len(istable)>0:
                        self.WritePG(istable[0][0])
                        doesTable = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+
                                                str(table_names[etd1][0])+"'")
                '''
                if len(doesTable)>0:
                    self.WritePG("begin")
                    for i in range(len(data)):
                        statement = ("insert into "+str(table_names[etd1][0])+"("+self.ConglomerateFields(pragma_info)+") values("+
                                    self.ConglomerateStatements(pragma_info,data[i])+")"+self.ConglomerateUpdate(pragma_info,data[i],mode))
                        #print(statement)
                        self.WritePG(statement)
                    self.WritePG("commit")
            # insert or replace into
            elif mode==2:
                doesTable = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+
                                        str(table_names[etd1][0])+"'")
                if len(doesTable)<1:
                    self.WritePG(self.makeCreatePG(str(table_names[etd1][0]),pragma_info))
                    alist = []
                    alist.append(str(table_names[etd1][0]))
                    doesTable.append(alist)
                '''
                if len(doesTable)<1 and tblfilter!=None and tblfilter!="":
                    istable = ReadSQL("select sql from pgtables"+tblfilter+" and hostip = '"+str(pg_data[2])+
                                      "' and port = "+str(pg_data[3])+" and username = '"+str(pg_data[1])+"' and dbname = '"+
                                      str(pg_data[0])+"' and password = '"+str(pg_data[4])+"';")
                    if len(istable)>0:
                        self.WritePG(istable[0][0])
                        doesTable = self.ReadPG("select table_name from information_schema.tables where table_schema = 'public' and table_name = '"+
                                                str(table_names[etd1][0])+"'")
                '''
                if len(doesTable)>0:
                    self.WritePG("begin")
                    for i in range(len(data)):
                        statement = ("insert into "+str(table_names[etd1][0])+"("+self.ConglomerateFields(pragma_info)+") values("+
                                    self.ConglomerateStatements(pragma_info,data[i])+")"+self.ConglomerateUpdate(pragma_info,data[i],mode))
                        #print(statement)
                        self.WritePG(statement)
                    self.WritePG("commit")
            
            self.conn.close()
        attachedconnection.close(True)

    def ReadPG(self,query,attempts=12,sec=1):
        val = []
        try:
            self.cursor.execute(query)
            results = self.cursor.fetchall()
            for result in results:
                val.append(list(result))
            return(val)
        except Exception as e4:
            print(query)
            print(str(e4))
            return(val)
            '''
            for i in range(0,attempts):
                try:
                    self.cursor.execute(query)
                    results = self.cursor.fetchall()
                    for result in results:
                        val.append(list(result))
                    return(val)
                except Exception as e4:
                    if i==attempts-1:
                        print(query)
                        print(str(e4))
                        return(val)
                    else:
                        time.sleep(sec)
                        continue
            '''

    # Write to Postgres
    def WritePG(self,query,attempts=12,sec=1,box=False):
        try:
            self.cursor.execute(query)
            self.conn.commit()
        except Exception as e4:
            print(query)
            print(str(e4))
            self.conn.rollback()
            '''
            for i in range(0,attempts):
                try:
                    self.cursor.execute(query)
                    self.conn.commit()
                except Exception as e4:
                    if i==attempts-1:
                        print(query)
                        print(str(e4))
                        return(val)
                    else:
                        time.sleep(sec)
                        continue
            '''

    def makeCreatePG(self,tblname,pragma):
        val = "CREATE TABLE IF NOT EXISTS "+str(tblname)+"("
        pk = ""
        for i in range(len(pragma)):
            if str(pragma[i][2][0:4])=="CHAR" or str(pragma[i][2][0:4])=="char":
                pragma[i][2] = "VAR"+pragma[i][2].upper()
            elif str(pragma[i][2])=="DATETIME" or str(pragma[i][2])=="datetime":
                pragma[i][2] = "TIMESTAMP WITHOUT TIME ZONE"
            elif str(pragma[i][2])=="INTEGER" or str(pragma[i][2])=="integer":
                pragma[i][2] = "BIGINT"
            elif str(pragma[i][2])=="DOUBLE" or str(pragma[i][2])=="double":
                pragma[i][2] = "NUMERIC"
            if pragma[i][3]>0:
                val += str(pragma[i][1])+" "+str(pragma[i][2])+" NOT NULL,"
            else:
                val += str(pragma[i][1])+" "+str(pragma[i][2])+","
            if pragma[i][5]>0:
                if pk=="":
                    pk = ",PRIMARY KEY("
                pk += str(pragma[i][1])+","
        val = val[:-1]
        if pk!="":
            pk = pk[:-1]+")"
            val += pk
        val += ");"
        #print(val)
        return(val)
            
    def ConglomerateFields(self,pragma):
        val = ""
        for i in range(len(pragma)):
            pragma[i][1] = str(pragma[i][1])
            val += str(pragma[i][1])+","
        if len(val)>0:
            val = val[:-1]
        return val

    def ConglomerateStatements(self,pragma,data):
        val = ""
        for i in range(len(pragma)):
            val += self.isSQLiteText(pragma[i][2],data[i])+","
        if len(val)>0:
            val = val[:-1]
        return val

    def ConglomeratePK(self,pragma,mode):
        val = ""
        if mode!=0:
            for i in range(len(pragma)):
                if pragma[i][5]>0:
                    val += str(pragma[i][1])+","
            if len(val)>0:
                val = val[:-1]
                return " on conflict ("+val+") do "
            else:
                return ""
        else:
            return ""

    def ConglomerateUpdate(self,pragma,data,mode):
        pk = self.ConglomeratePK(pragma,mode)
        if len(pk)<1:
            return ""
        else:
            val = ""
            if mode==0:
                return ""
            elif mode==1:
                return pk+"nothing"
            elif mode==2:
                if val=="" and len(pragma)>0:
                    val += "update set "
                for i in range(len(pragma)):
                    val += pragma[i][1]+" = "+self.isSQLiteText(pragma[i][2],data[i])+", "
                val = pk+val[:-2]
                return val
            
    def isSQLiteText(self,n,field):
        if field==None:
            return("null")
        elif n[0:4]=="char" or n[0:4]=="CHAR" or n[0:7]=="varchar" or n[0:7]=="VARCHAR" or n=="text" or n=="TEXT" or n=="datetime" or n=="DATETIME":
            field = str(field)
            if "'" in field:
                field = field.replace("'","''")
            return "'"+field+"'"
        else:
            field = str(field)
            if "'" in field:
                field = field.replace("'","''")
            return field

    def PG_Function(self,runfile,pg_data,inputdata,err,mode):
        pg_data = pg_data.split('|')
        txt = open(runfile).read()
        if mode==0:
            os.system("SET PGPASSWORD="+str(PG_PASS)+"& psql -h "+str(pg_data[2])+" -p "+str(pg_data[3])+" -U "+str(pg_data[1])+" -d "+str(pg_data[0])+" < \""+runfile+"\"")
        elif mode==1:
            if inputdata==None or inputdata=="":
                ctypes.windll.user32.MessageBoxW(0,"No variables specified in source?","Failed sql: "+bname+"! \\"+err,0)
            else:
                customfile = str(os.path.dirname(runfile))+"\\%%%"+str(os.path.basename(runfile))
                inputdata = inputdata.split('|')
                for i in range(len(inputdata)):
                    txt = txt.replace("%var"+str(i+1)+"%",inputdata[i])
                os.system("type NUL > "+customfile)
                if os.path.exists(customfile)==True:
                    newtxt = open(customfile,'w+')
                    newtxt.write(txt)
                    newtxt.close()
                try:
                    os.system("SET PGPASSWORD="+str(PG_PASS)+"& psql -h "+str(pg_data[2])+" -p "+str(pg_data[3])+" -U "+str(pg_data[1])+" -d "+str(pg_data[0])+" < \""+customfile+"\"")
                    os.remove(customfile)
                except OSError as e:
                    ctypes.windll.user32.MessageBoxW(0,"Error removing "+customfile+", permission issue? \\"+err,"",1)
    
    def Create_Button_Dialog(self,fname,tname,bname,dimension,dialogname,mode):
        allbuttons = ReadSQL("select buttonname from buttondialogs where "+
                             "source_formname = '"+fname+"' and source_tab = '"+
                             tname+"' and source_buttonname = '"+bname+
                             "' order by buttonsequence asc")
        if len(allbuttons)>0:
            dto = QtWidgets.QDialog()
            dto.b = Ui_ButtonDialog()
            dto.b.setupDialog(dto)
            if dialogname!=None and dialogname!="":
                print(dialogname)
                dto.setWindowTitle(str(dialogname))
            if dimension!=None and dimension!="":
                dimension = str(dimension)
                dimension = dimension.split(',')
                if len(dimension)>1:
                    try:
                        xx = int(dimension[0])
                        yy = int(dimension[1])
                        dto.resize(xx,yy)
                    except:
                        pass

            for i in reversed(range(dto.b.gridLayout.count())):
                dto.b.gridLayout.itemAt(i).widget().setParent(None)

            tabgrid = ReadSQL("select max(columnnum) from buttondialogs where "+
                              "source_formname = '"+fname+"' and source_tab = '"+
                              tname+"' and source_buttonname = '"+bname+"'")
            if len(tabgrid)<1:
                tabgrid = 1
            else:
                tabgrid = int(tabgrid[0][0])

            formbuttons = ReadSQL("select buttonname,formname,tab from buttondialogs where "+
                                  "source_formname = '"+fname+"' and source_tab = '"+
                                  tname+"' and source_buttonname = '"+bname+
                                  "' and columnnum is null order by buttonsequence asc")
            formbuttons2 = ReadSQL("select buttonname,formname,tab from buttondialogs where "+
                                   "source_formname = '"+fname+"' and source_tab = '"+
                                   tname+"' and source_buttonname = '"+bname+
                                   "' and columnnum is not null order by buttonsequence asc")
            buttonsordered = []
            for i in range(len(formbuttons)):
                buttonsordered.append(formbuttons[i])
            for i in range(len(formbuttons2)):
                buttonsordered.append(formbuttons2[i])
            buttoncol = ReadSQL("select columnnum from buttondialogs where "+
                                "source_formname = '"+fname+"' and source_tab = '"+
                                tname+"' and source_buttonname = '"+bname+
                                "' and columnnum is not null order by buttonsequence asc")
            buttoncolordered = []
            for i in range(len(formbuttons)):
                buttoncolordered.append(None)
            for i in range(len(formbuttons2)):
                buttoncolordered.append(buttoncol[i])
            buttonsdesc = ReadSQL("select buttondesc from buttondialogs where "+
                                  "source_formname = '"+fname+"' and source_tab = '"+
                                  tname+"' and source_buttonname = '"+bname+
                                  "' and columnnum is null order by buttonsequence asc")
            buttonsdesc2 = ReadSQL("select buttondesc from buttondialogs where "+
                                   "source_formname = '"+fname+"' and source_tab = '"+
                                   tname+"' and source_buttonname = '"+bname+
                                   "' and columnnum is not null order by buttonsequence asc")
            buttondescordered = []
            for i in range(len(buttonsdesc)):
                buttondescordered.append(buttonsdesc[i])
            for i in range(len(buttonsdesc2)):
                buttondescordered.append(buttonsdesc2[i])

            accumbuttons = []
            for bn in range(len(buttonsordered)):
                if bn!=0:
                    dto.b.gridLayout.setRowStretch(bn,3)
                dto.b.button = QtWidgets.QPushButton(dto.b.scrollAreaWidgetContents)
                dto.b.button.setToolTip(str(buttondescordered[bn][0]))
                dto.b.button.setObjectName(bname+"_Button_"+str(1)+"_"+str(bn+1))
                dto.b.button.setText(str(buttonsordered[bn][0]))

                if buttoncolordered[bn]==None:
                    x = 0
                    y = 0
                    if tabgrid<1:
                        tabgrid = 1
                        x = bn//tabgrid
                        y = bn%tabgrid
                        while len(accumbuttons)<1:
                            accumbuttons.append(0)
                        accumbuttons[0] += 1
                    else:
                        if tabgrid<2:
                            x = bn//tabgrid
                            y = bn%tabgrid
                            while len(accumbuttons)<1:
                                accumbuttons.append(0)
                            accumbuttons[0] += 1
                        else:
                            x = bn%math.ceil(len(buttonsordered)/tabgrid)
                            y = bn//math.ceil(len(buttonsordered)/tabgrid)
                            while len(accumbuttons)<y+1:
                                accumbuttons.append(0)
                            accumbuttons[y] += 1
                else:
                    if len(formbuttons2)>0:
                        if len(accumbuttons)<tabgrid:
                            lenaccumbuttons = len(accumbuttons)
                            for i in range(tabgrid):
                                if i>=lenaccumbuttons:
                                    accumbuttons.append(0)
                        
                        y = buttoncolordered[bn][0]-1
                        if tabgrid<2:
                            tabgrid = 1
                        if y>tabgrid-1:
                            y = tabgrid-1
                        elif y<0:
                            y = 0
                        x = accumbuttons[y]
                        
                        accumbuttons[y] += 1

                dto.b.gridLayout.addWidget(dto.b.button, x, y, 1, 1)
                dto.b.button.setStyleSheet("QPushButton { background-color: none }"
                                           "QPushButton:hover { background-color: lightblue }"
                                           "QPushButton:focus { background-color: tomato }" )
                dto.b.button.clicked.connect(partial(self.on_click_button_plus,
                                                     str(buttonsordered[bn][1]),
                                                     str(buttonsordered[bn][2]),
                                                     str(buttonsordered[bn][0]),
                                                     bname+"_Button_"+str(1)+"_"+str(bn+1),dto,mode,2))
        
            result = dto.exec_()

    def on_click_button_plus(self,pname,tname,bname,objn,d,typemode,mode=0):
        self.on_click_button(pname,tname,bname,objn,mode)
        if typemode==1:
            d.reject()


#type handlers 
    def type_copy(self, batch_name,file_name,file_source,file_destination,run_sequence,new_name): 
        full_file_source = file_source+"\\"+file_name 
        full_file_destination = file_destination+"\\"+new_name
        fail_handler = "Failed copy: "+batch_name+"! \\"+str(run_sequence)    
        if not os.path.exists(full_file_source):
            ctypes.windll.user32.MessageBoxW(0,full_file_source+" source does not exist?",fail_handler,0)
            return False
        if not os.path.exists(file_destination): 
            os.system("mkdir \""+file_destination+"\"") 
            if os.path.exists(file_destination)==False: 
                ctypes.windll.user32.MessageBoxW(0,file_destination+" target does not exist?",fail_handler,0) 
                return False
        try:
            shutil.copy2(full_file_source,full_file_destination) #copy file
            return True
        except:
            if not os.path.exists(full_file_destination): 
                ctypes.windll.user32.MessageBoxW(0,f"Problem copying {full_file_source} to {full_file_destination}?",fail_handler,0)  
                return False
            else: 
                ctypes.windll.user32.MessageBoxW(0,f"Problem copying {full_file_source} to {full_file_destination}. Check if file is open.",fail_handler,0) 
                return False

        return(False)




class Edit_Popup(QDialog):
    def __init__(self, parent, bseq, pname, tname, bname, objn):
        
        super().__init__(parent)
        self.resize(600, 300)
        self.setWindowTitle("Edit")
        self.layout = QFormLayout()
        self.setLayout(self.layout)
        self.pname_orig = pname
        self.tname_orig = tname
        self.bname_orig = bname
        self.parent = parent
        self.bseq = bseq[0]
        db_columns = ["folderpath","filename","type","source","target","databasepath","databasename","keypath","keyfile","runsequence","treepath"]
        db_buttons_columns = ["buttonsequence","columnnum","buttondesc","buttongroup","active","treepath"]
        item_dict = {}
        self.qlineeditdict = {}
        for item in range(len(self.bseq)):
            if self.bseq[item] == None:
                self.bseq[item] = ""
        for item in range(len(self.bseq)):
            item_dict[db_columns[item]] = bseq[0][item]
        #take each item up to database path from db coumns from item dict and put into button_item
        self.batchsequence_items = {"runsequence":item_dict["runsequence"],"folderpath":item_dict["folderpath"],"filename":item_dict["filename"],"type":item_dict["type"],"source":item_dict["source"],"target":item_dict["target"],"databasepath":item_dict["databasepath"], "databasename":item_dict["databasename"],"keypath":item_dict["keypath"],"keyfile":item_dict["keyfile"],"treepath":item_dict["treepath"]}
        self.buttons_items = self.grab_buttons_data()
        for item in self.buttons_items:
            if self.buttons_items[item] == None:
                self.buttons_items[item] = ""

        #add pname tname bname objn as text
        self.pname = QtWidgets.QLineEdit(self)
        self.pname.setText(str(pname))  
        self.tname = QtWidgets.QLineEdit(self)
        self.tname.setText((tname))
        self.bname = QtWidgets.QLineEdit(self)
        self.bname.setText((bname))
        self.objn = QtWidgets.QLineEdit(self)
        self.objn.setText((objn))
        #add button save
        self.save = QtWidgets.QPushButton(self)
        self.save.setText("Update")
        self.save.clicked.connect(self.on_click_save)
        self.delete = QtWidgets.QPushButton(self)
        self.delete.setText("Delete")
        self.delete.clicked.connect(self.on_click_delete)

        self.layout.addRow(self.pname, self.tname)
        self.layout.addRow("Form name", self.pname)
        self.layout.addRow("Tab name", self.tname)
        self.layout.addRow("Button name", self.bname)
        self.layout.addRow("Object name", self.objn)



        font = QtGui.QFont()
        font.setBold(True)

        btndata = QtWidgets.QLabel("Button Data")
        btndata.setFont(font)

        self.layout.addRow(btndata)
        #bold btndata
        #import Qfont


        self.buttoneditdict = {}

        for item in self.buttons_items:

            qlineedit = QtWidgets.QLineEdit(self)
            qlineedit.setText(str(self.buttons_items[item]))
            self.buttoneditdict.update({item:qlineedit})
            self.layout.addRow(item, qlineedit)


        btnbseq = QtWidgets.QLabel("Batchsequence Data")

        btnbseq.setFont(font)

        self.layout.addRow(btnbseq)

        for item in self.batchsequence_items:
            qlineedit = QtWidgets.QLineEdit(self)
            qlineedit.setText(str(self.batchsequence_items[item]))
            self.qlineeditdict.update({item:qlineedit})
            self.layout.addRow(item, qlineedit)




        self.layout.addRow(self.save)
        self.layout.addRow(self.delete)

    def on_click_save(self):
        batchsequence_changes_dict = {}
        buttons_changes_dict = {}
        queries = []
        for item in self.qlineeditdict:
            #compare to self.batchsequence_items doctopmaru
            if str(self.qlineeditdict[item].text()) != str(self.batchsequence_items[item]):
                logger.debug(self.qlineeditdict[item].text(), self.batchsequence_items[item])
                batchsequence_changes_dict.update({item:self.qlineeditdict[item].text()})

        for item in self.buttoneditdict:
            if str(self.buttoneditdict[item].text()) != str(self.buttons_items[item]):
                logger.debug(self.buttoneditdict[item].text(), self.buttons_items[item])
                buttons_changes_dict.update({item:self.buttoneditdict[item].text()})
                
        
        changed_button_data = {"formname":self.pname.text(),"tab":self.tname.text(),"buttonname":self.bname.text()}   
        button_original = {"formname":self.pname_orig,"tab":self.tname_orig,"buttonname":self.bname_orig}
        new_button_data = {"formname":self.pname.text(),"tab":self.tname.text(),"buttonname":self.bname.text()}

        for item in button_original:
            if item in changed_button_data and changed_button_data[item] == button_original[item]:
                changed_button_data.pop(item)

        
        if len(changed_button_data) > 0:
            queries.append(self.make_update_query(changed_button_data, button_original, "buttons"))
            queries.append(self.make_update_query(changed_button_data, button_original, "batchsequence"))
        
        if len(buttons_changes_dict) > 0:
            queries.append(self.make_update_query(buttons_changes_dict, new_button_data, "buttons"))

        if len(batchsequence_changes_dict) > 0:
            queries.append(self.make_update_query(batchsequence_changes_dict, new_button_data, "batchsequence"))


        if len(queries) > 0:
            for query in queries:
                WriteSQL(query, msg=True)
        #exit popup
        
        self.parent.Refresh()
        self.close()

    def grab_buttons_data(self):
        data = ReadSQL(f"SELECT buttonsequence,columnnum,buttondesc,buttongroup,active,treepath FROM buttons WHERE formname = '{self.pname_orig}' AND tab = '{self.tname_orig}' AND buttonname = '{self.bname_orig}'",msg=False)[0]
        db_buttons_columns = ["buttonsequence","columnnum","buttondesc","buttongroup","active","treepath"]
        data_dict = {}
        for item in range(len(data)):
            data_dict[db_buttons_columns[item]] = data[item]
        return data_dict



    def on_click_delete(self):
        #popup that asks to confirm if you want to delete
        #if yes, delete
        #if no, exit
        qm = QtWidgets.QMessageBox()
        qm.setText("Are you sure you want to delete?")
        qm.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        qm.setDefaultButton(QtWidgets.QMessageBox.No)
        qm.setWindowTitle("Delete?")
        ret = qm.exec_()
        if ret == QtWidgets.QMessageBox.Yes:
            #delete button
            #delete batchsequence
            queries = []
            queries.append(self.make_delete_query(self.pname_orig, self.tname_orig, self.bname_orig, "buttons"))
            queries.append(self.make_delete_query(self.pname_orig, self.tname_orig, self.bname_orig, "batchsequence"))
            for query in queries:
                WriteSQL(query, msg=True)
            self.parent.Refresh()
            self.close()

    def make_delete_query(self, pname, tname, bname, table):
        query = "DELETE FROM " + table + " WHERE formname = '" + pname + "' AND tab = '" + tname + "' AND buttonname = '" + bname + "'"
        logger.debug(query)
        return(query)

    def make_update_query(self, changes_dict, search_dict, table):
        query = "UPDATE "+table+" SET "
        for item in changes_dict:
            query += item+" = '"+changes_dict[item]+"', "
        query = query[:-2]
        query += " WHERE "
        for item in search_dict:
            query += item+" = '"+search_dict[item]+"' AND "
        query = query[:-5]
        logger.debug(query)
        return query

class Create_sequence(QDialog):
    def __init__(self, parent):
        self._parent = parent
        self.orderedbuttons = []
        super().__init__(parent)
        self.resize(600, 300)
        self.setWindowTitle("Sequence Builder")
        self.layout = QFormLayout()
        self.setLayout(self.layout)
        self.button_items = []
        font = QtGui.QFont()
        font.setBold(True)

        self.button_name = QtWidgets.QLineEdit(self)
        self.button_name.setPlaceholderText("Enter Button Name")

        self.button_tabs = QtWidgets.QComboBox(self)

        tablist = self.get_tabs()
        self.button_tabs.addItems(tablist)
        #        button_tabs.setFixedWidth(20)
        #        button_tabs.adjustSize()

        self.info_grid = QGridLayout()
        self.info_grid.setSpacing(10)
        self.info_grid.setContentsMargins(10, 10, 10, 10)
        self.info_grid.addWidget(self.button_name, 0, 0)
        self.info_grid.addWidget(self.button_tabs, 0, 1)
        self.layout.addRow(self.info_grid)

        self.btn1x = QtCore.pyqtSignal(int)
        self.btn1y = QtCore.pyqtSignal(int)


        select_btn_text = QtWidgets.QLabel("Select Buttons you want to use on main tab")
        select_btn_text.setFont(font)
        self.layout.addRow(select_btn_text)


    def get_tabs(self):
        #using self._parent.ReadSQL
        #SELECT formname, tab, tabsequence FROM tabs where formname is 

        tabs = ReadSQL(f"SELECT tab FROM tabs where formname = '{self._parent.title}' ORDER BY tabsequence") 
        tablist = []
        for tab in tabs:
            tablist.append(tab[0])
        return(tablist)

    def add_button(self,bseq,pname,tname,bname,objn,mode):
        new_button = DragButton(bname, self)
        new_button.fake_init(self, bname)
        new_button_x = QPushButton("X")
        #run function self.clicked and send the button obj

        new_button_x.setFixedWidth(20)
        new_button_x.adjustSize()

        new_button_x.setStyleSheet("color: red; font-weight: bold;")
        
        grid = QGridLayout()
        grid.addWidget(new_button, 0, 0)
        grid.addWidget(new_button_x, 0, 1)
        self.layout.addRow(grid)
        x = len(self.orderedbuttons)
        arr = [x, new_button, new_button_x]
        dict_data = {"arrpos":x,"bseq":bseq,"pname":pname,"tname":tname,"bname":bname,"objn":objn,"mode":mode, "btn_item":new_button, "x":new_button.x(), "y":new_button.y(), "x_btn":new_button_x, "grid":grid}
        self.orderedbuttons.append(arr)
        self.button_items.append(dict_data)
        new_button_x.clicked.connect(lambda: self.remove_button(dict_data))
    
    def remove_button(self, data_dict):
        self.layout.removeRow(data_dict["grid"])
        self.button_items.remove(data_dict)
        for item in self.orderedbuttons:
            if item[1] == data_dict["btn_item"]:
                self.orderedbuttons.remove(item)
                break



    def closeEvent(self, event):
        qm = QtWidgets.QMessageBox()
        qm.setText("Are you sure you want to exit the sequence builder?")
        qm.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        qm.setDefaultButton(QtWidgets.QMessageBox.No)
        qm.setWindowTitle("Exit?")
        ret = qm.exec_()
        if ret == QtWidgets.QMessageBox.Yes:
            self._parent.cur_seq = None
            event.accept()
        else:
            event.ignore()
    
    def set_others_lower(self, button):
        for item in self.button_items:
            if button != item["bname"]:
                item["btn_item"].lower()

    def reset_layout(self, buttons):
        #remove from layout
        for item in self.button_items:

            item["grid"].removeWidget(item["btn_item"])
            item["grid"].removeWidget(item["x_btn"])
            self.layout.removeRow(item["grid"])
            logger.debug(self.orderedbuttons)

        self.orderedbuttons = buttons
        for item in buttons:
            grid = QGridLayout()
            grid.addWidget(item[1], 0, 0)
            grid.addWidget(item[2], 0, 1)
            self.layout.addRow(grid)
            #replace grid in button_items
            for button in self.button_items:
                if button["btn_item"] == item[1]:
                    button["grid"] = grid
                    break
    

        #add to layout

    
class DragButton(QPushButton):

    def fake_init(self, parent_, bname):
        self.parent_ = parent_
        self.bname = bname

    def mousePressEvent(self, event):
        self.__mousePressPos = None
        self.__mouseMovePos = None
        if event.button() == QtCore.Qt.LeftButton:
            self.__mousePressPos = event.globalPos()
            self.__mouseMovePos = event.globalPos()
    
        super(DragButton, self).mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if event.buttons() == QtCore.Qt.LeftButton:
            # adjust offset from clicked point to origin of widget
            self.show()
            currPos = self.mapToGlobal(self.pos())
            globalPos = event.globalPos()
            self.parent_.set_others_lower(self.bname)
            diff = globalPos - self.__mouseMovePos
            diff.setX(0)

            newPos = self.mapFromGlobal(currPos + diff)
            logger.debug(f"{newPos.x()}, {newPos.y()}, {self.parent_.height()}")
            if newPos.y() < 80:
                newPos.setY(80)
            if newPos.y() > self.parent_.height() - 35:
                newPos.setY(self.parent_.height() - 35)
            self.__mouseMovePos = globalPos
            self.move(newPos)
            for item in self.parent_.button_items:
                if item["bname"] == self.bname:
                    item["x"] = newPos.x()
                    item["y"] = newPos.y()
                    item["x_btn"].move(newPos.x()+self.width(), newPos.y())

        super(DragButton, self).mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self.__mousePressPos is not None:
            buttons = []
            for item in self.parent_.button_items:

                
                #add item to buttons array ordered by y position
                buttons.append([item['btn_item'].y(), item['btn_item'], item['x_btn']])

            buttons.sort(key=lambda x: x[0])
            self.parent_.reset_layout(buttons)
            moved = event.globalPos() - self.__mousePressPos 
            if moved.manhattanLength() > 3:
                event.ignore()
        super(DragButton, self).mouseReleaseEvent(event)

def main():
    app = QApplication(sys.argv)
    form = DB_Window()
    form.show()
    
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()

# Disconnect from DB
connection.close(True)


