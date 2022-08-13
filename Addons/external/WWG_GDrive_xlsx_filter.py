from __important.PluginInterface import PluginInterface


from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from pandas import read_excel, to_datetime
import openpyxl
#import get_column_letter
from openpyxl.utils import get_column_letter
#import Table
from openpyxl.worksheet.table import Table




class WWG_GDrive_xlsx_filter(PluginInterface):
    load = True
    types = {"keyfile":8, "target":4}
    type_types = {"keyfile":{"type":"drag_drop_file", "description":"please select the excel document"}, "target":{"type":"drag_drop_folder", "description":"please select the save location folder"}}

    callname = "wwgdrive_filter"
    hooks_handler = ["log"]



    def load_self(self, hooks):
        self.logger = hooks["log"]
        return True


    # "keyfile":8,"runsequence":9,"treepath":10,"buttonname":11}





    def main(self,xlsx_loc, save_loc, Popups):
        self.logger.success(f"xlsx_loc: {xlsx_loc}")
        #load the excel file and get the dataframe
        try:
            df = read_excel(xlsx_loc)
        except Exception as e:
            Popups.alert("Error", str(e))
            return False
        

            

        dialog = Dialog
        x = Popups.custom(dialog)
        if x == (None,):
            return
        keywords, folder, filetype, start_date, end_date, is_date_filtered, file_name = x
        if not is_date_filtered:
            start_date = None
            end_date = None


        if keywords is None:
            Popups.alert("Please enter keywords")
            return
        keywords = keywords.split(",")


        #df headers: File	File Type	Folder Location	Link	Path	URL	Modified	Modified By	Created
        #modified/created format = 2022-01-03T10:06:24.558Z
        #filter on df
        #print df headers
        df.columns = df.iloc[0]
        df = df.drop(df.index[0])

        #delete first column
        df = df.drop(df.columns[0], axis=1)
    
        df = df[df["File"].str.contains("|".join(keywords), case=False)]
        df = df[df["Main Folder"].str.contains(folder, case=False)] if folder not in ["Any", ""] else df
        df = df[df["File Type"].str.contains(filetype, case=False)] if filetype not in ["Any", ""] else df

        #convert modified date to datetime
        if start_date is not None:
            df["Modified"] = to_datetime(df["Modified"], utc=False)
            df["Created"] = to_datetime(df["Created"], utc=False)
            start_date = to_datetime(start_date, utc=True)
            end_date = to_datetime(end_date, utc=True)
            df = df[(df["Modified"] >= start_date) & (df["Modified"] <= end_date)] if is_date_filtered else df
            df = df[(df["Created"] >= start_date) & (df["Created"] <= end_date)] if is_date_filtered else df
        
            #set timezones to unaware
            df["Modified"] = df["Modified"].dt.tz_localize(None)
            df["Created"] = df["Created"].dt.tz_localize(None)
        #replace LINK with =HYPERLINK("URL", "URL")
        #df["URL"].apply(lambda x: f"=HYPERLINK(\"{x}\", \"{x}\")")
        df["Link"] = df["URL"].apply(lambda x: f"=HYPERLINK(\"{x}\", \"link\")")

        #change order to this:Link,File,File Type,Folder Location,Main Folder,Modified By,Modified,Created,URL and remove path

        df = df.drop(df.columns[4], axis=1)
        #drop URL
        df = df[["Link", "File", "File Type", "Folder Location", "Main Folder", "Modified By", "Modified", "Created", "URL"]]
        df = df.drop(df.columns[8], axis=1)



        #add column filter to df
        df["Filter"] = ",".join(keywords)
        #save to excel file with auto

        
        while True:
            try:
                file_name =  file_name + ".xlsx" if file_name not in (None, False, "") else "WWG_GDrive_Filtered.xlsx"

                df.to_excel(save_loc + "\\" + file_name, index=False)
            except Exception as e:
                Popups.alert(str(e), "Error")
                x = Popups.yesno( f"ERROR:{str(e)}, Do you want to try again?")
                if x is False:
                    return False
            self.logger.success(f"Saved to: {save_loc} as {file_name}")
            break


###NEW ACTION FOR THIS######
        #load file in with openpyxl
        wb = openpyxl.load_workbook(save_loc + "\\" + file_name)
        ws = wb.active
        #column a width = 3.8, b = 63, c = 8, d = 45, e = 12, f = 14, g = 15, h = 9.7, i = 13
        ws.column_dimensions['A'].width = 3.86
        ws.column_dimensions['B'].width = 68
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 49
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 15.7
        ws.column_dimensions['H'].width = 9.7
        ws.column_dimensions['I'].width = 13
        #convert to table
        #a -> i
        #rename sheet to WWG_GD
        ws.title = "WWG_GD"

        #table_setter:
        table = Table(displayName="tbl_fltr", ref="A1:"  + get_column_letter(ws.max_column) + str(ws.max_row))
        ws.add_table(table)




        wb.save(save_loc + "\\" + file_name)
        

class Dialog(QDialog):
    NumGridRows = 3
    NumButtons = 4
    signal = pyqtSignal(tuple)
    _done = False

    def __init__(self, parent=None):
        super(Dialog, self).__init__(parent)
        self.is_date_filtered = False

        keywords_label = QLabel("Keywords:")
        self.keywords = QLineEdit()
        self.keywords.setPlaceholderText("Enter keywords separated by commas")

        folder_label = QLabel("Folder:")
        self.folder_combobox = QComboBox()
        self.folder_combobox.addItem("Any")

        filetype_label = QLabel("File Type:")
        self.filetype_combobox = QComboBox()
        self.filetype_combobox.addItem("Any")

        start_date_label = QLabel("Start Date:")
        self.start_date = QDateEdit(QDate.currentDate())
        self.start_date.setDisplayFormat("dd/MM/yyyy")
        self.start_date.setCalendarPopup(True)
        self.start_date.setEnabled(False)

        end_date_label = QLabel("End Date:")
        self.end_date = QDateEdit(QDate.currentDate())
        self.end_date.setDisplayFormat("dd/MM/yyyy")
        self.end_date.setCalendarPopup(True)
        self.end_date.setEnabled(False)


        filter_date_label = QLabel("Filter on Date?")
        self.filter_date = QCheckBox()
        self.filter_date.setChecked(False)
        self.filter_date.stateChanged.connect(self.filter_date_changed)


        self.filter_button = QPushButton("Filter")
        self.filter_button.clicked.connect(self.accept)

        self.file_name_label = QLabel("File Name:")
        self.file_name = QLineEdit()


        
        object_layout = QGridLayout()
        object_layout.addWidget(keywords_label, 0, 0)
        object_layout.addWidget(self.keywords, 0, 1)
        object_layout.addWidget(folder_label, 1, 0)
        object_layout.addWidget(self.folder_combobox, 1, 1)
        object_layout.addWidget(filetype_label, 2, 0)
        object_layout.addWidget(self.filetype_combobox, 2, 1)
        object_layout.addWidget(filter_date_label, 3, 0)
        object_layout.addWidget(self.filter_date, 3, 1)
        object_layout.addWidget(start_date_label, 4, 0)
        object_layout.addWidget(self.start_date, 4, 1)
        object_layout.addWidget(end_date_label, 5, 0)
        object_layout.addWidget(self.end_date, 5, 1)
        object_layout.addWidget(self.filter_button, 6, 1)
        object_layout.addWidget(self.file_name_label, 7, 0)
        object_layout.addWidget(self.file_name, 7, 1)

        main_layout = QVBoxLayout()
        main_layout.addLayout(object_layout)
        main_layout.addWidget(self.filter_button)
        self.add_dropdown_data()
        
        self.setLayout(main_layout)
        self.setWindowTitle("WWG GDrive XLSX Filter")
        
        self.setWindowTitle("Filter XLSX")
        
    def filter_date_changed(self, state):
        if state == 2:
            self.start_date.setEnabled(True)
            self.end_date.setEnabled(True)
            self.is_date_filtered = True
        else:
            self.start_date.setEnabled(False)
            self.end_date.setEnabled(False)
            self.is_date_filtered = False

    def add_dropdown_data(self):
        folders = ["1. WWG",
                    "2. Client & Partner Accounts",
                    "3. Projects",
                    "4. Business Development",
                    "5. Marketing",
                    "6. Success",
                    "7. Sustainability",
                    "12. Executive Committee",
                    "15. G17Eco",
                    "19. WWG Academy",
                    "CT Light",
                    "Project Tracker App",
                    "Project Ubuntu",
                    "Sustainability Reseach",]
        folders = [x.lower() for x in folders]
        self.folder_combobox.addItems(folders)
        filetypes = ['csv','db','doc','docx','GIF','html''ico','jpg','JPEG','json','lnk','msg','pdf','pkl','png','ppt','pptx','pst','py','pyc','rtf','svg','txt','url','vsd','xls','xlsb','xlsm','xlsx','yml','zip']
        self.filetype_combobox.addItems(filetypes)

            

    def closeEvent(self, a0) -> None:
        if self._done:
            self.signal.emit((self.keywords.text().lower(), self.folder_combobox.currentText(), self.filetype_combobox.currentText(), self.start_date.date().toPyDate(), self.end_date.date().toPyDate(), self.is_date_filtered, self.file_name.text()))
        else:
            self.signal.emit((None,))
            self.close()
        self.close()
        a0.accept()


    def accept(self):
        self._done = True
        self.close()



