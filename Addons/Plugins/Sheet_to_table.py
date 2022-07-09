from .__important.PluginInterface import PluginInterface
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
import ctypes
class Sheet_to_table(PluginInterface):
    load = True
    types = {"source":3,"target":4}
    type_types = {"source":["drag_drop_folder", "please select source folder"],"target":["drag_drop_folder", "please select destination folder"], "__Name":"Sheet to table"}
    callname = "sht2tbl"
    hooks_handler = ["log"]

    def load_self(self, hooks):
        self.logger = hooks["log"]
        return True


    # "keyfile":8,"runsequence":9,"treepath":10,"buttonname":11}

    def main(self, source, target, Popups) -> bool:
        for file in os.listdir(source):
            if file.endswith(".xlsx") or file.endswith(".xls"):
                save_name = os.path.join(target, file)
                res = self.Convert_table(os.path.join(source,file), save_name)
                if not res:
                    return False

        self.logger.success(f"Converted all files in {source}")
        return True

    def Convert_table(self, file_name, targetfilename):
        self.logger.debug(f"Converting {file_name}")
        wb = load_workbook(file_name)
        #load first worksheet
        sheet = wb.active
        #create a table

        table = Table(displayName="table", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
        try:
            sheet.add_table(table)
        except Exception as e:
            self.logger.warning(f"Error: {e} in {file_name}")
            return False

        #save
        try:
            wb.save(targetfilename)
            self.logger.success(f"Converted {file_name}")
        except PermissionError as e:
            self.logger.warning(f"Permission error, please close {targetfilename}")
            self.logger.error(e)
            x = ctypes.windll.user32.MessageBoxW(0, f"Permission error, please close {file_name}", "Error", 0)
            #if the popup is exited
            try:
                wb.save(targetfilename)
                self.logger.success(f"Converted {file_name}")
            except PermissionError as e:
                self.logger.error("permission error, second attempt failed. Exiting")
                self.logger.error(e)
                return False
        return True
